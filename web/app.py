from contextlib import asynccontextmanager
from fastapi import FastAPI, Request, Form, HTTPException, UploadFile, File
from fastapi.responses import RedirectResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import csv
import io
import os
import json
import socket
import threading as _threading
import importlib
import traceback as _traceback
from datetime import date, datetime, timezone

from db import get_connection, create_device, update_device, delete_device, get_devices, get_device
from db import (list_global_users, create_global_user, update_global_user, delete_global_user,
                find_global_user_by_global_id, get_global_user)
from db import upsert_employee
from db import (get_employee_with_device, get_employees_with_device,
                delete_employee_record, bulk_delete_employee_records,
                get_all_departments)
import db as db_mod
import puller as puller_mod
import re
import time as _time
import concurrent.futures
import psycopg2.extras
import config as _cfg_mod
from config import SCHEDULE_TIMES, SCHEDULER_TIMEZONE, load_db_config
from config import COMPANY_NAME, COMPANY_ADDRESS, COMPANY_EMAIL, COMPANY_WEBSITE
from urllib.parse import urlencode
from web.flash import redirect_with_flash
from web.helpers import render, device_config_from_row, attendance_to_dict, action_label, _get_company_settings
from web.auth import (get_secret_key, find_user_by_username, verify_password,
                      get_session_user, hash_password)
from db import (get_web_user_by_username, get_web_user_by_id, get_all_web_users,
                create_web_user, update_web_user, update_web_user_login,
                add_web_audit_log, get_web_audit_logs)
import nepali_utils

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "templates")

# ---- In-process scheduler -----------------------------------------------
_web_scheduler: BackgroundScheduler | None = None
_pull_lock = _threading.Lock()


def _scheduler_jobs_info() -> list:
    if _web_scheduler is None:
        return []
    out = []
    for j in _web_scheduler.get_jobs():
        nrt = j.next_run_time
        out.append({
            "name": j.name,
            "next_run": nrt.strftime("%Y-%m-%d %H:%M %Z") if nrt else "—",
        })
    return out


def _safe_pull() -> bool:
    """Run one pull cycle unless one is already in progress. Returns True if it ran."""
    if not _pull_lock.acquire(blocking=False):
        return False
    try:
        from main import run_pull_cycle
        run_pull_cycle()
        return True
    finally:
        _pull_lock.release()


def _restart_web_scheduler():
    global _web_scheduler
    if _web_scheduler and _web_scheduler.running:
        _web_scheduler.shutdown(wait=False)
    tz = _cfg_mod.SCHEDULER_TIMEZONE

    conn = get_connection()
    try:
        active_times = db_mod.get_pull_schedule(conn, active_only=True)
    finally:
        conn.close()
    times = [(row['hour'], row['minute']) for row in active_times]

    sched = BackgroundScheduler(timezone=tz)
    for hour, minute in times:
        sched.add_job(
            _safe_pull,
            trigger=CronTrigger(hour=hour, minute=minute, timezone=tz),
            id=f"zkteco_pull_{hour:02d}{minute:02d}",
            name=f"ZKTeco Pull {hour:02d}:{minute:02d}",
            max_instances=1,
            misfire_grace_time=300,
            coalesce=True,
        )
    sched.start()
    _web_scheduler = sched


@asynccontextmanager
async def _app_lifespan(fastapi_app: FastAPI):
    # Ensure schema is up to date (idempotent) and backfill missing BS dates
    try:
        _conn = get_connection()
        db_mod.init_schema(_conn)
        _conn.commit()
        filled = db_mod.backfill_bs_dates(_conn)
        if filled:
            import logging as _lg
            _lg.getLogger(__name__).info("Backfilled BS dates for %d attendance rows.", filled)
        _conn.close()
    except Exception as _e:
        import logging as _lg
        _lg.getLogger(__name__).warning("Startup DB init/backfill failed: %s", _e)
    _restart_web_scheduler()
    yield
    if _web_scheduler and _web_scheduler.running:
        _web_scheduler.shutdown(wait=False)


app = FastAPI(title="ZKTeco Puller — Web UI", lifespan=_app_lifespan)
app.mount("/static", StaticFiles(directory=os.path.join(os.path.dirname(__file__), "static")), name="static")
templates = Jinja2Templates(directory=TEMPLATES_DIR)
nepali_utils.register_filters(templates)

# ─── Shared helpers ───────────────────────────────────────────────────────────


def _today_bs() -> str:
    """Return today's BS date string in NPT."""
    return db_mod._today_bs()


# ─── Auth helpers ──────────────────────────────────────────────────────────────


def _current_user(request: Request) -> dict | None:
    return get_session_user(request)


def _current_user_id(request: Request) -> int:
    u = _current_user(request)
    return u['id'] if u else 0


_ADMIN_ONLY_PREFIXES = (
    '/devices', '/sync', '/migrate', '/schedule',
    '/pull-sessions', '/settings', '/web-users', '/payroll',
)
_VIEWER_ALLOWED_PREFIXES = (
    '/', '/attendance', '/reports/daily', '/reports/absent',
    '/profile', '/logout',
)
_EMPLOYEE_ALLOWED_PREFIXES = (
    '/my-attendance', '/my-leaves', '/profile', '/logout',
)


async def _auth_gate_dispatch(request: Request, call_next):
    path = request.url.path
    if path == "/login" or path.startswith("/static"):
        return await call_next(request)

    if not request.session.get("user_id"):
        if path.startswith("/api/"):
            return JSONResponse({"error": "Not authenticated"}, status_code=401)
        return RedirectResponse(url=f"/login?next={path}", status_code=302)

    role = request.session.get("role", "admin")

    # admin-only paths
    if any(path == p or path.startswith(p + '/') for p in _ADMIN_ONLY_PREFIXES):
        if role != "admin":
            if path.startswith("/api/"):
                return JSONResponse({"error": "Access denied"}, status_code=403)
            return RedirectResponse(url="/?no_access=1", status_code=302)

    # viewer: restrict to allowed paths
    if role == "viewer":
        allowed = any(
            path == p or path.startswith(p + '/') or path.startswith(p + '?')
            for p in _VIEWER_ALLOWED_PREFIXES
        ) or path == "/"
        if not allowed:
            if path.startswith("/api/"):
                return JSONResponse({"error": "Access denied"}, status_code=403)
            return RedirectResponse(url="/attendance", status_code=302)

    # employee: can only see their own daily attendance report + profile
    if role == "employee":
        allowed = any(
            path == p or path.startswith(p + '/') or path.startswith(p + '?')
            for p in _EMPLOYEE_ALLOWED_PREFIXES
        )
        if not allowed:
            if path.startswith("/api/"):
                return JSONResponse({"error": "Access denied"}, status_code=403)
            return RedirectResponse(url="/my-attendance", status_code=302)

    return await call_next(request)


# Middleware registration order matters for Starlette's LIFO stack:
# add_middleware(A) then add_middleware(B) → stack is B(A(router))
# We want: Session(AuthGate(router)) — Session runs first to set up request.session,
# then AuthGate can safely read it.
# So: register AuthGate FIRST, Session SECOND.
from starlette.middleware.base import BaseHTTPMiddleware
app.add_middleware(BaseHTTPMiddleware, dispatch=_auth_gate_dispatch)
app.add_middleware(SessionMiddleware, secret_key=get_secret_key(), session_cookie="zk_session", max_age=86400 * 7)


# ─── Login / Logout ────────────────────────────────────────────────────────────

def _get_client_ip(request: Request) -> str:
    xff = request.headers.get("X-Forwarded-For", "")
    if xff:
        return xff.split(",")[0].strip()
    return request.client.host if request.client else ""


@app.get("/login")
def login_get(request: Request, next: str | None = None):
    if request.session.get("user_id"):
        return RedirectResponse(url=next or "/", status_code=302)
    return templates.TemplateResponse(request, "login.html", {"error": None, "username": ""})


@app.post("/login")
def login_post(request: Request,
               username: str = Form(...),
               password: str = Form(...),
               next: str = Form("/login")):
    ip = _get_client_ip(request)
    ua = request.headers.get("User-Agent", "")[:255]
    conn = None
    try:
        conn = get_connection()
        # Try DB first
        db_user = get_web_user_by_username(conn, username)
        if db_user and verify_password(password, db_user["password_hash"]):
            request.session["user_id"]        = db_user["id"]
            request.session["username"]       = db_user["username"]
            request.session["display_name"]   = db_user.get("display_name") or db_user["username"]
            request.session["role"]           = db_user.get("role", "viewer")
            request.session["user_source"]    = "db"
            request.session["global_user_id"] = db_user.get("global_user_id")
            update_web_user_login(conn, db_user["id"], ip)
            add_web_audit_log(conn, db_user["id"], db_user["username"],
                              "login", ip, ua)
            default_dest = "/my-attendance" if db_user.get("role") == "employee" else "/"
            dest = next if next and next not in ("/login", "") else default_dest
            if db_user.get("must_change_pwd"):
                return redirect_with_flash("/profile", "warning",
                    "For security, please change your password — it's currently set to your attendance ID.")
            return RedirectResponse(url=dest, status_code=302)
        if db_user:
            # Username matched but wrong password
            add_web_audit_log(conn, db_user["id"], username,
                              "login_failed", ip, ua, {"reason": "bad_password"})
        else:
            # Try users.json fallback
            json_user = find_user_by_username(username)
            if json_user and verify_password(password, json_user["password_hash"]):
                request.session["user_id"]        = json_user["id"]
                request.session["username"]       = json_user["username"]
                request.session["display_name"]   = json_user.get("display_name", json_user["username"])
                request.session["role"]           = json_user.get("role", "admin")
                request.session["user_source"]    = "json"
                request.session["global_user_id"] = None
                dest = next if next and next not in ("/login", "") else "/"
                return RedirectResponse(url=dest, status_code=302)
            # No match anywhere — log failed attempt if we can find any partial match
            add_web_audit_log(conn, None, username,
                              "login_failed", ip, ua, {"reason": "user_not_found"})
    except Exception:
        pass
    finally:
        if conn:
            conn.close()
    return templates.TemplateResponse(request, "login.html", {
        "error": "Invalid username or password.",
        "username": username,
    })


@app.post("/logout")
def logout(request: Request):
    uid    = request.session.get("user_id")
    uname  = request.session.get("username", "")
    source = request.session.get("user_source", "json")
    ip     = _get_client_ip(request)
    if uid and source == "db":
        try:
            conn = get_connection()
            add_web_audit_log(conn, uid, uname, "logout", ip)
            conn.close()
        except Exception:
            pass
    request.session.clear()
    return RedirectResponse(url="/login", status_code=302)


@app.get("/logout")
def logout_get(request: Request):
    uid    = request.session.get("user_id")
    uname  = request.session.get("username", "")
    source = request.session.get("user_source", "json")
    ip     = _get_client_ip(request)
    if uid and source == "db":
        try:
            conn = get_connection()
            add_web_audit_log(conn, uid, uname, "logout", ip)
            conn.close()
        except Exception:
            pass
    request.session.clear()
    return RedirectResponse(url="/login", status_code=302)


# ─── Profile ──────────────────────────────────────────────────────────────────


@app.get("/profile")
def profile_get(request: Request):
    user = _current_user(request)
    linked = None
    last_login_at = None
    last_login_ip = None
    if user and user.get("source") == "db":
        conn = get_connection()
        try:
            db_u = get_web_user_by_id(conn, user["id"])
            if db_u:
                last_login_at = db_u.get("last_login_at")
                last_login_ip = db_u.get("last_login_ip")
                if db_u.get("global_user_id"):
                    from db import get_global_user
                    linked = get_global_user(conn, db_u["global_user_id"])
        finally:
            conn.close()
    return render(templates, request, "profile.html", {
        "current_user":  user,
        "linked":        linked,
        "last_login_at": last_login_at,
        "last_login_ip": last_login_ip,
        "COMPANY_NAME":  COMPANY_NAME,
    })


@app.post("/profile/change-password")
def profile_change_password(request: Request,
                             current_password: str = Form(...),
                             new_password:     str = Form(...),
                             confirm_password: str = Form(...)):
    user = _current_user(request)
    if not user or user.get("source") != "db":
        return redirect_with_flash("/profile", "error",
                                   "Password change is only available for DB accounts.")
    if new_password != confirm_password:
        return redirect_with_flash("/profile", "error", "New passwords do not match.")
    if len(new_password) < 6:
        return redirect_with_flash("/profile", "error", "Password must be at least 6 characters.")
    conn = get_connection()
    try:
        db_u = get_web_user_by_id(conn, user["id"])
        if not db_u or not verify_password(current_password, db_u["password_hash"]):
            return redirect_with_flash("/profile", "error", "Current password is incorrect.")
        update_web_user(conn, user["id"], user["id"],
                        password_hash=hash_password(new_password),
                        must_change_pwd=False)
        ip = _get_client_ip(request)
        add_web_audit_log(conn, user["id"], user["username"],
                          "password_changed", ip,
                          request.headers.get("User-Agent", "")[:255])
    finally:
        conn.close()
    return redirect_with_flash("/profile", "success", "Password updated successfully.")


# ─── Web User Management (admin only) ─────────────────────────────────────────


@app.get("/web-users")
def web_users_list(request: Request):
    conn = get_connection()
    try:
        users = get_all_web_users(conn)
    finally:
        conn.close()
    return render(templates, request, "web_users.html", {
        "users":        users,
        "COMPANY_NAME": COMPANY_NAME,
    })


@app.post("/web-users/backfill-employees")
def web_users_backfill_employees(request: Request):
    """One-time/occasional bulk action: create an 'employee' self-service
    login for every Global User that doesn't have a web login yet."""
    from db import get_global_user_ids_without_web_login, create_employee_login
    conn = get_connection()
    created, skipped = 0, 0
    try:
        gu_ids = get_global_user_ids_without_web_login(conn)
        for gu_id in gu_ids:
            result = create_employee_login(conn, gu_id, created_by=_current_user_id(request))
            conn.commit()
            if result.get('created'):
                created += 1
            else:
                skipped += 1
    finally:
        conn.close()
    return redirect_with_flash("/web-users", "success",
        f"Created {created} employee login(s); skipped {skipped} (already had a login or no attendance ID).")


@app.get("/web-users/new")
def web_user_new_get(request: Request):
    from db import get_all_global_users_with_dept
    conn = get_connection()
    try:
        global_users = get_all_global_users_with_dept(conn)
    finally:
        conn.close()
    return render(templates, request, "web_user_form.html", {
        "form_user":    None,
        "global_users": global_users,
        "COMPANY_NAME": COMPANY_NAME,
    })


@app.post("/web-users/new")
def web_user_new_post(request: Request,
                      username:       str = Form(...),
                      display_name:   str = Form(""),
                      password:       str = Form(...),
                      confirm_pwd:    str = Form(...),
                      role:           str = Form("viewer"),
                      global_user_id: str = Form(""),
                      must_change_pwd: str = Form("")):
    if password != confirm_pwd:
        return redirect_with_flash("/web-users/new", "error", "Passwords do not match.")
    if len(password) < 6:
        return redirect_with_flash("/web-users/new", "error", "Password must be at least 6 characters.")
    conn = get_connection()
    try:
        if get_web_user_by_username(conn, username):
            return redirect_with_flash("/web-users/new", "error",
                                       f"Username '{username}' already exists.")
        gid = int(global_user_id) if global_user_id.isdigit() else None
        uid = create_web_user(conn,
                              username     = username.strip(),
                              password_hash= hash_password(password),
                              display_name = display_name.strip() or username.strip(),
                              role         = role,
                              global_user_id = gid,
                              created_by   = _current_user_id(request))
        if must_change_pwd:
            update_web_user(conn, uid, _current_user_id(request), must_change_pwd=True)
        ip = _get_client_ip(request)
        add_web_audit_log(conn, uid, username, "created", ip,
                          request.headers.get("User-Agent", "")[:255],
                          {"created_by": request.session.get("username")})
    finally:
        conn.close()
    return redirect_with_flash("/web-users", "success",
                                f"User '{username}' created successfully.")


@app.get("/web-users/{user_id}/edit")
def web_user_edit_get(request: Request, user_id: int):
    from db import (get_all_global_users_with_dept, get_all_departments,
                    get_all_sections, get_all_units, get_all_shifts)
    conn = get_connection()
    try:
        form_user    = get_web_user_by_id(conn, user_id)
        global_users = get_all_global_users_with_dept(conn)
        departments  = get_all_departments(conn)
        sections     = get_all_sections(conn)
        units        = get_all_units(conn)
        shifts       = get_all_shifts(conn)
    finally:
        conn.close()
    if not form_user:
        return redirect_with_flash("/web-users", "error", "User not found.")
    return render(templates, request, "web_user_form.html", {
        "form_user":    form_user,
        "global_users": global_users,
        "departments":  departments,
        "sections":     sections,
        "units":        units,
        "shifts":       shifts,
        "COMPANY_NAME": COMPANY_NAME,
    })


@app.post("/web-users/{user_id}/edit")
def web_user_edit_post(request: Request,
                       user_id:        int,
                       display_name:   str = Form(""),
                       role:           str = Form("viewer"),
                       global_user_id: str = Form(""),
                       is_active:      str = Form(""),
                       new_password:   str = Form(""),
                       confirm_pwd:    str = Form(""),
                       # global user fields
                       gu_employee_id:      str = Form(""),
                       gu_name:             str = Form(""),
                       gu_designation:      str = Form(""),
                       gu_department_id:    str = Form(""),
                       gu_section_id:       str = Form(""),
                       gu_unit_id:          str = Form(""),
                       gu_shift_id:         str = Form(""),
                       gu_emp_type:         str = Form("PERMANENT"),
                       gu_emp_status:       str = Form("ACTIVE"),
                       gu_join_date:        str = Form(""),
                       gu_level_grade:      str = Form(""),
                       gu_email:            str = Form(""),
                       gu_phone:            str = Form(""),
                       gu_bank_number:      str = Form(""),
                       gu_appointment_date: str = Form(""),
                       gu_dob:              str = Form(""),
                       gu_gender:           str = Form(""),
                       must_change_pwd:     str = Form("")):
    conn = get_connection()
    try:
        # ── Update web_users table ──────────────────────────────────────
        fields: dict = {
            "display_name":   display_name.strip(),
            "role":           role,
            "global_user_id": int(global_user_id) if global_user_id.isdigit() else None,
            "is_active":      bool(is_active),
            "must_change_pwd": bool(must_change_pwd),
        }
        if new_password:
            if new_password != confirm_pwd:
                return redirect_with_flash(f"/web-users/{user_id}/edit", "error",
                                           "Passwords do not match.")
            if len(new_password) < 6:
                return redirect_with_flash(f"/web-users/{user_id}/edit", "error",
                                           "Password must be at least 6 characters.")
            fields["password_hash"] = hash_password(new_password)
        update_web_user(conn, user_id, _current_user_id(request), **fields)

        # ── Update global_users table if linked ─────────────────────────
        gu_id = int(global_user_id) if global_user_id.isdigit() else None
        if gu_id:
            from db import update_global_user
            gu_data = {
                "global_user_id":   gu_id,
                "employee_id":      gu_employee_id.strip() or None,
                "name":             gu_name.strip(),
                "designation":      gu_designation.strip() or None,
                "department_id":    int(gu_department_id) if gu_department_id.isdigit() else None,
                "section_id":       int(gu_section_id) if gu_section_id.isdigit() else None,
                "unit_id":          int(gu_unit_id) if gu_unit_id.isdigit() else None,
                "shift_id":         int(gu_shift_id) if gu_shift_id.isdigit() else None,
                "emp_type":         gu_emp_type or 'PERMANENT',
                "emp_status":       gu_emp_status or 'ACTIVE',
                "join_date":        gu_join_date or None,
                "level_grade":      gu_level_grade.strip() or None,
                "email":            gu_email.strip() or None,
                "phone":            gu_phone.strip() or None,
                "bank_number":      gu_bank_number.strip() or None,
                "appointment_date": gu_appointment_date or None,
                "dob":              gu_dob or None,
                "gender":           gu_gender or None,
            }
            update_global_user(conn, gu_id, gu_data, app_user_id=_current_user_id(request))

        ip = _get_client_ip(request)
        add_web_audit_log(conn, user_id, "", "updated", ip,
                          request.headers.get("User-Agent", "")[:255],
                          {"updated_by": request.session.get("username"),
                           "pwd_changed": bool(new_password)})
    finally:
        conn.close()
    return redirect_with_flash("/web-users", "success", "User updated successfully.")


@app.post("/web-users/{user_id}/toggle-active")
def web_user_toggle_active(request: Request, user_id: int):
    conn = get_connection()
    try:
        u = get_web_user_by_id(conn, user_id)
        if not u:
            return redirect_with_flash("/web-users", "error", "User not found.")
        new_active = not u["is_active"]
        update_web_user(conn, user_id, _current_user_id(request), is_active=new_active)
        action = "enabled" if new_active else "disabled"
        add_web_audit_log(conn, user_id, u["username"], action, _get_client_ip(request))
    finally:
        conn.close()
    msg = f"User '{u['username']}' {'enabled' if new_active else 'disabled'}."
    return redirect_with_flash("/web-users", "success", msg)


@app.get("/web-users/audit-log")
def web_users_audit_log(request: Request, uid: int | None = None):
    conn = get_connection()
    try:
        logs  = get_web_audit_logs(conn, web_user_id=uid, limit=300)
        users = get_all_web_users(conn)
    finally:
        conn.close()
    return render(templates, request, "web_users_audit.html", {
        "logs":         logs,
        "users":        users,
        "filter_uid":   uid,
        "COMPANY_NAME": COMPANY_NAME,
    })


def _fmt_schedule() -> list[str]:
    """Active pull times, DB-backed (see pull_schedule table) — the single
    source of truth used everywhere a schedule summary is shown."""
    try:
        conn = get_connection()
        try:
            rows = db_mod.get_pull_schedule(conn, active_only=True)
        finally:
            conn.close()
        return [f"{row['hour']:02d}:{row['minute']:02d}" for row in rows]
    except Exception:
        return [f"{hour:02d}:{minute:02d}" for hour, minute in SCHEDULE_TIMES]


def _db_status() -> dict:
    cfg = load_db_config()
    try:
        conn = get_connection()
        conn.close()
        return {
            "ok": True,
            "label": "Connected",
            "detail": f"{cfg.get('host')}:{cfg.get('port')} / {cfg.get('dbname')}",
        }
    except Exception as exc:
        return {
            "ok": False,
            "label": "Disconnected",
            "detail": str(exc),
        }


def _ping_device(ip: str, port: int, timeout: int = 3) -> dict:
    """TCP connect check. Returns {ok, ms, checked_at (UTC datetime)}."""
    start = _time.monotonic()
    try:
        sock = socket.create_connection((ip, port), timeout=timeout)
        sock.close()
        ms = int((_time.monotonic() - start) * 1000)
        return {"ok": True, "ms": ms, "checked_at": datetime.now(timezone.utc)}
    except Exception:
        return {"ok": False, "ms": None, "checked_at": datetime.now(timezone.utc)}


def _ping_devices_parallel(devices: list) -> None:
    """Ping all devices concurrently; adds 'ping' key to each device dict in-place."""
    def _ping_one(d):
        d["ping"] = _ping_device(d["ip_address"], int(d.get("port", 4370)))

    with concurrent.futures.ThreadPoolExecutor(max_workers=16) as executor:
        list(executor.map(_ping_one, devices))


def _dashboard_data(conn) -> dict:
    from db import get_global_user_count
    with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
        cur.execute("SELECT COUNT(*) FROM devices")
        device_count = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM devices WHERE is_active")
        active_device_count = cur.fetchone()[0]
        # Same population as /users (Global Users) and /reports/monthly —
        # keeps the "employee count" consistent across the whole app.
        employee_count = get_global_user_count(conn)
        cur.execute("""
            SELECT COALESCE(d.name, 'Unassigned') AS department_name, COUNT(*) AS cnt
            FROM global_users gu
            LEFT JOIN departments d ON d.id = gu.department_id
            WHERE gu.emp_status IS DISTINCT FROM 'DELETED'
            GROUP BY d.name
            ORDER BY cnt DESC
            LIMIT 6
        """)
        dept_rows = [dict(row) for row in cur.fetchall()]
        dept_total = sum(r['cnt'] for r in dept_rows) or 1
        department_distribution = [
            {**r, 'pct': round(r['cnt'] * 100 / dept_total)} for r in dept_rows
        ]
        cur.execute("""
            SELECT COUNT(DISTINCT user_id)
            FROM attendance_logs
            WHERE "timestamp" >= CURRENT_DATE
              AND "timestamp" < CURRENT_DATE + INTERVAL '1 day'
        """)
        punches_today = cur.fetchone()[0]
        cur.execute("""
            SELECT ps.*, d.name AS device_name
            FROM pull_sessions ps
            JOIN devices d ON ps.device_id = d.id
            ORDER BY ps.started_at DESC
            LIMIT 1
        """)
        latest = cur.fetchone()
        cur.execute("""
            SELECT d.id, d.name, d.ip_address, d.port, d.model, d.is_active,
                   ps.status AS last_status, ps.started_at AS last_started_at,
                   ps.completed_at AS last_completed_at, ps.records_pulled,
                   ps.new_inserts, ps.error_message
            FROM devices d
            LEFT JOIN LATERAL (
                SELECT *
                FROM pull_sessions ps
                WHERE ps.device_id = d.id
                ORDER BY ps.started_at DESC
                LIMIT 1
            ) ps ON TRUE
            ORDER BY d.name
        """)
        devices = [dict(row) for row in cur.fetchall()]
    _ping_devices_parallel(devices)
    with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
        cur.execute("""
            SELECT al.timestamp, COALESCE(al.name, e.name, 'Unknown') AS name,
                   al.user_id, al.punch_label, d.name AS device_name
            FROM attendance_logs al
            LEFT JOIN LATERAL (
                SELECT name FROM employees
                WHERE device_id = al.device_id AND user_id = al.user_id
                ORDER BY id LIMIT 1
            ) e ON TRUE
            JOIN devices d ON al.device_id = d.id
            ORDER BY al.timestamp DESC
            LIMIT 12
        """)
        recent_punches = [dict(row) for row in cur.fetchall()]
    return {
        "device_count": device_count,
        "active_device_count": active_device_count,
        "employee_count": employee_count,
        "punches_today": punches_today,
        "latest_session": dict(latest) if latest else None,
        "devices": devices,
        "recent_punches": recent_punches,
        "department_distribution": department_distribution,
    }


def write_devices_json_from_db(conn):
    devices = get_devices(conn)
    out = []
    for d in devices:
        out.append({
            "name": d["name"],
            "ip": d["ip_address"],
            "port": int(d.get("port", 4370)),
            "password": d.get("password", ""),
            "model": d.get("model", "unknown"),
            "is_active": bool(d.get("is_active", True)),
            "connection_timeout": 10,
        })
    path = os.path.join(BASE_DIR, "devices.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2)


@app.get("/")
def index(request: Request):
    db_status = _db_status()
    if not db_status["ok"]:
        return render(templates, request, "devices.html", {
            "db_status": db_status,
            "schedule_times": _fmt_schedule(),
            "scheduler_timezone": SCHEDULER_TIMEZONE,
            "stats": None,
            "devices": [],
        })
    conn = get_connection()
    try:
        stats = _dashboard_data(conn)
        return render(templates, request, "devices.html", {
            "db_status": db_status,
            "schedule_times": _fmt_schedule(),
            "scheduler_timezone": SCHEDULER_TIMEZONE,
            "stats": stats,
            "devices": stats["devices"],
        })
    finally:
        conn.close()


@app.get("/devices/add")
def device_add_form(request: Request):
    return render(templates, request, "device_form.html", {"device": None})


@app.post("/devices/add")
def device_add(request: Request,
               name: str = Form(...),
               ip: str = Form(...),
               port: int = Form(4370),
               password: str = Form(""),
               model: str = Form(""),
               is_active: str = Form("off"),
               force_udp: str = Form("off"),
               connection_timeout: int = Form(10)):
    conn = get_connection()
    try:
        device = {
            "name": name,
            "ip_address": ip,
            "port": int(port),
            "password": password,
            "model": model,
            "is_active": is_active == "on",
            "force_udp": force_udp == "on",
            "connection_timeout": max(5, int(connection_timeout)),
        }
        create_device(conn, device, app_user_id=_current_user_id(request))
        conn.commit()
        write_devices_json_from_db(conn)
        return redirect_with_flash("/", "success", f'Device "{name}" was added.')
    finally:
        conn.close()


@app.get("/devices/{device_id}/edit")
def device_edit_form(request: Request, device_id: int):
    conn = get_connection()
    try:
        device = get_device(conn, device_id)
        if not device:
            raise HTTPException(status_code=404, detail="Device not found")
        return render(templates, request, "device_form.html", {"device": device})
    finally:
        conn.close()


@app.post("/devices/{device_id}/edit")
def device_edit(request: Request, device_id: int,
                name: str = Form(...), ip: str = Form(...), port: int = Form(4370),
                password: str = Form(""), model: str = Form(""),
                is_active: str = Form("off"), force_udp: str = Form("off"),
                connection_timeout: int = Form(10)):
    conn = get_connection()
    try:
        device = {
            "name": name,
            "ip_address": ip,
            "port": int(port),
            "password": password,
            "model": model,
            "is_active": is_active == "on",
            "force_udp": force_udp == "on",
            "connection_timeout": max(5, int(connection_timeout)),
        }
        update_device(conn, device_id, device, app_user_id=_current_user_id(request))
        conn.commit()
        write_devices_json_from_db(conn)
        return redirect_with_flash("/", "success", f'Device "{name}" was updated.')
    finally:
        conn.close()


@app.post("/devices/{device_id}/delete")
def device_delete(request: Request, device_id: int):
    conn = get_connection()
    try:
        d = get_device(conn, device_id)
        if not d:
            raise HTTPException(status_code=404, detail="Device not found")
        snapshot = {
            "device_name": d["name"],
            "ip_address": d["ip_address"],
            "port": d.get("port", 4370),
        }
        delete_device(conn, device_id)
        conn.commit()
        write_devices_json_from_db(conn)
        return render(templates, request, "delete_result.html", snapshot)
    finally:
        conn.close()


@app.post("/devices/{device_id}/test")
def device_test(device_id: int):
    conn = get_connection()
    try:
        d = get_device(conn, device_id)
        if not d:
            raise HTTPException(status_code=404, detail="Device not found")
        ip = d["ip_address"]
        port = int(d.get("port", 4370))
        try:
            sock = socket.create_connection((ip, port), timeout=5)
            sock.close()
            return {"ok": True, "message": f"Connected to {ip}:{port}"}
        except Exception as exc:
            return {"ok": False, "message": str(exc)}
    finally:
        conn.close()


# ---- Users ---------------------------------------------------------------


_EMP_SORT_COLS = {
    'name':       'e.name',
    'user_id':    'e.user_id',
    'uid':        'e.uid',
    'created_at': 'e.created_at',
    'device':     'd.name',
}


def _int_param(v) -> int | None:
    """Convert a query-string value to int, returning None for empty / None inputs."""
    if v is None:
        return None
    s = str(v).strip()
    return int(s) if s else None


def _build_emp_query(device_id, search, date_str, sort_by, sort_dir, limit=1000, offset=0):
    col   = _EMP_SORT_COLS.get(sort_by, 'e.name')
    direc = 'DESC' if sort_dir == 'desc' else 'ASC'
    where = []
    params = []
    if device_id:
        where.append("d.id = %s")
        params.append(device_id)
    if search:
        where.append("(e.name ILIKE %s OR e.user_id ILIKE %s)")
        params += [f'%{search}%', f'%{search}%']
    if date_str:
        where.append("DATE(e.created_at) = %s")
        params.append(date_str)
    w = f"WHERE {' AND '.join(where)}" if where else ""
    sql = f"""
        SELECT e.id, e.uid, e.user_id, e.name, e.privilege, e.card,
               e.global_user_id, d.name AS device_name, d.id AS device_id,
               e.created_at
        FROM employees e
        JOIN devices d ON e.device_id = d.id
        {w}
        ORDER BY {col} {direc}
        LIMIT {limit} OFFSET {offset}
    """
    return sql, params


def _build_emp_count_query(device_id, search, date_str):
    where = []
    params = []
    if device_id:
        where.append("d.id = %s")
        params.append(device_id)
    if search:
        where.append("(e.name ILIKE %s OR e.user_id ILIKE %s)")
        params += [f'%{search}%', f'%{search}%']
    if date_str:
        where.append("DATE(e.created_at) = %s")
        params.append(date_str)
    w = f"WHERE {' AND '.join(where)}" if where else ""
    sql = f"""
        SELECT COUNT(*) FROM employees e
        JOIN devices d ON e.device_id = d.id {w}
    """
    return sql, params


_USERS_PER_PAGE = 25


def _gu_sort_key(u, sort_by: str):
    def _num(val):
        s = (val or '').strip()
        try: return (0, int(s), s)
        except: return (1, 0, s.lower())
    if sort_by == 'att_id':    return _num(u.get('global_user_id'))
    if sort_by == 'emp_id':    return _num(u.get('employee_id'))
    if sort_by == 'name':      return (u.get('name') or '').lower()
    if sort_by == 'department':return ((u.get('department_name') or '').lower(), (u.get('name') or '').lower())
    if sort_by == 'section':   return ((u.get('section_name') or '').lower(), (u.get('name') or '').lower())
    if sort_by == 'shift':     return ((u.get('shift_name') or '').lower(), (u.get('name') or '').lower())
    return _num(u.get('global_user_id'))


def _gu_query_string(gu_search, gu_directorate, gu_department, gu_section,
                      gu_unit, gu_shift, gu_status) -> str:
    """Filter-only query string (no page/sort) so pagination and sort links
    can append their own &page=/&gu_sort_by= safely."""
    from urllib.parse import urlencode
    parts = {
        'tab': 'global',
        'gu_search': gu_search or '',
        'gu_directorate': gu_directorate or '',
        'gu_department': gu_department or '',
        'gu_section': gu_section or '',
        'gu_unit': gu_unit or '',
        'gu_shift': gu_shift or '',
        'gu_status': gu_status or '',
    }
    return urlencode({k: v for k, v in parts.items() if v})


@app.get("/users")
def users_index(
    request: Request,
    tab:              str = 'global',
    # global user filters + sort
    gu_search:        str | None = None,
    gu_directorate:   str | None = None,
    gu_department:    str | None = None,
    gu_section:       str | None = None,
    gu_unit:          str | None = None,
    gu_shift:         str | None = None,
    gu_status:        str | None = None,
    gu_sort_by:       str = 'att_id',
    gu_sort_dir:      str = 'asc',
    page:             str | None = None,
    # device employee filters
    device_id:        str | None = None,
    search:           str | None = None,
    date_str:         str | None = None,
    sort_by:          str = 'uid',
    sort_dir:         str = 'asc',
    emp_page:         str | None = None,
):
    _VALID_GU_SORTS = {'att_id', 'emp_id', 'name', 'department', 'section', 'shift'}
    gu_sort_by  = gu_sort_by  if gu_sort_by  in _VALID_GU_SORTS else 'att_id'
    gu_sort_dir = 'desc' if gu_sort_dir == 'desc' else 'asc'

    conn = get_connection()
    try:
        from db import (get_all_directorates, get_all_departments,
                        get_all_sections, get_all_units, get_all_shifts)
        # Global users with filters + sort + pagination
        all_users = list_global_users(
            conn,
            search=gu_search or None,
            directorate_id=_int_param(gu_directorate),
            department_id=_int_param(gu_department),
            section_id=_int_param(gu_section),
            unit_id=_int_param(gu_unit),
            shift_id=_int_param(gu_shift),
            emp_status=gu_status or None,
        )
        all_users.sort(key=lambda u: _gu_sort_key(u, gu_sort_by),
                       reverse=(gu_sort_dir == 'desc'))

        total_users  = len(all_users)
        page_num     = max(1, _int_param(page) or 1)
        total_pages  = max(1, (total_users + _USERS_PER_PAGE - 1) // _USERS_PER_PAGE)
        page_num     = min(page_num, total_pages)
        users        = all_users[(page_num - 1) * _USERS_PER_PAGE : page_num * _USERS_PER_PAGE]

        # Device employees with pagination
        cnt_sql, cnt_params = _build_emp_count_query(_int_param(device_id), search, date_str)
        with conn.cursor() as cur:
            cur.execute(cnt_sql, cnt_params)
            total_employees = cur.fetchone()[0]

        emp_page_num  = max(1, _int_param(emp_page) or 1)
        emp_total_pages = max(1, (total_employees + _USERS_PER_PAGE - 1) // _USERS_PER_PAGE)
        emp_page_num  = min(emp_page_num, emp_total_pages)

        sql, params = _build_emp_query(
            _int_param(device_id), search, date_str, sort_by, sort_dir,
            limit=_USERS_PER_PAGE, offset=(emp_page_num - 1) * _USERS_PER_PAGE,
        )
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(sql, params)
            employees = [dict(row) for row in cur.fetchall()]

        devices      = get_devices(conn)
        directorates = get_all_directorates(conn)
        departments  = get_all_departments(conn)
        sections     = get_all_sections(conn)
        units        = get_all_units(conn)
        shifts       = get_all_shifts(conn)

        return render(templates, request, "users.html", {
            "tab":            tab if tab in ('global', 'device') else 'global',
            "users":          users,
            "total_users":    total_users,
            "page":           page_num,
            "total_pages":    total_pages,
            "page_range":     _page_range(page_num, total_pages),
            "gu_search":      gu_search or '',
            "gu_directorate": gu_directorate or '',
            "gu_department":  gu_department or '',
            "gu_section":     gu_section or '',
            "gu_unit":        gu_unit or '',
            "gu_shift":       gu_shift or '',
            "gu_status":      gu_status or '',
            "gu_sort_by":     gu_sort_by,
            "gu_sort_dir":    gu_sort_dir,
            "gu_qs":          _gu_query_string(gu_search, gu_directorate, gu_department,
                                                gu_section, gu_unit, gu_shift, gu_status),
            "employees":      employees,
            "total_employees": total_employees,
            "emp_page":       emp_page_num,
            "emp_total_pages": emp_total_pages,
            "emp_page_range": _page_range(emp_page_num, emp_total_pages),
            "devices":        devices,
            "directorates":   directorates,
            "departments":    departments,
            "sections":       sections,
            "units":          units,
            "shifts":         shifts,
            "selected_device_id": device_id or '',
            "search":         search or '',
            "date_str":       date_str or '',
            "sort_by":        sort_by,
            "sort_dir":       sort_dir,
        })
    finally:
        conn.close()


@app.get("/users/export")
def users_export(
    device_id: str | None = None,
    search: str | None = None,
    date_str: str | None = None,
    sort_by: str = 'name',
    sort_dir: str = 'asc',
):
    conn = get_connection()
    try:
        sql, params = _build_emp_query(_int_param(device_id), search, date_str, sort_by, sort_dir, limit=10000)
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(sql, params)
            rows = [dict(r) for r in cur.fetchall()]
    finally:
        conn.close()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['Device', 'UID', 'User ID', 'Name', 'Privilege', 'Card',
                     'Linked to Global', 'Created (AD)', 'Created (BS)'])
    for r in rows:
        writer.writerow([
            r.get('device_name', ''),
            r.get('uid', ''),
            r.get('user_id', ''),
            r.get('name', ''),
            r.get('privilege', 0),
            r.get('card', '') or '',
            'Yes' if r.get('global_user_id') else 'No',
            nepali_utils.jinja_fmt_dt(r.get('created_at')),
            nepali_utils.jinja_bs_datetime(r.get('created_at')),
        ])

    fname = f"employees_{date.today().isoformat()}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


@app.get("/users/add")
def user_add_form(request: Request):
    conn = get_connection()
    try:
        from db import (get_all_directorates, get_all_departments,
                        get_all_sections, get_all_units, get_all_shifts)
        return render(templates, request, "user_form.html", {
            "user":         None,
            "directorates": get_all_directorates(conn),
            "departments":  get_all_departments(conn),
            "sections":     get_all_sections(conn),
            "units":        get_all_units(conn),
            "shifts":       get_all_shifts(conn),
        })
    finally:
        conn.close()


_EXTRA_PROFILE_FIELDS = [
    'name_nep', 'citizenship_no', 'national_id_card_no', 'full_address',
    'local_body', 'state', 'ward_no', 'pan_no', 'bank_name', 'bank_branch',
    'initial_appointment_date', 'retirement_date',
    'dob', 'gender', 'designation', 'level_grade', 'emp_type', 'emp_status',
    'join_date',
]


def _extract_profile_fields(form) -> dict:
    """Only include fields actually present in the submission — update_global_user
    merges onto the current row, so a field genuinely absent from the form
    (vs. present-but-cleared) must not overwrite existing data with None."""
    out = {}
    for f in _EXTRA_PROFILE_FIELDS:
        if f in form:
            v = (form.get(f) or '').strip()
            out[f] = v or None
    # Checkboxes are always rendered on this form; unlike text fields, an
    # absent checkbox means "unchecked", not "field not on this form".
    out['is_technical'] = bool(form.get('is_technical'))
    return out


@app.post("/users/add")
async def user_add(request: Request):
    form            = await request.form()
    global_user_id  = (form.get('global_user_id') or '').strip()
    name            = (form.get('name') or '').strip()
    privilege       = int(form.get('privilege') or 0)
    card            = (form.get('card') or '').strip() or None
    employee_id     = (form.get('employee_id') or '').strip() or None
    bank_number     = (form.get('bank_number') or '').strip() or None
    email           = (form.get('email') or '').strip() or None
    phone           = (form.get('phone') or '').strip() or None
    department_id   = _int_param(form.get('department_id'))
    section_id      = _int_param(form.get('section_id'))
    unit_id         = _int_param(form.get('unit_id'))
    shift_id        = _int_param(form.get('shift_id'))
    enroll_devices  = (form.get('enroll_devices') or '').strip()

    if not global_user_id or not name:
        return redirect_with_flash('/users/add', 'error', 'Attendance ID and name are required.')

    conn = get_connection()
    try:
        gid = create_global_user(
            conn, global_user_id, name, privilege, card,
            app_user_id=_current_user_id(request),
            employee_id=employee_id, bank_number=bank_number,
            email=email, phone=phone,
            department_id=department_id, section_id=section_id,
            unit_id=unit_id, shift_id=shift_id,
        )
        conn.commit()
        extra = _extract_profile_fields(form)
        update_global_user(conn, gid, extra, app_user_id=_current_user_id(request))
        conn.commit()
        if enroll_devices:
            ids = [int(x) for x in enroll_devices.split(",") if x.strip()]
            for did in ids:
                d = get_device(conn, did)
                if not d:
                    continue
                device_cfg = device_config_from_row(d)
                try:
                    puller_mod.push_global_user_to_device(
                        device_cfg,
                        {"global_user_id": global_user_id, "name": name,
                         "privilege": privilege, "card": card}
                    )
                except Exception:
                    pass

        # Auto-create a self-service "employee" web login for this user
        # (username = their attendance device ID). Never blocks user creation.
        login_msg = ''
        try:
            from db import create_employee_login
            result = create_employee_login(conn, gid, created_by=_current_user_id(request))
            conn.commit()
            if result.get('created'):
                login_msg = (f' Web login created — username "{result["username"]}", '
                             f'temporary password same as the username (must be changed on first login).')
        except Exception:
            conn.rollback()

        return redirect_with_flash("/users", "success", f'User "{name}" was created.{login_msg}')
    finally:
        conn.close()


@app.get("/users/{global_id}/edit")
def user_edit_form(request: Request, global_id: int):
    conn = get_connection()
    try:
        from db import (get_all_directorates, get_all_departments,
                        get_all_sections, get_all_units, get_all_shifts)
        user = get_global_user(conn, global_id)
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        return render(templates, request, "user_form.html", {
            "user":          user,
            "directorates":  get_all_directorates(conn),
            "departments":   get_all_departments(conn),
            "sections":      get_all_sections(conn),
            "units":         get_all_units(conn),
            "shifts":        get_all_shifts(conn),
        })
    finally:
        conn.close()


@app.post("/users/{global_id}/edit")
async def user_edit(request: Request, global_id: int):
    form = await request.form()
    data = {
        'global_user_id': (form.get('global_user_id') or '').strip(),
        'employee_id':    (form.get('employee_id') or '').strip() or None,
        'name':           (form.get('name') or '').strip(),
        'privilege':      int(form.get('privilege') or 0),
        'card':           (form.get('card') or '').strip() or None,
        'bank_number':    (form.get('bank_number') or '').strip() or None,
        'email':          (form.get('email') or '').strip() or None,
        'phone':          (form.get('phone') or '').strip() or None,
        'department_id':  _int_param(form.get('department_id')),
        'section_id':     _int_param(form.get('section_id')),
        'unit_id':        _int_param(form.get('unit_id')),
        'shift_id':       _int_param(form.get('shift_id')),
    }
    data.update(_extract_profile_fields(form))
    if not data['global_user_id'] or not data['name']:
        return redirect_with_flash(f'/users/{global_id}/edit', 'error', 'Attendance ID and name are required.')
    conn = get_connection()
    try:
        update_global_user(conn, global_id, data, app_user_id=_current_user_id(request))
        conn.commit()
        return redirect_with_flash('/users', 'success', f'User "{data["name"]}" updated.')
    finally:
        conn.close()


@app.post("/users/{global_id}/delete")
def user_delete(request: Request, global_id: int):
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT global_user_id, name FROM global_users WHERE id = %s", (global_id,))
            row = cur.fetchone()
            if not row:
                return redirect_with_flash("/users", "warning", "User was not found.")
            global_user_id, user_name = row[0], row[1]
        devices = get_devices(conn)
        for d in devices:
            device_cfg = device_config_from_row(d)
            try:
                puller_mod.delete_user_from_device(device_cfg, global_user_id)
            except Exception:
                pass
        delete_global_user(conn, global_id)
        conn.commit()
        return redirect_with_flash("/users", "success", f'User "{user_name}" was deleted from the database and devices.')
    finally:
        conn.close()


@app.post("/users/{global_id}/soft-delete")
def user_soft_delete(request: Request, global_id: int):
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT global_user_id, name FROM global_users WHERE id = %s", (global_id,))
            row = cur.fetchone()
            if not row:
                return redirect_with_flash("/users", "warning", "User not found.")
            global_user_id, user_name = row[0], row[1]
        devices = get_devices(conn)
        for d in devices:
            device_cfg = device_config_from_row(d)
            try:
                puller_mod.delete_user_from_device(device_cfg, global_user_id)
            except Exception:
                pass
        from db import update_global_user
        update_global_user(conn, global_id, {'emp_status': 'DELETED'}, app_user_id=_current_user_id(request))
        conn.commit()
        return redirect_with_flash("/users", "success", f'User "{user_name}" soft-deleted (removed from devices, kept in software).')
    finally:
        conn.close()


@app.post("/users/{global_id}/restore")
def user_restore(request: Request, global_id: int):
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT name FROM global_users WHERE id = %s", (global_id,))
            row = cur.fetchone()
            if not row:
                return redirect_with_flash("/users", "warning", "User not found.")
        from db import update_global_user
        update_global_user(conn, global_id, {'emp_status': 'ACTIVE'}, app_user_id=_current_user_id(request))
        conn.commit()
        return redirect_with_flash("/users", "success", f'User restored to Active status.')
    finally:
        conn.close()


@app.get("/users/{global_id}/push")
def user_push_form(request: Request, global_id: int):
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("SELECT id, global_user_id, name, privilege, card FROM global_users WHERE id = %s", (global_id,))
            user = cur.fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        devices = get_devices(conn)
        return render(templates, request, 'user_push.html', {'user': dict(user), 'devices': devices})
    finally:
        conn.close()


@app.post("/users/{global_id}/push")
def user_push(request: Request, global_id: int, device_id: int = Form(...)):
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT global_user_id, name, privilege, card FROM global_users WHERE id = %s", (global_id,))
            row = cur.fetchone()
            if not row:
                return redirect_with_flash('/users', 'warning', 'User was not found.')
            global_user_id, name, privilege, card = row
        d = get_device(conn, device_id)
        if not d:
            raise HTTPException(status_code=404, detail='Device not found')
        device_cfg = device_config_from_row(d)
        res = puller_mod.push_global_user_to_device(device_cfg, {'global_user_id': global_user_id, 'name': name, 'privilege': privilege, 'card': card})
        return render(templates, request, 'user_push_result.html', {'result': res, 'user': {'global_user_id': global_user_id, 'name': name}})
    finally:
        conn.close()


# ---- Employees page ----------------------------------------------------


_EMP_PAGE_SIZE = 25


@app.get("/employees")
def employees_page(
    request: Request,
    search:         str | None = None,
    device_id:      str | None = None,
    link_status:    str | None = None,
    department_id:  str | None = None,
    page:           str | None = None,
):
    """Employee management page with search, device filter, link-status filter, and attendance stats."""
    conn = get_connection()
    try:
        devices = get_devices(conn)
        device_id_int = _int_param(device_id)
        dept_id_int = _int_param(department_id)
        page_num = max(1, _int_param(page) or 1)
        search_q = (search or '').strip()

        # ── Count totals ──
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM employees")
            total_employees = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM employees WHERE global_user_id IS NOT NULL")
            total_linked = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM employees WHERE global_user_id IS NULL")
            total_unlinked = total_employees - total_linked
            cur.execute("""
                SELECT COUNT(DISTINCT user_id)
                FROM attendance_logs
                WHERE "timestamp" >= CURRENT_DATE
                  AND "timestamp" < CURRENT_DATE + INTERVAL '1 day'
            """)
            punches_today = cur.fetchone()[0]

        # ── Build query ──
        where = []
        params = []
        if device_id_int:
            where.append("e.device_id = %s")
            params.append(device_id_int)
        if search_q:
            where.append("(e.name ILIKE %s OR e.user_id ILIKE %s)")
            params += [f'%{search_q}%', f'%{search_q}%']
        if link_status == 'linked':
            where.append("e.global_user_id IS NOT NULL")
        elif link_status == 'unlinked':
            where.append("e.global_user_id IS NULL")
        if dept_id_int:
            where.append("gu.department_id = %s")
            params.append(dept_id_int)

        where_clause = f"WHERE {' AND '.join(where)}" if where else ""

        # Fetch all departments for filter dropdown
        all_departments = get_all_departments(conn)

        with conn.cursor() as cur:
            cur.execute(
                f"SELECT COUNT(*) FROM employees e {where_clause}",
                tuple(params))
            total_filtered = cur.fetchone()[0]

        total_pages = max(1, (total_filtered + _EMP_PAGE_SIZE - 1) // _EMP_PAGE_SIZE)
        page_num = min(page_num, total_pages)
        offset = (page_num - 1) * _EMP_PAGE_SIZE

        # ── Fetch employees ──
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(f"""
                SELECT e.id, e.uid, e.user_id, e.name, e.privilege, e.card,
                       e.global_user_id, d.name AS device_name, d.id AS device_id,
                       gu.employee_id, gu.email, gu.phone, gu.bank_number,
                       gu.emp_type, gu.emp_status, gu.join_date, gu.designation,
                       gu.department_id, dep.name AS department_name,
                       gu.section_id, sec.name AS section_name,
                       gu.unit_id, unt.name AS unit_name
                FROM employees e
                JOIN devices d ON e.device_id = d.id
                LEFT JOIN global_users gu ON gu.id = e.global_user_id
                LEFT JOIN departments dep ON dep.id = gu.department_id
                LEFT JOIN sections sec ON sec.id = gu.section_id
                LEFT JOIN units unt ON unt.id = gu.unit_id
                {where_clause}
                ORDER BY e.name ASC
                LIMIT %s OFFSET %s
            """, tuple(params) + (_EMP_PAGE_SIZE, offset))
            employees = [dict(row) for row in cur.fetchall()]

        # ── Fetch today's attendance counts per employee ──
        emp_attendance = {}
        if employees:
            emp_pairs = [(e['device_id'], e['user_id']) for e in employees]
            placeholders = ','.join(['(%s, %s)'] * len(emp_pairs))
            flat_params = []
            for did, uid in emp_pairs:
                flat_params += [did, uid]
            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                cur.execute(f"""
                    SELECT device_id, user_id, punch_label, COUNT(*) as cnt
                    FROM attendance_logs
                    WHERE "timestamp" >= CURRENT_DATE
                      AND "timestamp" < CURRENT_DATE + INTERVAL '1 day'
                      AND (device_id, user_id) IN ({placeholders})
                    GROUP BY device_id, user_id, punch_label
                """, tuple(flat_params))
                for row in cur.fetchall():
                    key = (row['device_id'], row['user_id'])
                    emp_attendance.setdefault(key, {})
                    emp_attendance[key][row['punch_label']] = row['cnt']

        schedule_times = _fmt_schedule()

        return render(templates, request, "employees.html", {
            "employees":        employees,
            "devices":          devices,
            "all_departments":  all_departments,
            "total_employees":  total_employees,
            "total_linked":     total_linked,
            "total_unlinked":   total_unlinked,
            "punches_today":    punches_today,
            "total_filtered":   total_filtered,
            "page":             page_num,
            "total_pages":      total_pages,
            "page_range":       _page_range(page_num, total_pages),
            "search":           search_q,
            "selected_device_id": device_id or '',
            "selected_department_id": department_id or '',
            "link_status":      link_status or '',
            "emp_attendance":   emp_attendance,
            "schedule_times":   schedule_times,
        })
    finally:
        conn.close()


@app.get("/employees/{emp_id}/edit")
def employee_edit_page(request: Request, emp_id: int):
    """Employee edit page showing cross-device attendance records."""
    conn = get_connection()
    try:
        emp = get_employee_with_device(conn, emp_id)
        if not emp:
            return redirect_with_flash("/employees", "error", "Employee not found.")

        # Fetch all employee records across devices for this user_id
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("""
                SELECT e.id, e.uid, e.user_id, e.name, e.privilege, e.card,
                       e.global_user_id, d.name AS device_name, d.id AS device_id
                FROM employees e
                JOIN devices d ON e.device_id = d.id
                WHERE e.user_id = %s
                ORDER BY d.name
            """, (emp['user_id'],))
            cross_device_records = [dict(row) for row in cur.fetchall()]

        # Fetch today's attendance
        today_attendance = []
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("""
                SELECT al.*, d.name AS device_name
                FROM attendance_logs al
                JOIN devices d ON al.device_id = d.id
                WHERE al.user_id = %s
                  AND al."timestamp" >= CURRENT_DATE
                  AND al."timestamp" < CURRENT_DATE + INTERVAL '1 day'
                ORDER BY al.timestamp
            """, (emp['user_id'],))
            today_attendance = [dict(row) for row in cur.fetchall()]

        # All devices for set-global form
        devices = get_devices(conn)

        # Fetch all global users for the linking dropdown
        all_global_users = list_global_users(conn)

        return render(templates, request, "employee_edit.html", {
            "employee":             emp,
            "cross_device_records": cross_device_records,
            "today_attendance":     today_attendance,
            "devices":              devices,
            "global_users":         all_global_users,
        })
    finally:
        conn.close()


@app.post("/employees/{emp_id}/edit")
async def employee_edit_post(request: Request, emp_id: int):
    """Update employee name on device and in DB."""
    form = await request.form()
    new_name = (form.get('name') or '').strip()
    if not new_name:
        return redirect_with_flash(f"/employees/{emp_id}/edit", "error", "Name is required.")

    conn = get_connection()
    try:
        emp = get_employee_with_device(conn, emp_id)
        if not emp:
            return redirect_with_flash("/employees", "error", "Employee not found.")

        # Update in DB
        with conn.cursor() as cur:
            cur.execute("UPDATE employees SET name = %s WHERE id = %s", (new_name, emp_id))
        conn.commit()

        return redirect_with_flash(f"/employees/{emp_id}/edit", "success", f"Employee updated to '{new_name}'.")
    finally:
        conn.close()


@app.post("/employees/{emp_id}/delete")
def employee_delete(request: Request, emp_id: int):
    conn = get_connection()
    try:
        emp = get_employee_with_device(conn, emp_id)
        if not emp:
            return redirect_with_flash("/employees", "warning", "Employee not found.")
        device_cfg = device_config_from_row(emp)
        result = puller_mod.delete_employee_by_uid(device_cfg, emp["uid"])
        delete_employee_record(conn, emp_id)
        conn.commit()
        if result["ok"]:
            msg = f'Removed {emp["name"] or emp["user_id"]} (UID {emp["uid"]}) from {emp["device_name"]}.'
        else:
            msg = f'Removed from DB; device said: {result["message"]}'
        return redirect_with_flash("/employees", "success" if result["ok"] else "warning", msg)
    finally:
        conn.close()


@app.post("/employees/bulk-delete")
async def employees_bulk_delete(request: Request):
    form = await request.form()
    ids_raw = form.getlist("ids")
    emp_ids = [int(x) for x in ids_raw if x.strip()]
    if not emp_ids:
        return redirect_with_flash("/employees", "warning", "No employees selected.")
    conn = get_connection()
    try:
        rows = get_employees_with_device(conn, emp_ids)
        ok_count = 0
        fail_count = 0
        for emp in rows:
            device_cfg = device_config_from_row(emp)
            result = puller_mod.delete_employee_by_uid(device_cfg, emp["uid"])
            if result["ok"]:
                ok_count += 1
            else:
                fail_count += 1
        bulk_delete_employee_records(conn, emp_ids)
        conn.commit()
        msg = f"Deleted {ok_count} employee(s) from device(s)."
        if fail_count:
            msg += f" {fail_count} had device errors but were removed from DB."
        return redirect_with_flash("/employees", "success", msg)
    finally:
        conn.close()


# ---- Employee migrate to global user ------------------------------------


@app.post("/employees/{emp_id}/migrate")
async def employee_migrate(request: Request, emp_id: int):
    form = await request.form()

    global_user_id = (form.get('global_user_id') or '').strip()
    name           = (form.get('name') or '').strip()
    employee_id    = (form.get('employee_id') or '').strip() or None
    bank_number    = (form.get('bank_number') or '').strip() or None
    email          = (form.get('email') or '').strip() or None
    phone          = (form.get('phone') or '').strip() or None
    department_id  = _int_param(form.get('department_id'))
    section_id     = _int_param(form.get('section_id'))
    unit_id        = _int_param(form.get('unit_id'))
    shift_id       = _int_param(form.get('shift_id'))

    if not global_user_id or not name:
        return redirect_with_flash('/users?tab=device', 'error', 'Attendance ID and name are required.')

    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("SELECT * FROM employees WHERE id = %s", (emp_id,))
            emp = cur.fetchone()
        if not emp:
            return redirect_with_flash('/users?tab=device', 'error', 'Employee not found.')

        existing = find_global_user_by_global_id(conn, global_user_id)
        if existing:
            gid = existing['id']
        else:
            gid = create_global_user(
                conn, global_user_id, name,
                int(emp['privilege'] or 0), emp.get('card') or None,
                app_user_id=_current_user_id(request),
                employee_id=employee_id, bank_number=bank_number,
                email=email, phone=phone,
                department_id=department_id, section_id=section_id,
                unit_id=unit_id, shift_id=shift_id,
            )
        with conn.cursor() as cur:
            cur.execute("UPDATE employees SET global_user_id = %s WHERE id = %s", (gid, emp_id))
        conn.commit()
        return redirect_with_flash('/users?tab=device', 'success',
                                   f'Employee "{name}" migrated to global users.')
    except Exception as exc:
        return redirect_with_flash('/users?tab=device', 'error', str(exc))
    finally:
        conn.close()


# ---- Global users CSV / Excel / Print export -----------------------------


def _filtered_global_users(gu_search, gu_directorate, gu_department,
                            gu_section, gu_unit, gu_shift, gu_status):
    conn = get_connection()
    try:
        return list_global_users(
            conn,
            search=gu_search or None,
            directorate_id=_int_param(gu_directorate),
            department_id=_int_param(gu_department),
            section_id=_int_param(gu_section),
            unit_id=_int_param(gu_unit),
            shift_id=_int_param(gu_shift),
            emp_status=gu_status or None,
        )
    finally:
        conn.close()


_GU_EXPORT_COLUMNS = [
    ('SN',            None),
    ('Employee ID',   'employee_id'),
    ('Att. Device ID','global_user_id'),
    ('Name',          'name'),
    ('Nepali Name',   'name_nep'),
    ('Directorate',   'directorate_name'),
    ('Department',    'department_name'),
    ('Section',       'section_name'),
    ('Unit',          'unit_name'),
    ('Default Shift', 'shift_name'),
    ('Designation',   'designation'),
    ('Emp. Type',     'emp_type'),
    ('Status',        'emp_status'),
    ('Email',         'email'),
    ('Phone',         'phone'),
    ('Gender',        'gender'),
    ('DOB',           'dob'),
    ('Join Date',     'join_date'),
    ('Citizenship No.', 'citizenship_no'),
    ('National ID No.', 'national_id_card_no'),
    ('PAN No.',       'pan_no'),
    ('Bank Name',     'bank_name'),
    ('Bank Branch',   'bank_branch'),
    ('Bank Number',   'bank_number'),
    ('Full Address',  'full_address'),
]


@app.get("/users/global-export")
def global_users_export(
    gu_search:      str | None = None,
    gu_directorate: str | None = None,
    gu_department:  str | None = None,
    gu_section:     str | None = None,
    gu_unit:        str | None = None,
    gu_shift:       str | None = None,
    gu_status:      str | None = None,
):
    rows = _filtered_global_users(gu_search, gu_directorate, gu_department,
                                   gu_section, gu_unit, gu_shift, gu_status)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([label for label, _ in _GU_EXPORT_COLUMNS])
    for i, r in enumerate(rows, 1):
        writer.writerow([
            (i if key is None else (r.get(key) or ''))
            for _, key in _GU_EXPORT_COLUMNS
        ])

    fname = f"global_users_{date.today().isoformat()}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


@app.get("/users/global-export/excel")
def global_users_export_excel(
    gu_search:      str | None = None,
    gu_directorate: str | None = None,
    gu_department:  str | None = None,
    gu_section:     str | None = None,
    gu_unit:        str | None = None,
    gu_shift:       str | None = None,
    gu_status:      str | None = None,
):
    rows = _filtered_global_users(gu_search, gu_directorate, gu_department,
                                   gu_section, gu_unit, gu_shift, gu_status)

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Global Users"

    headers = [label for label, _ in _GU_EXPORT_COLUMNS]
    ws.append(headers)
    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows, 1):
        ws.append([
            (i if key is None else (r.get(key) or ''))
            for _, key in _GU_EXPORT_COLUMNS
        ])

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"global_users_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


@app.get("/users/global-print")
def global_users_print(
    request: Request,
    gu_search:      str | None = None,
    gu_directorate: str | None = None,
    gu_department:  str | None = None,
    gu_section:     str | None = None,
    gu_unit:        str | None = None,
    gu_shift:       str | None = None,
    gu_status:      str | None = None,
):
    rows = _filtered_global_users(gu_search, gu_directorate, gu_department,
                                   gu_section, gu_unit, gu_shift, gu_status)
    return render(templates, request, 'users_print.html', {
        'rows':        rows,
        'total':       len(rows),
        'now_str':     _npt_now_str(),
        'COMPANY_NAME': COMPANY_NAME,
        'filters': {
            'search': gu_search or '', 'status': gu_status or '',
        },
    })


# ---- Device backup download ---------------------------------------------


@app.get("/devices/{device_id}/backup")
def device_backup(device_id: int):
    conn = get_connection()
    try:
        d = get_device(conn, device_id)
        if not d:
            raise HTTPException(status_code=404, detail="Device not found")
    finally:
        conn.close()

    device_cfg = device_config_from_row(d)
    result = puller_mod.get_device_backup(device_cfg)
    if not result["ok"]:
        raise HTTPException(status_code=502, detail=result.get("message", "Backup failed"))

    filename = f"backup_{d['name']}_{date.today().isoformat()}.json"
    payload = json.dumps(result["data"], indent=2, ensure_ascii=False)
    return StreamingResponse(
        iter([payload]),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---- Migrate users between devices --------------------------------------


@app.get("/migrate")
def migrate_form(request: Request, source_id: int | None = None):
    conn = get_connection()
    try:
        devices = get_devices(conn)
        source_users = None
        source_device = None
        error = None
        if source_id:
            source_device = get_device(conn, source_id)
            if source_device:
                device_cfg = device_config_from_row(source_device)
                try:
                    raw = puller_mod.list_device_users(device_cfg)
                    source_users = [
                        {
                            "uid": getattr(u, "uid", ""),
                            "user_id": getattr(u, "user_id", ""),
                            "name": getattr(u, "name", ""),
                            "card": getattr(u, "card", ""),
                        }
                        for u in raw
                    ]
                except Exception as exc:
                    msg = str(exc)
                    if "timed out" in msg.lower() or "timeout" in msg.lower():
                        msg = (f"Could not load device users: connection timed out. "
                               f"Check that the device is powered on and reachable on the network. "
                               f"(raw: {msg})")
                    else:
                        msg = f"Could not load device users: {msg}"
                    error = msg
        return render(templates, request, "migrate.html", {
            "devices": devices,
            "source_users": source_users,
            "source_device": source_device,
            "source_id": source_id,
            "error": error,
        })
    finally:
        conn.close()


@app.post("/migrate")
async def migrate_execute(request: Request):
    form = await request.form()
    source_device_id = int(form.get("source_device_id") or 0)
    target_device_id = int(form.get("target_device_id") or 0)
    uids_raw = form.getlist("uids")
    uids = [int(u) for u in uids_raw if u.strip()] if uids_raw else None

    if not source_device_id or not target_device_id:
        return redirect_with_flash("/migrate", "warning", "Select both source and target devices.")
    if source_device_id == target_device_id:
        return redirect_with_flash("/migrate", "warning", "Source and target must be different devices.")

    conn = get_connection()
    try:
        source_d = get_device(conn, source_device_id)
        target_d = get_device(conn, target_device_id)
        if not source_d or not target_d:
            raise HTTPException(status_code=404, detail="Device not found")
    finally:
        conn.close()

    source_cfg = device_config_from_row(source_d)
    target_cfg = device_config_from_row(target_d)
    result = puller_mod.migrate_users_to_device(source_cfg, target_cfg, uids)
    return render(templates, request, "migrate_result.html", {
        "result": result,
        "source_device": source_d,
        "target_device": target_d,
    })


# ---- Pull / Sync endpoints ----------------------------------------------


@app.post("/devices/{device_id}/pull")
def device_pull(request: Request, device_id: int):
    conn = get_connection()
    try:
        d = get_device(conn, device_id)
        if not d:
            raise HTTPException(status_code=404, detail="Device not found")
        device_cfg = device_config_from_row(d)
        started_at = datetime.now(timezone.utc)
        session_id = None
        error_message = None
        success = False
        user_count = 0
        records_pulled = 0
        new_inserts = 0
        completed_at = None

        try:
            device_db_id = db_mod.upsert_device(conn, device_cfg)
            conn.commit()
            session_id = db_mod.start_pull_session(conn, device_db_id, started_at)
            conn.commit()
            result = puller_mod.pull_device(device_cfg)

            if not result.success:
                error_message = result.error
                db_mod.complete_pull_session(conn, session_id, 0, 0, 'failed',
                                             result.error, result.error_traceback)
                conn.commit()
            else:
                for user in result.users:
                    gu = db_mod.find_global_user_by_global_id(conn, str(user.user_id))
                    if gu:
                        try:
                            setattr(user, 'global_user_id', gu['id'])
                        except Exception:
                            pass
                    try:
                        db_mod.upsert_employee(conn, device_db_id, user)
                    except Exception:
                        conn.rollback()

                user_count = len(result.users)
                employee_map = db_mod.build_employee_map(conn, device_db_id)
                records = [attendance_to_dict(a) for a in result.attendance]
                records_pulled = len(records)
                new_inserts = db_mod.insert_attendance_batch(conn, device_db_id, records, employee_map)
                db_mod.complete_pull_session(conn, session_id, records_pulled, new_inserts, 'success')
                conn.commit()
                success = True

                # Settle attendance_daily for past 45 days (covers current + previous month)
                try:
                    from datetime import date as _sd, timedelta as _std
                    _to   = _sd.today().isoformat()
                    _from = (_sd.today() - _std(days=45)).isoformat()
                    _sr = db_mod.settle_all_attendance_daily(conn, _from, _to)
                    conn.commit()
                except Exception as _se:
                    conn.rollback()
                    import logging as _lg
                    _lg.getLogger(__name__).warning("attendance_daily settlement failed: %s", _se)

            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                cur.execute(
                    "SELECT completed_at FROM pull_sessions WHERE id = %s",
                    (session_id,),
                )
                row = cur.fetchone()
                completed_at = row["completed_at"] if row else None

        except Exception as exc:
            conn.rollback()
            error_message = str(exc)
            if session_id:
                try:
                    db_mod.complete_pull_session(conn, session_id, 0, 0, 'failed',
                                                 str(exc), _traceback.format_exc())
                    conn.commit()
                except Exception:
                    conn.rollback()

        return render(templates, request, "pull_result.html", {
            "device": d,
            "success": success,
            "session_id": session_id,
            "user_count": user_count,
            "records_pulled": records_pulled,
            "new_inserts": new_inserts,
            "error_message": error_message,
            "started_at": started_at,
            "completed_at": completed_at,
        })
    finally:
        conn.close()


@app.get("/devices/{device_id}/manual-pull")
def device_manual_pull_form(request: Request, device_id: int):
    conn = get_connection()
    try:
        d = get_device(conn, device_id)
        if not d:
            raise HTTPException(status_code=404, detail="Device not found")
    finally:
        conn.close()
    from datetime import date as _sd, timedelta as _std
    return render(templates, request, "device_manual_pull.html", {
        "device":   d,
        "sel_from": (_sd.today() - _std(days=7)).isoformat(),
        "sel_to":   _sd.today().isoformat(),
    })


@app.post("/devices/{device_id}/manual-pull")
def device_manual_pull(request: Request, device_id: int,
                        from_ad: str = Form(...), to_ad: str = Form(...)):
    conn = get_connection()
    try:
        d = get_device(conn, device_id)
        if not d:
            raise HTTPException(status_code=404, detail="Device not found")
        device_cfg = device_config_from_row(d)
        started_at = datetime.now(timezone.utc)
        session_id = None
        error_message = None
        success = False
        user_count = 0
        records_pulled = 0
        new_inserts = 0
        completed_at = None

        try:
            import zoneinfo
            npt_tz = zoneinfo.ZoneInfo('Asia/Kathmandu')
        except Exception:
            npt_tz = timezone.utc

        try:
            device_db_id = db_mod.upsert_device(conn, device_cfg)
            conn.commit()
            session_id = db_mod.start_pull_session(conn, device_db_id, started_at)
            conn.commit()
            result = puller_mod.pull_device(device_cfg)

            if not result.success:
                error_message = result.error
                db_mod.complete_pull_session(conn, session_id, 0, 0, 'failed',
                                             result.error, result.error_traceback)
                conn.commit()
            else:
                for user in result.users:
                    gu = db_mod.find_global_user_by_global_id(conn, str(user.user_id))
                    if gu:
                        try:
                            setattr(user, 'global_user_id', gu['id'])
                        except Exception:
                            pass
                    try:
                        db_mod.upsert_employee(conn, device_db_id, user)
                    except Exception:
                        conn.rollback()

                user_count = len(result.users)
                employee_map = db_mod.build_employee_map(conn, device_db_id)

                # Only keep records whose Nepal-time calendar date falls in the
                # requested range — same filtering approach as pull_month.py.
                all_records = [attendance_to_dict(a) for a in result.attendance]
                records = []
                for r in all_records:
                    ts = r.get('timestamp')
                    if ts is None:
                        continue
                    npt_date = ts.astimezone(npt_tz).date().isoformat()
                    if from_ad <= npt_date <= to_ad:
                        records.append(r)

                records_pulled = len(records)
                new_inserts = db_mod.insert_attendance_batch(conn, device_db_id, records, employee_map) if records else 0
                db_mod.complete_pull_session(conn, session_id, records_pulled, new_inserts, 'success')
                conn.commit()
                success = True

                try:
                    _sr = db_mod.settle_all_attendance_daily(conn, from_ad, to_ad)
                    conn.commit()
                except Exception as _se:
                    conn.rollback()
                    import logging as _lg
                    _lg.getLogger(__name__).warning("attendance_daily settlement failed: %s", _se)

            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                cur.execute(
                    "SELECT completed_at FROM pull_sessions WHERE id = %s",
                    (session_id,),
                )
                row = cur.fetchone()
                completed_at = row["completed_at"] if row else None

        except Exception as exc:
            conn.rollback()
            error_message = str(exc)
            if session_id:
                try:
                    db_mod.complete_pull_session(conn, session_id, 0, 0, 'failed',
                                                 str(exc), _traceback.format_exc())
                    conn.commit()
                except Exception:
                    conn.rollback()

        return render(templates, request, "pull_result.html", {
            "device": d,
            "success": success,
            "session_id": session_id,
            "user_count": user_count,
            "records_pulled": records_pulled,
            "new_inserts": new_inserts,
            "error_message": error_message,
            "started_at": started_at,
            "completed_at": completed_at,
            "manual_range": {"from_ad": from_ad, "to_ad": to_ad},
        })
    finally:
        conn.close()


@app.get("/devices/{device_id}/users")
def device_users_view(request: Request, device_id: int):
    conn = get_connection()
    try:
        d = get_device(conn, device_id)
        if not d:
            raise HTTPException(status_code=404, detail="Device not found")
        device_cfg = device_config_from_row(d)
        try:
            device_users = puller_mod.list_device_users(device_cfg)
            users = [
                {
                    "uid": getattr(u, "uid", ""),
                    "user_id": getattr(u, "user_id", ""),
                    "name": getattr(u, "name", ""),
                    "privilege": getattr(u, "privilege", ""),
                    "card": getattr(u, "card", ""),
                }
                for u in device_users
            ]
            error = None
        except Exception as exc:
            users = []
            msg = str(exc)
            if "timed out" in msg.lower() or "timeout" in msg.lower():
                msg = (f"Could not load device users: connection timed out. "
                       f"Check that the device is powered on and reachable on the network. "
                       f"(raw: {msg})")
            else:
                msg = f"Could not load device users: {msg}"
            error = msg
        return render(templates, request, "device_users.html", {
            "device": d,
            "users": users,
            "error": error,
        })
    finally:
        conn.close()


@app.post("/devices/{device_id}/import-users")
def import_device_users(
    request: Request,
    device_id: int,
    selected: list[str] | None = Form(None),
    all: str | None = Form(None),
):
    conn = get_connection()
    try:
        d = get_device(conn, device_id)
        if not d:
            raise HTTPException(status_code=404, detail="Device not found")
        device_cfg = device_config_from_row(d)
        selected_uids = {int(uid) for uid in selected} if selected else set()
        import_all = all == "1"
        results = []
        # Try to get fingerprint data if available
        fp_map = {}
        try:
            fp_map = puller_mod.get_fingerprint_map(device_cfg)
        except Exception:
            pass
        for user in puller_mod.list_device_users(device_cfg):
            uid = int(getattr(user, "uid", 0))
            if not import_all and uid not in selected_uids:
                continue
            user_id = str(getattr(user, "user_id", "") or "")
            name = str(getattr(user, "name", "") or "")
            privilege = int(getattr(user, "privilege", 0) or 0)
            card = str(getattr(user, "card", "") or "") or None
            if not user_id:
                results.append({"uid": uid, "name": name, "ok": False, "message": "Skipped blank user ID"})
                continue
            global_user = find_global_user_by_global_id(conn, user_id)
            if global_user:
                global_user_id = global_user["id"]
            else:
                fp_data = fp_map.get(uid)
                fp_json = json.dumps(fp_data) if fp_data else None
                global_user_id = create_global_user(
                    conn, user_id, name, privilege, card,
                    app_user_id=_current_user_id(request),
                    fingerprint_data=fp_json,
                )
                conn.commit()
            setattr(user, "global_user_id", global_user_id)
            upsert_employee(conn, device_id, user)
            conn.commit()
            results.append({"uid": uid, "name": name, "ok": True, "message": "Imported"})
        return render(templates, request, "import_result.html", {
            "device": d,
            "results": results,
        })
    finally:
        conn.close()


@app.get("/attendance")
def attendance_view(
    request: Request,
    from_date: str | None = None,
    to_date:   str | None = None,
    from_bs:   str | None = None,
    to_bs:     str | None = None,
    date_str:  str | None = None,   # legacy compat
    device_id: str | None = None,
    name:      str | None = None,
    page:      str | None = None,
    log_page:  str | None = None,
):
    import nepali_utils as _nu
    # BS date overrides
    if from_bs:
        _ad = _nu.bs_to_ad(from_bs)
        if _ad:
            from_date = _ad
    if to_bs:
        _ad = _nu.bs_to_ad(to_bs)
        if _ad:
            to_date = _ad
    # legacy single-date compat
    if date_str and not from_date:
        from_date = date_str
    if date_str and not to_date:
        to_date = date_str
    today = date.today().isoformat()
    from_date = from_date or today
    to_date   = to_date   or today
    if to_date < from_date:
        to_date = from_date

    per_page      = 100
    page_num      = max(1, _int_param(page)     or 1)
    log_page_num  = max(1, _int_param(log_page) or 1)
    device_id_int = _int_param(device_id)
    name_clean    = name.strip() if name else None

    conn = get_connection()
    try:
        devices = get_devices(conn)

        # ── Summary (all rows, then slice for pagination) ──
        summary_all  = db_mod.get_attendance_summary_filtered(
            conn, from_date, to_date, device_id_int, name_clean)
        total_summary = len(summary_all)
        total_pages   = max(1, (total_summary + per_page - 1) // per_page)
        page_num      = min(page_num, total_pages)
        summary       = summary_all[(page_num - 1) * per_page : page_num * per_page]

        # ── Raw punch log (server-side paginated) ──
        where:  list = ["DATE(al.timestamp AT TIME ZONE 'Asia/Kathmandu') BETWEEN %s AND %s"]
        params: list = [from_date, to_date]
        if device_id_int:
            where.append("al.device_id = %s")
            params.append(device_id_int)
        if name_clean:
            where.append(
                "(al.name ILIKE %s OR al.user_id ILIKE %s "
                "OR EXISTS (SELECT 1 FROM employees _e "
                "           WHERE _e.device_id = al.device_id AND _e.user_id = al.user_id "
                "           AND _e.name ILIKE %s))"
            )
            params += [f'%{name_clean}%', f'%{name_clean}%', f'%{name_clean}%']

        _lat = ("LEFT JOIN LATERAL (SELECT name FROM employees "
                "WHERE device_id = al.device_id AND user_id = al.user_id "
                "ORDER BY id LIMIT 1) e ON TRUE ")

        with conn.cursor() as cur:
            cur.execute(
                f"SELECT COUNT(*) FROM attendance_logs al "
                + _lat +
                f"JOIN devices d ON al.device_id = d.id "
                f"WHERE {' AND '.join(where)}", tuple(params))
            total_records = cur.fetchone()[0]

        total_log_pages = max(1, (total_records + per_page - 1) // per_page)
        log_page_num    = min(log_page_num, total_log_pages)
        log_offset      = (log_page_num - 1) * per_page

        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(
                f"SELECT al.timestamp, al.uid, al.user_id, "
                f"       COALESCE(al.name, e.name, 'Unknown') AS name, "
                f"       al.status, al.punch, al.punch_label, d.name AS device_name "
                f"FROM attendance_logs al "
                + _lat +
                f"JOIN devices d ON al.device_id = d.id "
                f"WHERE {' AND '.join(where)} "
                f"ORDER BY al.timestamp DESC LIMIT %s OFFSET %s",
                tuple(params) + (per_page, log_offset))
            records = [dict(r) for r in cur.fetchall()]

        filter_qs = urlencode({
            'from_date': from_date, 'to_date': to_date,
            'from_bs': from_bs or '', 'to_bs': to_bs or '',
            'device_id': device_id or '', 'name': name or '',
        })
        return render(templates, request, "attendance.html", {
            "records": records, "summary": summary,
            "devices": devices,
            "from_date": from_date, "to_date": to_date,
            "from_bs": from_bs or '', "to_bs": to_bs or '',
            "selected_device_id": device_id_int,
            "name_search": name or '',
            # pagination — summary
            "page": page_num, "total_pages": total_pages, "total_summary": total_summary,
            # pagination — log
            "log_page": log_page_num, "total_log_pages": total_log_pages, "total_records": total_records,
            "per_page": per_page,
            "filter_qs": filter_qs,
            # company
            "COMPANY_NAME": COMPANY_NAME, "COMPANY_ADDRESS": COMPANY_ADDRESS,
            "COMPANY_EMAIL": COMPANY_EMAIL, "COMPANY_WEBSITE": COMPANY_WEBSITE,
        })
    finally:
        conn.close()


@app.get("/attendance/export/excel")
def attendance_export_excel(
    from_date: str | None = None, to_date:   str | None = None,
    from_bs:   str | None = None, to_bs:     str | None = None,
    date_str:  str | None = None,
    device_id: str | None = None, name: str | None = None,
):
    import nepali_utils as _nu
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    if from_bs:
        _ad = _nu.bs_to_ad(from_bs)
        if _ad: from_date = _ad
    if to_bs:
        _ad = _nu.bs_to_ad(to_bs)
        if _ad: to_date = _ad
    if date_str and not from_date: from_date = date_str
    if date_str and not to_date:   to_date   = date_str
    today = date.today().isoformat()
    from_date = from_date or today
    to_date   = to_date   or today
    device_id_int = _int_param(device_id)
    name_clean = name.strip() if name else None

    conn = get_connection()
    try:
        summary = db_mod.get_attendance_summary_filtered(conn, from_date, to_date, device_id_int, name_clean)
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Summary"

    thin = Side(style='thin')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Company header rows
    header_font = Font(bold=True, size=14)
    sub_font    = Font(size=10)
    ws.merge_cells('A1:J1'); ws['A1'] = COMPANY_NAME
    ws['A1'].font = header_font; ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:J2'); ws['A2'] = COMPANY_ADDRESS
    ws['A2'].font = sub_font; ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A3:J3'); ws['A3'] = f"Email: {COMPANY_EMAIL}  |  Website: {COMPANY_WEBSITE}"
    ws['A3'].font = sub_font; ws['A3'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A4:J4'); ws['A4'] = (
        f"Attendance Report  |  {from_date} to {to_date}"
        f"  |  {_nu.bs_date_str(from_date)} BS  to  {_nu.bs_date_str(to_date)} BS")
    ws['A4'].font = Font(italic=True, size=10)
    ws['A4'].alignment = Alignment(horizontal='center')
    ws.append([])  # blank row 5

    col_headers = ['SN', 'Name', 'User ID',
                   'First Check-In (NPT)', 'First Check-In (BS)',
                   'Last Check-Out (NPT)', 'Last Check-Out (BS)',
                   'All Punch Times', 'Total Punches', 'Devices']
    ws.append(col_headers)
    hdr_fill = PatternFill("solid", fgColor="1769E0")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[6]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = bdr

    for i, s in enumerate(summary, 1):
        punch_str = "\n".join(
            f"{_nu.jinja_fmt_dt(p['ts'])} {p['label']}" for p in s.get('punches', []))
        row_data = [
            i,
            s['name'], str(s['user_id']),
            _nu.jinja_fmt_dt(s['first_in']),
            _nu.jinja_bs_datetime(s['first_in']),
            _nu.jinja_fmt_dt(s['last_out']),
            _nu.jinja_bs_datetime(s['last_out']),
            punch_str,
            s['total_punches'],
            s['devices'] or '',
        ]
        ws.append(row_data)
        row_idx = ws.max_row
        for cell in ws[row_idx]:
            cell.border = bdr
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        if i % 2 == 0:
            fill = PatternFill("solid", fgColor="EEF6FF")
            for cell in ws[row_idx]: cell.fill = fill

    col_widths = [6, 30, 12, 22, 28, 22, 28, 40, 10, 24]
    for idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    for r in range(1, 5):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[6].height = 20
    ws.freeze_panes = 'A7'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"attendance_{from_date}_to_{to_date}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@app.get("/attendance/export/pdf")
def attendance_export_pdf(
    from_date: str | None = None, to_date:   str | None = None,
    from_bs:   str | None = None, to_bs:     str | None = None,
    date_str:  str | None = None,
    device_id: str | None = None, name: str | None = None,
):
    import nepali_utils as _nu
    from fpdf import FPDF
    if from_bs:
        _ad = _nu.bs_to_ad(from_bs)
        if _ad: from_date = _ad
    if to_bs:
        _ad = _nu.bs_to_ad(to_bs)
        if _ad: to_date = _ad
    if date_str and not from_date: from_date = date_str
    if date_str and not to_date:   to_date   = date_str
    today = date.today().isoformat()
    from_date = from_date or today
    to_date   = to_date   or today
    device_id_int = _int_param(device_id)
    name_clean = name.strip() if name else None

    conn = get_connection()
    try:
        summary = db_mod.get_attendance_summary_filtered(conn, from_date, to_date, device_id_int, name_clean)
    finally:
        conn.close()

    def _pdf_safe(v):
        """Sanitize a value to Latin-1 safe string for fpdf Helvetica font."""
        if v is None:
            return '-'
        s = str(v)
        s = s.replace('—', '-').replace('–', '-').replace('…', '...') \
             .replace('’', "'").replace('‘', "'").replace('“', '"').replace('”', '"')
        return s.encode('latin-1', errors='replace').decode('latin-1')

    # Column widths (mm) for landscape A4 (277mm usable)
    COL_W = [10, 52, 18, 32, 40, 32, 40, 11, 42]  # total = 277
    HEADERS = ['SN', 'Name', 'User ID', 'First In (NPT)', 'First In (BS)',
               'Last Out (NPT)', 'Last Out (BS)', 'Punches', 'Devices']
    ROW_H = 7

    class AttPDF(FPDF):
        def header(self):
            self.set_font('Helvetica', 'B', 13)
            self.cell(0, 7, _pdf_safe(COMPANY_NAME), align='C', new_x='LMARGIN', new_y='NEXT')
            self.set_font('Helvetica', '', 9)
            self.cell(0, 5, _pdf_safe(COMPANY_ADDRESS), align='C', new_x='LMARGIN', new_y='NEXT')
            self.cell(0, 5, _pdf_safe(f"Email: {COMPANY_EMAIL}  |  Website: {COMPANY_WEBSITE}"),
                      align='C', new_x='LMARGIN', new_y='NEXT')
            self.set_draw_color(23, 105, 224)
            self.set_line_width(0.5)
            self.line(self.l_margin, self.get_y() + 1,
                      self.w - self.r_margin, self.get_y() + 1)
            self.ln(4)
            self.set_font('Helvetica', 'I', 8)
            bs_from = _nu.bs_date_str(from_date)
            bs_to   = _nu.bs_date_str(to_date)
            self.cell(0, 5,
                _pdf_safe(f"Attendance Report  |  {from_date}  to  {to_date}"
                f"   ({bs_from}  to  {bs_to})"),
                align='C', new_x='LMARGIN', new_y='NEXT')
            self.ln(2)
            # Table header
            self.set_fill_color(23, 105, 224)
            self.set_text_color(255, 255, 255)
            self.set_font('Helvetica', 'B', 7)
            for h, w in zip(HEADERS, COL_W):
                self.cell(w, ROW_H, h, border=1, align='C', fill=True)
            self.set_text_color(0, 0, 0)
            self.ln()

        def footer(self):
            self.set_y(-12)
            self.set_font('Helvetica', 'I', 7)
            self.cell(0, 5, f"Page {self.page_no()} — Generated by ZKTeco Attendance Console", align='C')

    pdf = AttPDF(orientation='L', unit='mm', format='A4')
    pdf.set_margins(10, 10, 10)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font('Helvetica', '', 7)

    for i, s in enumerate(summary, 1):
        if pdf.get_y() + ROW_H > pdf.page_break_trigger:
            pdf.add_page()
        fill = (i % 2 == 0)
        if fill:
            pdf.set_fill_color(238, 246, 255)
        def _t(v):
            txt = str(v or '')
            return txt[:40] + '…' if len(txt) > 40 else txt
        row_vals = [
            str(i),
            _pdf_safe(_t(s['name'])),
            _pdf_safe(s['user_id']),
            _pdf_safe(_nu.jinja_fmt_dt(s['first_in'])),
            _pdf_safe(_nu.jinja_bs_datetime(s['first_in'])),
            _pdf_safe(_nu.jinja_fmt_dt(s['last_out'])),
            _pdf_safe(_nu.jinja_bs_datetime(s['last_out'])),
            str(s['total_punches']),
            _pdf_safe(_t(s['devices'] or '')),
        ]
        for val, w in zip(row_vals, COL_W):
            pdf.cell(w, ROW_H, val, border=1, fill=fill)
        pdf.ln()

    buf = io.BytesIO(pdf.output())
    fname = f"attendance_{from_date}_to_{to_date}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@app.get("/pull-sessions")
def pull_sessions_view(request: Request, device_id: str | None = None, status: str | None = None, days: str | None = None):
    """Show pull sessions with optional filters: device_id, status, days (lookback)."""
    device_id_int = _int_param(device_id)
    days_int = _int_param(days) or 7
    conn = get_connection()
    try:
        sql = "SELECT ps.*, d.name as device_name FROM pull_sessions ps JOIN devices d ON ps.device_id = d.id"
        where = []
        params = []
        if device_id_int:
            where.append("ps.device_id = %s")
            params.append(device_id_int)
        if status:
            where.append("ps.status = %s")
            params.append(status)
        where.append("ps.started_at >= NOW() - (%s * INTERVAL '1 day')")
        params.append(days_int)
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY ps.started_at DESC LIMIT 1000"
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(sql, tuple(params))
            sessions = [dict(r) for r in cur.fetchall()]
        # load devices for filter dropdown
        devices = db_mod.get_devices(conn) if hasattr(db_mod, 'get_devices') else get_devices(conn)
        return render(templates, request, "pull_sessions.html", {
            "sessions": sessions,
            "devices": devices,
            "selected_device_id": device_id_int,
            "selected_status": status,
            "selected_days": days_int,
        })
    finally:
        conn.close()


@app.get('/bulk-enroll')
def bulk_enroll_form(request: Request):
    conn = get_connection()
    try:
        users = list_global_users(conn)
        devices = get_devices(conn)
        return render(templates, request, 'bulk_enroll.html', {'users': users, 'devices': devices})
    finally:
        conn.close()


@app.get('/devices/{device_id}/sync')
def device_sync_view(request: Request, device_id: int):
    """Show diff between DB and device users for a device."""
    conn = get_connection()
    try:
        d = get_device(conn, device_id)
        if not d:
            raise HTTPException(status_code=404, detail='Device not found')
        device_cfg = device_config_from_row(d)
        # fetch device users
        try:
            device_users = puller_mod.list_device_users(device_cfg)
        except Exception as exc:
            return render(templates, request, 'sync.html', {'device': d, 'error': str(exc)})

        # Map device users by user_id (global_user_id) and uid
        dev_by_userid = {str(getattr(u, 'user_id')): u for u in device_users}
        dev_by_uid = {int(getattr(u, 'uid')): u for u in device_users}

        # fetch DB employees for this device
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute('SELECT id, uid, user_id, name, privilege, card, global_user_id FROM employees WHERE device_id = %s', (device_id,))
            db_employees = [dict(r) for r in cur.fetchall()]

        db_by_userid = {str(r['user_id']): r for r in db_employees if r.get('user_id')}
        db_by_uid = {int(r['uid']): r for r in db_employees}

        users_only_on_device = [v for k, v in dev_by_userid.items() if k not in db_by_userid]
        users_only_in_db = [v for k, v in db_by_userid.items() if k not in dev_by_userid]

        return render(templates, request, 'sync.html', {
            'device': d,
            'device_users': device_users,
            'db_employees': db_employees,
            'only_on_device': users_only_on_device,
            'only_in_db': users_only_in_db,
        })
    finally:
        conn.close()


@app.post('/devices/{device_id}/sync')
def device_sync_action(request: Request, device_id: int, action: str = Form(...), selected: list[str] | None = Form(None)):
    """Resolve diffs: action='push_missing' pushes DB users missing on device; action='import_unknown' imports device-only users to global_users and employees."""
    conn = get_connection()
    try:
        d = get_device(conn, device_id)
        if not d:
            raise HTTPException(status_code=404, detail='Device not found')
        device_cfg = device_config_from_row(d)

        ids = [s.strip() for s in selected] if selected else []
        results = []

        if action == 'push_missing':
            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                if ids:
                    cur.execute(
                        'SELECT global_user_id, name, privilege, card FROM global_users WHERE id = ANY(%s)',
                        ([int(x) for x in ids],),
                    )
                    rows = [dict(r) for r in cur.fetchall()]
                else:
                    rows = []
            for gu in rows:
                try:
                    r = puller_mod.push_global_user_to_device(device_cfg, gu)
                    results.append({'global_user_id': gu.get('global_user_id'), 'result': r})
                except Exception as exc:
                    results.append({'global_user_id': gu.get('global_user_id'), 'result': {'ok': False, 'message': str(exc)}})

        elif action == 'import_unknown':
            # import device-only users to global_users and employees
            try:
                device_users = puller_mod.list_device_users(device_cfg)
            except Exception as exc:
                return render(templates, request, 'sync_result.html', {
                    'device': d,
                    'results': [{'ok': False, 'message': str(exc)}],
                    'action': action,
                    'action_label': action_label(action),
                    'ok_count': 0,
                    'fail_count': 1,
                })
            device_users_filter = device_users
            if ids:
                selected_uids = set(int(x) for x in ids)
                device_users_filter = [u for u in device_users if int(getattr(u, 'uid')) in selected_uids]
            for u in device_users_filter:
                uid = int(getattr(u, 'uid'))
                user_id = str(getattr(u, 'user_id') or '')
                if user_id == '':
                    # skip blank ids
                    continue
                # if already in global_users skip
                with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                    cur.execute('SELECT id FROM global_users WHERE global_user_id = %s', (user_id,))
                    row = cur.fetchone()
                    if row:
                        gu_id = row['id']
                    else:
                        gu_id = create_global_user(conn, user_id, getattr(u, 'name') or '', int(getattr(u, 'privilege') or 0), getattr(u, 'card') or None)
                        conn.commit()
                # upsert employee for this device
                try:
                    # build a simple user-like dict
                    user_obj = {'uid': uid, 'user_id': user_id, 'name': getattr(u, 'name') or '', 'privilege': int(getattr(u, 'privilege') or 0), 'card': getattr(u, 'card') or None, 'global_user_id': gu_id}
                    db_mod.upsert_employee(conn, device_id, user_obj)
                    conn.commit()
                    results.append({'uid': uid, 'global_user_id': user_id, 'ok': True})
                except Exception as exc:
                    conn.rollback()
                    results.append({'uid': uid, 'global_user_id': user_id, 'ok': False, 'error': str(exc)})

        else:
            return render(templates, request, 'sync_result.html', {
                'device': d,
                'results': [{'ok': False, 'message': 'Unknown action'}],
                'action': action,
                'action_label': action_label(action),
                'ok_count': 0,
                'fail_count': 1,
            })

        ok_count = sum(1 for r in results if (r.get('result') or {}).get('ok') or r.get('ok'))
        fail_count = len(results) - ok_count
        return render(templates, request, 'sync_result.html', {
            'device': d,
            'results': results,
            'action': action,
            'action_label': action_label(action),
            'ok_count': ok_count,
            'fail_count': fail_count,
        })
    finally:
        conn.close()


@app.post('/bulk-enroll')
def bulk_enroll(request: Request, device_id: int = Form(...), user_ids: str = Form(None)):
    """Enroll multiple global users (CSV of ids) to a single device."""
    conn = get_connection()
    try:
        devices = get_devices(conn)
        target = None
        for d in devices:
            if d['id'] == device_id:
                target = d
                break
        if not target:
            raise HTTPException(status_code=404, detail='Device not found')
        # resolve user list
        ids = [int(x) for x in user_ids.split(',') if x.strip()] if user_ids else []
        global_users = []
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            if ids:
                cur.execute('SELECT global_user_id, name, privilege, card FROM global_users WHERE id = ANY(%s)', (ids,))
            else:
                cur.execute('SELECT global_user_id, name, privilege, card FROM global_users')
            for row in cur.fetchall():
                global_users.append(dict(row))

        device_cfg = device_config_from_row(target)
        summary = puller_mod.push_bulk_users_to_device(device_cfg, global_users)
        return render(templates, request, 'bulk_result.html', {'summary': summary, 'device': target})
    finally:
        conn.close()


# ---- Nepali date conversion API -----------------------------------------


@app.get("/api/bs-to-ad")
def api_bs_to_ad(date: str = ""):
    """Convert BS date 'YYYY-MM-DD' → AD ISO date."""
    from nepali_utils import bs_to_ad
    ad = bs_to_ad(date) if date else None
    return {"bs": date, "ad": ad}


@app.get("/api/ad-to-bs")
def api_ad_to_bs(date: str = ""):
    """Convert AD date 'YYYY-MM-DD' → BS ISO date."""
    from nepali_utils import ad_to_bs
    bs = ad_to_bs(date) if date else None
    return {"ad": date, "bs": bs}


@app.get("/api/bs-month-info")
def api_bs_month_info(year: int = 2082, month: int = 1):
    """Return days count + first weekday (0=Sun) for a BS month."""
    from nepali_utils import bs_month_info
    info = bs_month_info(year, month)
    if info is None:
        raise HTTPException(status_code=400, detail="Invalid BS year/month")
    return info


# ---- Monthly attendance report ------------------------------------------


def _fmt_min(minutes: int | None) -> str:
    if minutes is None or minutes <= 0:
        return ''
    h, m = divmod(int(minutes), 60)
    return f"{h:02d}:{m:02d}"


def _time_to_min(t) -> int | None:
    """Convert a datetime.time or HH:MM string to minutes since midnight."""
    if t is None:
        return None
    try:
        if hasattr(t, 'hour'):
            return t.hour * 60 + t.minute
        parts = str(t).split(':')
        return int(parts[0]) * 60 + int(parts[1])
    except Exception:
        return None


def _compute_monthly_report(daily_rows: list, from_ad: str, to_ad: str,
                              default_si_min: int = 600, default_so_min: int = 1020,
                              shift_calendar: dict = None,
                              holiday_map: dict = None,
                              leave_map: set = None) -> list:
    """Build per-day dicts matching the 16-column ZKBioTime periodic attendance format.

    Columns: Work Date | Planned In | Planned Out | Work Time |
             Time In | Time Out | Break In | Break Out | Time |
             Actual | OT | LateIn | EarlyOut | EarlyIn | LateOut | Remark
    """
    from datetime import date, timedelta
    from nepali_utils import ad_to_bs_tuple

    NEPAL_DAYS = ['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday']

    punch_map = {r['work_date']: r for r in daily_rows}
    start = date.fromisoformat(from_ad)
    end   = date.fromisoformat(to_ad)

    def _pstr(pd) -> str:
        if not pd:
            return ''
        t = pd.get('time')
        return t.strftime('%H:%M') if hasattr(t, 'strftime') else str(t)[:5]

    days = []
    d = start
    while d <= end:
        bs_t = ad_to_bs_tuple(d)
        bs_str = f"{bs_t[2]:02d}/{bs_t[1]:02d}/{bs_t[0]}" if bs_t else ''
        nepal_dow = d.isoweekday() % 7
        day_name  = NEPAL_DAYS[nepal_dow]
        is_weekend = (nepal_dow == 6)

        # Per-day shift lookup (employee-specific or department-level via shift_calendar)
        shift_info   = (shift_calendar or {}).get(d)
        si_min       = shift_info['start_min'] if shift_info else default_si_min
        so_min       = shift_info['end_min']   if shift_info else default_so_min
        shift_name   = shift_info['name']      if shift_info else ''
        planned_work = so_min - si_min

        holiday_entry = (holiday_map or {}).get(d.isoformat())
        is_holiday    = bool(holiday_entry)
        holiday_type  = (holiday_entry or {}).get('holiday_type', 'public') if is_holiday else ''
        is_off_day    = is_weekend or is_holiday

        row = punch_map.get(d)

        pts = []
        if row:
            if row.get('all_punches_with_device'):
                pts = [p for p in row['all_punches_with_device'] if p]
            elif row.get('all_punch_times'):
                pts = [{'time': p, 'device_name': ''} for p in row['all_punch_times'] if p]

        first_punch = row['first_punch'] if row else None
        last_punch  = row['last_punch']  if row else None

        # 1 punch: Time In only
        # 2 punches: Time In + Break Out
        # 3 punches: Time In + Time Out + Break Out
        # 4+:        Time In + Time Out + Break In + Break Out
        time_in   = _pstr(pts[0])  if len(pts) >= 1 else ''
        time_out  = _pstr(pts[1])  if len(pts) >= 3 else ''
        break_in  = _pstr(pts[2])  if len(pts) >= 4 else ''
        break_out = _pstr(pts[-1]) if len(pts) >= 2 else ''

        ci_min = _time_to_min(first_punch)
        co_min = _time_to_min(last_punch) if (last_punch and first_punch and last_punch != first_punch) else None

        work_min = (co_min - ci_min) if (ci_min is not None and co_min is not None and co_min > ci_min) else None
        if work_min is not None:
            time_col = _fmt_min(work_min)
        elif len(pts) == 1:
            time_col = _pstr(pts[0])
        else:
            time_col = ''

        late_in = early_out = early_in = late_out = ot = ''
        if not is_off_day and ci_min is not None:
            if ci_min > si_min:
                late_in  = _fmt_min(ci_min - si_min)
            elif ci_min < si_min:
                early_in = _fmt_min(si_min - ci_min)
        if not is_off_day and co_min is not None:
            if co_min < so_min:
                early_out = _fmt_min(so_min - co_min)
            elif co_min > so_min:
                late_out  = _fmt_min(co_min - so_min)
        if not is_off_day and work_min and work_min > planned_work:
            ot = _fmt_min(work_min - planned_work)

        is_on_leave = (not row) and (d.isoformat() in (leave_map or set()))

        if is_weekend:
            remark = 'Weekend'
        elif is_holiday:
            remark = 'Festival' if holiday_type == 'festival' else 'Holiday'
        elif row:
            remark = 'Present'
        elif is_on_leave:
            remark = 'Leave'
        else:
            remark = 'Absent'

        days.append({
            'bs_date':      bs_str,
            'ad_date':      d.isoformat(),
            'day_name':     day_name,
            'shift_name':   shift_name,
            'planned_in':   _fmt_min(si_min)       if not is_off_day else '00:00',
            'planned_out':  _fmt_min(so_min)        if not is_off_day else '00:00',
            'planned_work': _fmt_min(planned_work)  if not is_off_day else '',
            'planned_min':  planned_work             if not is_off_day else 0,
            'time_in':      time_in,
            'time_out':     time_out,
            'break_in':     break_in,
            'break_out':    break_out,
            'time_col':     time_col,
            'actual':       time_col,
            'ot':           ot,
            'late_in':      late_in,
            'early_out':    early_out,
            'early_in':     early_in,
            'late_out':     late_out,
            'remark':       remark,
            'work_min':     work_min or 0,
        })
        d += timedelta(days=1)

    return days


def _monthly_totals(days: list, planned_work: int = 0) -> dict:
    tot_actual = tot_ot = tot_late_in = tot_early_out = tot_early_in = tot_late_out = 0
    tot_planned = 0

    counts = {'Present': 0, 'Absent': 0, 'Weekend': 0, 'Holiday': 0, 'Festival': 0, 'Leave': 0, 'Misc': 0}
    for d in days:
        tot_actual    += d.get('work_min', 0)
        # Sum per-day planned for all workdays (not off-days or leaves)
        if d['remark'] not in ('Weekend', 'Holiday', 'Festival', 'Leave'):
            tot_planned += d.get('planned_min', 0)
        def _parse(s):
            if not s: return 0
            try:
                p = str(s).split(':')
                return int(p[0]) * 60 + int(p[1])
            except Exception:
                return 0
        tot_ot        += _parse(d['ot'])
        tot_late_in   += _parse(d['late_in'])
        tot_early_out += _parse(d['early_out'])
        tot_early_in  += _parse(d['early_in'])
        tot_late_out  += _parse(d['late_out'])
        r = d['remark']
        if r in counts:
            counts[r] += 1

    working_days = len(days) - counts['Weekend'] - counts['Holiday'] - counts['Festival']
    return {
        'planned':      _fmt_min(tot_planned),
        'actual':       _fmt_min(tot_actual),
        'ot':           _fmt_min(tot_ot),
        'late_in':      _fmt_min(tot_late_in),
        'early_out':    _fmt_min(tot_early_out),
        'early_in':     _fmt_min(tot_early_in),
        'late_out':     _fmt_min(tot_late_out),
        'counts':       counts,
        'working_days': working_days,
        'total_days':   len(days),
    }


def _bs_defaults():
    try:
        import nepali_datetime
        t = nepali_datetime.date.today()
        return t.year, t.month
    except Exception:
        return 2082, 1


def _npt_now_str():
    import zoneinfo as _zi
    return datetime.now(_zi.ZoneInfo('Asia/Kathmandu')).strftime('%Y-%m-%d %H:%M') + ' NPT'


def _month_name(m: int) -> str:
    names = ['','Baisakh','Jestha','Ashadh','Shrawan','Bhadra','Ashwin',
             'Kartik','Mangsir','Poush','Magh','Falgun','Chaitra']
    return names[m] if 1 <= m <= 12 else str(m)


def _page_range(page: int, total_pages: int) -> list:
    """Return page number list with None for ellipsis gaps, for pagination UI."""
    if total_pages <= 9:
        return list(range(1, total_pages + 1))
    pages = set()
    pages.update([1, 2])
    pages.update(range(max(1, page - 2), min(total_pages + 1, page + 3)))
    pages.update([total_pages - 1, total_pages])
    sorted_p = sorted(pages)
    result, prev = [], None
    for p in sorted_p:
        if prev is not None and p - prev > 1:
            result.append(None)
        result.append(p)
        prev = p
    return result


@app.get("/reports/monthly")
def reports_monthly_list(
    request: Request,
    bs_year:        str | None = None,
    bs_month:       str | None = None,
    page:           str | None = None,
    search:         str | None = None,
    department_id:  str | None = None,
    section_id:     str | None = None,
    unit_id:        str | None = None,
    directorate_id: str | None = None,
):
    def_year, def_month = _bs_defaults()
    sel_year  = _int_param(bs_year)  or def_year
    sel_month = _int_param(bs_month) or def_month
    page_num  = max(1, _int_param(page) or 1)
    per_page  = 25

    f_dept  = _int_param(department_id)
    f_sec   = _int_param(section_id)
    f_unit  = _int_param(unit_id)
    f_dir   = _int_param(directorate_id)
    f_srch  = (search or '').strip().lower()

    conn = get_connection()
    try:
        from db import (get_employees_for_report as _gef,
                        get_all_departments, get_all_directorates,
                        get_all_sections, get_all_units, get_global_user_count)
        all_emps        = _gef(conn)
        all_depts       = get_all_departments(conn)
        all_dirs        = get_all_directorates(conn)
        all_sections    = get_all_sections(conn)
        all_units_l     = get_all_units(conn)
        # Same canonical total as the dashboard and /users (Global Users) —
        # keeps "total employees" consistent everywhere it's shown.
        global_user_total = get_global_user_count(conn)
    finally:
        conn.close()

    # Apply filters
    filtered = all_emps
    if f_dir:
        filtered = [e for e in filtered if e.get('directorate_id') == f_dir]
    if f_dept:
        filtered = [e for e in filtered if e.get('department_id') == f_dept]
    if f_sec:
        filtered = [e for e in filtered if e.get('section_id') == f_sec]
    if f_unit:
        filtered = [e for e in filtered if e.get('unit_id') == f_unit]
    if f_srch:
        filtered = [e for e in filtered
                    if f_srch in (e.get('display_name') or '').lower()
                    or f_srch in (e.get('company_id') or '').lower()
                    or f_srch in (e.get('department_name') or '').lower()
                    or f_srch in (e.get('section_name') or '').lower()]

    # Sort by company user ID (numeric where possible) then name
    def _emp_sort_key(e):
        cid = (e.get('company_id') or '').strip()
        try:
            return (0, int(cid), (e.get('display_name') or '').lower())
        except ValueError:
            return (1, 0, cid.lower() or (e.get('display_name') or '').lower())

    filtered    = sorted(filtered, key=_emp_sort_key)
    total       = len(filtered)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page_num    = min(page_num, total_pages)
    page_emps   = filtered[(page_num - 1) * per_page: page_num * per_page]

    return render(templates, request, 'reports_monthly.html', {
        'view':            'list',
        'page_emps':       page_emps,
        'total_employees': total,
        'total_unfiltered': global_user_total,
        'page':            page_num,
        'total_pages':     total_pages,
        'sel_bs_year':     sel_year,
        'sel_bs_month':    sel_month,
        'def_year':        def_year,
        'def_month':       def_month,
        'all_depts':       all_depts,
        'all_dirs':        all_dirs,
        'all_sections':    all_sections,
        'all_units':       all_units_l,
        'f_dept':          f_dept or '',
        'f_dir':           f_dir or '',
        'f_sec':           f_sec or '',
        'f_unit':          f_unit or '',
        'f_srch':          search or '',
        'report':          None,
        'error':           None,
        'now_str':         _npt_now_str(),
        'COMPANY_NAME':    COMPANY_NAME,
        'page_range':      _page_range(page_num, total_pages),
    })


@app.get("/reports/monthly/view")
def reports_monthly_view(
    request: Request,
    emp_key:   str | None = None,
    global_id: str | None = None,   # backward-compat
    bs_year:   str | None = None,
    bs_month:  str | None = None,
):
    def_year, def_month = _bs_defaults()
    bs_y = _int_param(bs_year)  or def_year
    bs_m = _int_param(bs_month) or def_month
    SI_MIN, SO_MIN = 600, 1020   # default 10:00–17:00 if no shift rule found

    conn = get_connection()
    try:
        from db import get_employees_for_report as _gef
        all_emps = _gef(conn)

        # Resolve emp_key: prefer explicit emp_key, fall back to global_id compat
        resolved_key = emp_key
        if not resolved_key and global_id:
            gid_int = _int_param(global_id)
            resolved_key = f"g{gid_int}" if gid_int else None

        emp_entry = None
        if resolved_key:
            for e in all_emps:
                if e['key'] == resolved_key:
                    emp_entry = e
                    break
        if emp_entry is None and all_emps:
            emp_entry    = all_emps[0]
            resolved_key = emp_entry['key']

        g_id = emp_entry['global_id'] if emp_entry else None

        report = None
        error  = None
        if emp_entry:
            from nepali_utils import bs_month_info as _bsmi
            mi = _bsmi(bs_y, bs_m)
            if mi is None:
                error = "Invalid BS year/month"
            else:
                from_ad = mi['first_ad']
                to_ad   = mi['last_ad']
                pairs   = [(dv['device_id'], dv['user_id']) for dv in emp_entry['devices']]

                from db import get_employee_daily_attendance_multi as _multi
                from db import get_shift_calendar as _gsc
                from db import get_holidays as _ghols
                from db import get_leave_applications as _gleaveapps
                from datetime import date as _ddate, timedelta as _dtd
                daily       = _multi(conn, pairs, from_ad, to_ad)
                shift_cal   = _gsc(conn, g_id, from_ad, to_ad)
                holiday_map = {
                    (h['holiday_ad'].isoformat() if hasattr(h['holiday_ad'], 'isoformat') else str(h['holiday_ad'])): h
                    for h in _ghols(conn, from_ad, to_ad)
                }
                leave_dates: set = set()
                for la in _gleaveapps(conn, global_user_id=g_id, status='approved',
                                      from_ad=from_ad, to_ad=to_ad):
                    ld = la['from_ad'] if isinstance(la['from_ad'], _ddate) else _ddate.fromisoformat(str(la['from_ad']))
                    le = la['to_ad']   if isinstance(la['to_ad'],   _ddate) else _ddate.fromisoformat(str(la['to_ad']))
                    while ld <= le:
                        leave_dates.add(ld.isoformat())
                        ld += _dtd(days=1)
                days   = _compute_monthly_report(daily, from_ad, to_ad, SI_MIN, SO_MIN,
                                                 shift_cal, holiday_map, leave_dates)
                totals = _monthly_totals(days)

                report = {
                    'days':         days,
                    'totals':       totals,
                    'emp_name':     emp_entry['display_name'],
                    'emp_user_id':  emp_entry['company_id'] or (
                                    emp_entry['devices'][0]['user_id'] if emp_entry['devices'] else ''),
                    'device_name':  ', '.join(dv['device_name'] for dv in emp_entry['devices']),
                    'department':   emp_entry.get('department_name', ''),
                    'directorate':  emp_entry.get('directorate_name', ''),
                    'section':      emp_entry.get('section_name', ''),
                    'unit':         emp_entry.get('unit_name', ''),
                    'bs_year':      bs_y,
                    'bs_month':     bs_m,
                    'month_name':   mi['month_name'],
                    'from_ad':      from_ad,
                    'to_ad':        to_ad,
                    'from_bs_disp': nepali_utils.bs_date_str(from_ad, fmt='long'),
                    'to_bs_disp':   nepali_utils.bs_date_str(to_ad,   fmt='long'),
                    'global_id':    g_id,
                }
        else:
            error = "No employees found. Pull data from devices first."

    except Exception as exc:
        error    = str(exc)
        report   = None
        all_emps = []
        resolved_key = None
    finally:
        conn.close()

    return render(templates, request, 'reports_monthly.html', {
        'view':             'report',
        'all_emps':         all_emps,
        'sel_bs_year':      bs_y,
        'sel_bs_month':     bs_m,
        'def_year':         def_year,
        'def_month':        def_month,
        'sel_emp_key':      resolved_key,
        'sel_emp_display':  emp_entry['display_name'] if emp_entry else None,
        'report':           report,
        'error':            error,
        'now_str':          _npt_now_str(),
        'COMPANY_NAME':     COMPANY_NAME,
    })


# ---- Employee self-service: own attendance + leaves ----------------------


def _session_global_user_id(request: Request) -> int | None:
    """The global_users.id linked to the logged-in web_user, if any."""
    return request.session.get("global_user_id")


@app.get("/my-attendance")
def my_attendance(request: Request, bs_year: str | None = None, bs_month: str | None = None):
    g_id = _session_global_user_id(request)
    def_year, def_month = _bs_defaults()
    bs_y = _int_param(bs_year)  or def_year
    bs_m = _int_param(bs_month) or def_month
    SI_MIN, SO_MIN = 600, 1020

    if not g_id:
        return render(templates, request, 'my_attendance.html', {
            'report': None, 'error': 'Your login is not linked to an employee record. Contact an administrator.',
            'sel_bs_year': bs_y, 'sel_bs_month': bs_m,
            'def_year': def_year, 'def_month': def_month,
            'COMPANY_NAME': COMPANY_NAME,
        })

    conn = get_connection()
    try:
        gu = get_global_user(conn, g_id)
        with conn.cursor() as cur:
            cur.execute("SELECT device_id, user_id FROM employees WHERE global_user_id = %s", (g_id,))
            pairs = [(row[0], row[1]) for row in cur.fetchall()]

        from nepali_utils import bs_month_info as _bsmi
        mi = _bsmi(bs_y, bs_m)
        error = None
        report = None
        if mi is None:
            error = "Invalid BS year/month"
        else:
            from_ad = mi['first_ad']
            to_ad   = mi['last_ad']
            from db import get_employee_daily_attendance_multi as _multi
            from db import get_shift_calendar as _gsc
            from db import get_holidays as _ghols
            from db import get_leave_applications as _gleaveapps
            from datetime import date as _ddate, timedelta as _dtd
            daily       = _multi(conn, pairs, from_ad, to_ad) if pairs else []
            shift_cal   = _gsc(conn, g_id, from_ad, to_ad)
            holiday_map = {
                (h['holiday_ad'].isoformat() if hasattr(h['holiday_ad'], 'isoformat') else str(h['holiday_ad'])): h
                for h in _ghols(conn, from_ad, to_ad)
            }
            leave_dates: set = set()
            for la in _gleaveapps(conn, global_user_id=g_id, status='approved',
                                  from_ad=from_ad, to_ad=to_ad):
                ld = la['from_ad'] if isinstance(la['from_ad'], _ddate) else _ddate.fromisoformat(str(la['from_ad']))
                le = la['to_ad']   if isinstance(la['to_ad'],   _ddate) else _ddate.fromisoformat(str(la['to_ad']))
                while ld <= le:
                    leave_dates.add(ld.isoformat())
                    ld += _dtd(days=1)
            days   = _compute_monthly_report(daily, from_ad, to_ad, SI_MIN, SO_MIN,
                                             shift_cal, holiday_map, leave_dates)
            totals = _monthly_totals(days)
            device_name = ''
            if pairs:
                with conn.cursor() as cur:
                    cur.execute("SELECT name FROM devices WHERE id = ANY(%s)", ([p[0] for p in pairs],))
                    device_name = ', '.join(r[0] for r in cur.fetchall())
            report = {
                'days':         days,
                'totals':       totals,
                'emp_name':     gu.get('name') if gu else '',
                'emp_user_id':  gu.get('global_user_id') if gu else '',
                'device_name':  device_name,
                'department':   gu.get('department_name') if gu else '',
                'directorate':  gu.get('directorate_name') if gu else '',
                'section':      gu.get('section_name') if gu else '',
                'unit':         gu.get('unit_name') if gu else '',
                'bs_year':      bs_y, 'bs_month': bs_m, 'month_name': mi['month_name'],
                'from_ad':      from_ad, 'to_ad': to_ad,
                'global_id':    g_id,
            }
    finally:
        conn.close()

    return render(templates, request, 'my_attendance.html', {
        'report': report, 'error': error,
        'sel_bs_year': bs_y, 'sel_bs_month': bs_m,
        'def_year': def_year, 'def_month': def_month,
        'now_str': _npt_now_str(),
        'COMPANY_NAME': COMPANY_NAME,
    })


@app.get("/my-leaves")
def my_leaves(request: Request):
    g_id = _session_global_user_id(request)
    if not g_id:
        return render(templates, request, 'my_leaves.html', {
            'error': 'Your login is not linked to an employee record. Contact an administrator.',
            'applications': [], 'balances': [], 'leave_types': [],
            'COMPANY_NAME': COMPANY_NAME,
        })
    conn = get_connection()
    try:
        from db import get_employee_leave_balance, get_leave_applications, get_all_leave_types
        bs_year = _bs_defaults()[0]
        applications = get_leave_applications(conn, global_user_id=g_id, limit=100)
        balances     = get_employee_leave_balance(conn, g_id, bs_year)
        leave_types  = get_all_leave_types(conn)
    finally:
        conn.close()
    return render(templates, request, 'my_leaves.html', {
        'error': None,
        'applications': applications,
        'balances': balances,
        'leave_types': leave_types,
        'bs_year': bs_year,
        'COMPANY_NAME': COMPANY_NAME,
    })


@app.post("/my-leaves/apply")
async def my_leaves_apply(request: Request):
    from db import create_leave_application, get_holidays, count_leave_working_days
    from nepali_utils import bs_to_ad

    g_id = _session_global_user_id(request)
    if not g_id:
        return redirect_with_flash('/my-leaves', 'error', 'Your login is not linked to an employee record.')

    form = await request.form()
    leave_type_id = _int_param(form.get('leave_type_id'))
    from_bs = (form.get('from_bs') or '').strip()
    to_bs   = (form.get('to_bs') or '').strip()
    reason  = (form.get('reason') or '').strip()

    if not leave_type_id or not from_bs or not to_bs:
        return redirect_with_flash('/my-leaves', 'error', 'Leave type and dates are required.')

    from_ad = bs_to_ad(from_bs)
    to_ad   = bs_to_ad(to_bs)
    if not from_ad or not to_ad:
        return redirect_with_flash('/my-leaves', 'error', 'Invalid BS dates.')
    if from_ad > to_ad:
        return redirect_with_flash('/my-leaves', 'error', 'From date must be before to date.')

    conn = get_connection()
    try:
        hols = get_holidays(conn, from_ad, to_ad)
        days_val = count_leave_working_days(from_ad, to_ad, hols)
        if days_val <= 0:
            return redirect_with_flash('/my-leaves', 'error',
                                       'No working days in selected range (check for holidays/weekend).')
        # Self-service applications always start pending — employees cannot self-approve.
        create_leave_application(conn, g_id, leave_type_id,
                                  from_bs, to_bs, from_ad, to_ad,
                                  days_val, reason, _today_bs(), 'pending',
                                  app_user_id=_current_user_id(request))
    except Exception as exc:
        return redirect_with_flash('/my-leaves', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/my-leaves', 'success',
                               f'Leave application submitted ({days_val} days) — pending approval.')


@app.get("/reports/monthly/print-all")
def reports_monthly_print_all(
    request: Request,
    bs_year:  str | None = None,
    bs_month: str | None = None,
):
    def_year, def_month = _bs_defaults()
    bs_y = _int_param(bs_year)  or def_year
    bs_m = _int_param(bs_month) or def_month
    SI_MIN, SO_MIN = 600, 1020

    conn = get_connection()
    try:
        from db import get_employees_for_report as _gef
        from nepali_utils import bs_month_info as _bsmi
        from db import get_employee_daily_attendance_multi as _multi
        all_emps = _gef(conn)
        mi = _bsmi(bs_y, bs_m)
        reports = []
        if mi:
            from_ad, to_ad = mi['first_ad'], mi['last_ad']
            from db import get_shift_calendar as _gsc
            from db import get_holidays as _ghols
            from db import get_leave_applications as _gleaveapps
            from datetime import date as _ddate, timedelta as _dtd
            holiday_map = {
                (h['holiday_ad'].isoformat() if hasattr(h['holiday_ad'], 'isoformat') else str(h['holiday_ad'])): h
                for h in _ghols(conn, from_ad, to_ad)
            }
            # Build per-employee leave date sets for the month
            all_la = _gleaveapps(conn, status='approved', from_ad=from_ad, to_ad=to_ad, limit=50000)
            _leave_by_emp: dict = {}
            for la in all_la:
                gid = la['global_user_id']
                ld  = la['from_ad'] if isinstance(la['from_ad'], _ddate) else _ddate.fromisoformat(str(la['from_ad']))
                le  = la['to_ad']   if isinstance(la['to_ad'],   _ddate) else _ddate.fromisoformat(str(la['to_ad']))
                if gid not in _leave_by_emp:
                    _leave_by_emp[gid] = set()
                while ld <= le:
                    _leave_by_emp[gid].add(ld.isoformat())
                    ld += _dtd(days=1)
            for emp in all_emps:
                pairs     = [(dv['device_id'], dv['user_id']) for dv in emp['devices']]
                daily     = _multi(conn, pairs, from_ad, to_ad)
                shift_cal = _gsc(conn, emp.get('global_id'), from_ad, to_ad)
                leave_map = _leave_by_emp.get(emp.get('global_id'), set())
                days      = _compute_monthly_report(daily, from_ad, to_ad, SI_MIN, SO_MIN,
                                                    shift_cal, holiday_map, leave_map)
                totals    = _monthly_totals(days)
                reports.append({
                    'emp_name':    emp['display_name'],
                    'emp_user_id': emp['company_id'] or (emp['devices'][0]['user_id'] if emp['devices'] else ''),
                    'device_name': ', '.join(dv['device_name'] for dv in emp['devices']),
                    'days':        days,
                    'totals':      totals,
                    'month_name':  mi['month_name'],
                    'bs_year':     bs_y,
                    'from_bs_disp': nepali_utils.bs_date_str(from_ad, fmt='long'),
                    'to_bs_disp':   nepali_utils.bs_date_str(to_ad,   fmt='long'),
                })
    finally:
        conn.close()

    return render(templates, request, 'reports_monthly_print_all.html', {
        'reports':      reports,
        'bs_year':      bs_y,
        'bs_month':     bs_m,
        'month_name':   mi['month_name'] if mi else '',
        'now_str':      _npt_now_str(),
        'COMPANY_NAME': COMPANY_NAME,
    })


@app.get("/reports/hajiri")
def reports_hajiri(
    request:       Request,
    bs_year:       str | None = None,
    bs_month:      str | None = None,
    department_id: str | None = None,
    section_id:    str | None = None,
    emp_type:      str | None = None,
    search:        str | None = None,
):
    def_year, def_month = _bs_defaults()
    sel_year   = _int_param(bs_year)  or def_year
    sel_month  = _int_param(bs_month) or def_month
    f_dept     = _int_param(department_id)
    f_sec      = _int_param(section_id)
    f_emp_type = (emp_type or '').strip()
    f_search   = (search  or '').strip()

    conn = get_connection()
    try:
        from nepali_utils import bs_month_info as _bsmi
        from db import get_all_departments, get_all_sections

        depts    = get_all_departments(conn)
        sections = get_all_sections(conn)
        mi       = _bsmi(sel_year, sel_month)

        if mi is None:
            return render(templates, request, 'reports_hajiri.html', {
                'error': 'Invalid BS year/month',
                'sel_bs_year': sel_year, 'sel_bs_month': sel_month,
                'departments': depts, 'sections': sections,
                'COMPANY_NAME': COMPANY_NAME,
            })

        from_ad = mi['first_ad']
        to_ad   = mi['last_ad']
        from datetime import date as _dt, timedelta as _td

        # ── Load all active global users (with filters) ───────────────────────
        all_users = list_global_users(
            conn,
            search=f_search or None,
            department_id=f_dept or None,
            section_id=f_sec or None,
        )
        if f_emp_type:
            all_users = [u for u in all_users
                         if (u.get('emp_type') or 'PERMANENT') == f_emp_type]
        all_users = [u for u in all_users
                     if (u.get('emp_status') or 'ACTIVE') == 'ACTIVE']

        user_ids = [u['id'] for u in all_users]
        att_map  = db_mod.get_hajiri_data_from_logs(conn, user_ids, from_ad, to_ad)

        # ── Build per-employee summary ────────────────────────────────────────
        employees = []
        for u in all_users:
            uid       = u['id']
            days_data = att_map.get(uid, {})
            counts = {'P': 0, 'A': 0, 'SAT': 0, 'PH': 0, 'FH': 0, 'NH': 0, 'OH': 0}
            leave_counts: dict = {}
            total_ot = 0
            for day in days_data.values():
                sc = day.get('status_code') or 'A'
                if sc in counts:
                    counts[sc] += 1
                elif sc not in ('device',):
                    leave_counts[sc] = leave_counts.get(sc, 0) + 1
                total_ot += (day.get('ot_minutes') or 0)
            holiday_days = counts['PH'] + counts['FH'] + counts['NH'] + counts['OH']
            employees.append({
                **u,
                'days_data':    days_data,
                'counts':       counts,
                'leave_counts': leave_counts,
                'total_ot_min': total_ot,
                'total_ot_h':   f"{total_ot // 60}:{total_ot % 60:02d}" if total_ot else '',
                'holiday_days': holiday_days,
            })

        # ── Build column day list ─────────────────────────────────────────────
        from db import get_holidays as _ghols
        holiday_dates: dict = {}   # {date_str: display_code}
        for h in _ghols(conn, from_ad, to_ad):
            hd = h['holiday_ad']
            hs = hd.isoformat() if hasattr(hd, 'isoformat') else str(hd)
            htype = h.get('holiday_type', 'public')
            holiday_dates[hs] = 'उत्' if htype == 'festival' else 'सा'

        day_list = []
        d = _dt.fromisoformat(from_ad)
        end_d = _dt.fromisoformat(to_ad)
        while d <= end_d:
            nepal_dow = d.isoweekday() % 7
            ds = d.isoformat()
            day_list.append({
                'date':       ds,
                'day_num':    d.day,
                'is_weekend': (nepal_dow == 6),
                'is_holiday': ds in holiday_dates,
                'hol_code':   holiday_dates.get(ds, ''),
            })
            d += _td(days=1)

        return render(templates, request, 'reports_hajiri.html', {
            'sel_bs_year':  sel_year,
            'sel_bs_month': sel_month,
            'month_name':   mi['month_name'],
            'bs_year':      sel_year,
            'from_ad':      from_ad,
            'to_ad':        to_ad,
            'total_days':   len(day_list),
            'employees':    employees,
            'day_list':     day_list,
            'departments':  depts,
            'sections':     sections,
            'f_dept':       f_dept or '',
            'f_sec':        f_sec or '',
            'f_emp_type':   f_emp_type,
            'f_search':     f_search,
            'now_str':      _npt_now_str(),
            'COMPANY_NAME': COMPANY_NAME,
            'error':        None,
        })
    finally:
        conn.close()


# ---- Settings (Org Hierarchy / Shifts / Shift Rules) ---------------------


@app.get("/settings")
def settings_page(request: Request):
    conn = get_connection()
    try:
        from db import (get_all_departments, get_all_shifts, get_all_shift_rules,
                        get_all_global_users_with_dept, get_employees_for_report as _gef,
                        get_all_directorates, get_all_sections, get_all_units,
                        get_devices as _gd)
        departments   = get_all_departments(conn)
        shifts        = get_all_shifts(conn)
        shift_rules   = get_all_shift_rules(conn)
        employees     = get_all_global_users_with_dept(conn)
        directorates  = get_all_directorates(conn)
        sections      = get_all_sections(conn)
        units         = get_all_units(conn)
        all_emps      = _gef(conn)
        devices_list  = _gd(conn)
        pull_times_str = '09:00,13:00,17:30'
    finally:
        conn.close()
    return render(templates, request, 'settings.html', {
        'departments':   departments,
        'shifts':        shifts,
        'shift_rules':   shift_rules,
        'employees':     employees,
        'directorates':  directorates,
        'sections':      sections,
        'units':         units,
        'all_emps':      all_emps,
        'devices_list':  devices_list,
        'pull_times_str': pull_times_str,
        'source_id':     None,
        'source_users':  None,
        'source_device': None,
        'migrate_error': None,
    })


# ── Directorates ──────────────────────────────────────────────────────────────

@app.post("/settings/directorates/add")
async def add_directorate(request: Request):
    form = await request.form()
    name = (form.get('name') or '').strip()
    if not name:
        return redirect_with_flash('/settings', 'error', 'Directorate name is required.')
    conn = get_connection()
    try:
        from db import create_directorate
        create_directorate(conn, name, app_user_id=_current_user_id(request))
    except Exception as exc:
        return redirect_with_flash('/settings', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/settings', 'success', f"Directorate '{name}' added.")

@app.post("/settings/directorates/{did}/delete")
async def delete_directorate_route(request: Request, did: int):
    conn = get_connection()
    try:
        from db import delete_directorate
        delete_directorate(conn, did)
    except Exception as exc:
        return redirect_with_flash('/settings', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/settings', 'success', 'Directorate deleted.')


# ── Departments ───────────────────────────────────────────────────────────────

@app.post("/settings/departments/add")
async def add_department(request: Request):
    form = await request.form()
    name    = (form.get('name') or '').strip()
    dir_id  = _int_param(form.get('directorate_id'))
    if not name:
        return redirect_with_flash('/settings', 'error', 'Department name is required.')
    conn = get_connection()
    try:
        from db import create_department
        create_department(conn, name, app_user_id=_current_user_id(request))
        if dir_id:
            with conn.cursor() as cur:
                cur.execute("UPDATE departments SET directorate_id=%s WHERE name=%s", (dir_id, name))
            conn.commit()
    except Exception as exc:
        return redirect_with_flash('/settings', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/settings', 'success', f"Department '{name}' added.")

@app.post("/settings/departments/{dept_id}/delete")
async def delete_department_route(request: Request, dept_id: int):
    conn = get_connection()
    try:
        from db import delete_department
        delete_department(conn, dept_id)
    except Exception as exc:
        return redirect_with_flash('/settings', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/settings', 'success', 'Department deleted.')


# ── Sections ──────────────────────────────────────────────────────────────────

@app.post("/settings/sections/add")
async def add_section(request: Request):
    form    = await request.form()
    name    = (form.get('name') or '').strip()
    dept_id = _int_param(form.get('department_id'))
    if not name or not dept_id:
        return redirect_with_flash('/settings', 'error', 'Section name and department are required.')
    conn = get_connection()
    try:
        from db import create_section
        create_section(conn, name, dept_id, app_user_id=_current_user_id(request))
    except Exception as exc:
        return redirect_with_flash('/settings', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/settings', 'success', f"Section '{name}' added.")

@app.post("/settings/sections/{sid}/delete")
async def delete_section_route(request: Request, sid: int):
    conn = get_connection()
    try:
        from db import delete_section
        delete_section(conn, sid)
    except Exception as exc:
        return redirect_with_flash('/settings', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/settings', 'success', 'Section deleted.')


# ── Units ─────────────────────────────────────────────────────────────────────

@app.post("/settings/units/add")
async def add_unit(request: Request):
    form    = await request.form()
    name    = (form.get('name') or '').strip()
    sec_id  = _int_param(form.get('section_id'))
    if not name or not sec_id:
        return redirect_with_flash('/settings', 'error', 'Unit name and section are required.')
    conn = get_connection()
    try:
        from db import create_unit
        create_unit(conn, name, sec_id, app_user_id=_current_user_id(request))
    except Exception as exc:
        return redirect_with_flash('/settings', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/settings', 'success', f"Unit '{name}' added.")

@app.post("/settings/units/{uid}/delete")
async def delete_unit_route(request: Request, uid: int):
    conn = get_connection()
    try:
        from db import delete_unit
        delete_unit(conn, uid)
    except Exception as exc:
        return redirect_with_flash('/settings', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/settings', 'success', 'Unit deleted.')


# ── Shifts ────────────────────────────────────────────────────────────────────

@app.post("/settings/shifts/add")
async def add_shift(request: Request):
    form  = await request.form()
    name  = (form.get('name') or '').strip()
    start = (form.get('start_time') or '').strip()
    end   = (form.get('end_time') or '').strip()
    if not (name and start and end):
        return redirect_with_flash('/settings', 'error', 'Name, start time and end time are required.')
    conn = get_connection()
    try:
        from db import create_shift
        create_shift(conn, name, start, end, app_user_id=_current_user_id(request))
    except Exception as exc:
        return redirect_with_flash('/settings', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/settings', 'success', f"Shift '{name}' added.")

@app.post("/settings/shifts/{shift_id}/delete")
async def delete_shift_route(request: Request, shift_id: int):
    conn = get_connection()
    try:
        from db import delete_shift
        delete_shift(conn, shift_id)
    except Exception as exc:
        return redirect_with_flash('/settings', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/settings', 'success', 'Shift deleted.')


# ── Shift Rules ───────────────────────────────────────────────────────────────

@app.post("/settings/shift-rules/add")
async def add_shift_rule(request: Request):
    form        = await request.form()
    shift_id    = _int_param(form.get('shift_id'))
    from_date   = (form.get('from_date') or '').strip()
    to_date     = (form.get('to_date') or '').strip() or None
    target_type = form.get('target_type', 'employee')

    if not shift_id or not from_date:
        return redirect_with_flash('/settings', 'error', 'Shift and from-date are required.')

    # Collect all target IDs based on type
    g_user_ids = []
    dept_id    = dir_id = sec_id = unit_id = None

    if target_type == 'employee':
        raw = form.getlist('global_user_id[]') or form.getlist('global_user_id')
        g_user_ids = [_int_param(v) for v in raw if _int_param(v)]
        if not g_user_ids:
            return redirect_with_flash('/settings', 'error', 'Select at least one employee.')
    elif target_type == 'department':
        dept_id = _int_param(form.get('department_id'))
        if not dept_id:
            return redirect_with_flash('/settings', 'error', 'Select a department.')
    elif target_type == 'section':
        sec_id = _int_param(form.get('section_id'))
        if not sec_id:
            return redirect_with_flash('/settings', 'error', 'Select a section.')
    elif target_type == 'unit':
        unit_id = _int_param(form.get('unit_id'))
        if not unit_id:
            return redirect_with_flash('/settings', 'error', 'Select a unit.')
    elif target_type == 'directorate':
        dir_id = _int_param(form.get('directorate_id'))
        if not dir_id:
            return redirect_with_flash('/settings', 'error', 'Select a directorate.')

    conn = get_connection()
    try:
        from db import create_shift_rule
        uid_who = _current_user_id(request)
        if g_user_ids:
            for gid in g_user_ids:
                create_shift_rule(conn, shift_id, from_date, to_date, gid,
                                  None, None, None, None, app_user_id=uid_who)
        else:
            create_shift_rule(conn, shift_id, from_date, to_date, None,
                              dept_id, dir_id, sec_id, unit_id, app_user_id=uid_who)
    except Exception as exc:
        return redirect_with_flash('/settings', 'error', str(exc))
    finally:
        conn.close()
    count = len(g_user_ids) if g_user_ids else 1
    return redirect_with_flash('/settings', 'success',
                                f"{count} shift rule(s) added.")

@app.post("/settings/shift-rules/{rule_id}/delete")
async def delete_shift_rule_route(request: Request, rule_id: int):
    conn = get_connection()
    try:
        from db import delete_shift_rule
        delete_shift_rule(conn, rule_id)
    except Exception as exc:
        return redirect_with_flash('/settings', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/settings', 'success', 'Shift rule deleted.')


# ── Employee Org Assignment ───────────────────────────────────────────────────

@app.post("/settings/employees/{global_id}/org")
async def set_emp_org(request: Request, global_id: int):
    form    = await request.form()
    dept_id = _int_param(form.get('department_id'))
    sec_id  = _int_param(form.get('section_id'))
    unit_id = _int_param(form.get('unit_id'))
    conn = get_connection()
    try:
        from db import set_employee_org
        set_employee_org(conn, global_id, dept_id, sec_id, unit_id)
    except Exception as exc:
        return redirect_with_flash('/settings', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/settings', 'success', 'Employee org assignment updated.')


# ── Company Settings ──────────────────────────────────────────────────────────

@app.post("/settings/company")
async def update_company(request: Request):
    """Update company settings (name, logo, address, etc.)."""
    form = await request.form()
    company_name = (form.get('company_name') or '').strip()
    logo_url = (form.get('logo_url') or '').strip()
    address = (form.get('address') or '').strip()
    phone = (form.get('phone') or '').strip()
    email = (form.get('email') or '').strip()
    website = (form.get('website') or '').strip()
    pan_number = (form.get('pan_number') or '').strip()
    fiscal_year_bs = (form.get('fiscal_year_bs') or '').strip()

    if not company_name:
        return redirect_with_flash('/settings', 'error', 'Company name is required.')

    conn = get_connection()
    try:
        from db import update_company_settings
        user_id = request.session.get('user_id', 1)
        update_company_settings(
            conn,
            updated_by=user_id,
            company_name=company_name,
            logo_url=logo_url or None,
            address=address or None,
            phone=phone or None,
            email=email or None,
            website=website or None,
            pan_number=pan_number or None,
            fiscal_year_bs=fiscal_year_bs or None,
        )
    finally:
        conn.close()
    return redirect_with_flash('/settings', 'success', 'Company settings updated.')


# ─── Leave Management ────────────────────────────────────────────────────────


@app.get("/leaves")
def leaves_page(request: Request,
                tab:            str = 'applications',
                status:         str | None = None,
                emp_key:        str | None = None,
                from_bs:        str | None = None,
                to_bs:          str | None = None,
                bs_year:        str | None = None,
                bs_month:       str | None = None,
                department_id:  str | None = None,
                section_id:     str | None = None,
                directorate_id: str | None = None,
                page:           str | None = None):
    from db import (get_all_leave_types, get_leave_applications,
                    get_all_global_users_with_dept, get_leave_balances,
                    get_employee_leave_balance, get_all_departments,
                    get_all_directorates, get_all_sections)
    from nepali_utils import bs_to_ad, NEPALI_MONTHS, bs_month_info

    today_bs    = _today_bs()
    cur_bs_year = int(today_bs[:4]) if today_bs else 2082

    sel_year  = _int_param(bs_year) or cur_bs_year
    sel_month = _int_param(bs_month) or 0
    f_dept    = _int_param(department_id)
    f_sec     = _int_param(section_id)
    f_dir     = _int_param(directorate_id)
    page_num  = max(1, _int_param(page) or 1)
    per_page  = 25

    # Month-wise filter overrides manual date range
    if sel_month:
        mi      = bs_month_info(sel_year, sel_month)
        from_ad = mi['first_ad'] if mi else None
        to_ad   = mi['last_ad']  if mi else None
        from_bs_disp = to_bs_disp = ''
    else:
        from_ad      = bs_to_ad(from_bs) if from_bs else None
        to_ad        = bs_to_ad(to_bs)   if to_bs   else None
        from_bs_disp = from_bs or ''
        to_bs_disp   = to_bs   or ''

    conn = get_connection()
    try:
        ltypes    = get_all_leave_types(conn)
        all_emps  = get_all_global_users_with_dept(conn)
        all_depts = get_all_departments(conn)
        all_dirs  = get_all_directorates(conn)
        all_sects = get_all_sections(conn)

        sel_gid = None
        if emp_key and emp_key.startswith('g'):
            try:
                sel_gid = int(emp_key[1:])
            except ValueError:
                pass

        apps = get_leave_applications(conn,
                                      global_user_id=sel_gid,
                                      status=status or None,
                                      from_ad=from_ad,
                                      to_ad=to_ad,
                                      department_id=f_dept,
                                      section_id=f_sec,
                                      directorate_id=f_dir,
                                      limit=5000)

        # Stats
        total_apps     = len(apps)
        total_approved = sum(1 for a in apps if a['status'] == 'approved')
        total_pending  = sum(1 for a in apps if a['status'] == 'pending')
        total_rejected = sum(1 for a in apps if a['status'] == 'rejected')
        total_days     = sum(float(a.get('days') or 0) for a in apps)
        approved_days  = sum(float(a.get('days') or 0) for a in apps if a['status'] == 'approved')

        # Dashboard breakdowns (computed from full unfiltered apps list)
        from collections import defaultdict as _dd
        _by_type: dict = _dd(lambda: {'count': 0, 'approved': 0, 'pending': 0, 'rejected': 0, 'days': 0.0, 'approved_days': 0.0})
        _by_dept: dict = _dd(lambda: {'count': 0, 'approved': 0, 'days': 0.0, 'approved_days': 0.0})
        for a in apps:
            lt   = a.get('leave_type_name') or 'Unknown'
            dept = a.get('department_name')  or '—'
            d    = float(a.get('days') or 0)
            s    = a.get('status', '')
            _by_type[lt]['count'] += 1
            _by_type[lt][s]       = _by_type[lt].get(s, 0) + 1
            _by_type[lt]['days']  += d
            _by_dept[dept]['count'] += 1
            _by_dept[dept]['days']  += d
            if s == 'approved':
                _by_type[lt]['approved_days'] += d
                _by_dept[dept]['approved']      = _by_dept[dept].get('approved', 0) + 1
                _by_dept[dept]['approved_days'] += d
        dash_by_type = sorted([{'name': k, **v} for k, v in _by_type.items()], key=lambda x: -x['days'])
        dash_by_dept = sorted([{'name': k, **v} for k, v in _by_dept.items()], key=lambda x: -x['days'])[:15]
        recent_apps  = apps[:10]

        # Pagination (for applications tab)
        total_pages = max(1, (total_apps + per_page - 1) // per_page)
        page_num    = min(page_num, total_pages)
        page_apps   = apps[(page_num - 1) * per_page: page_num * per_page]

        balances    = get_leave_balances(conn, sel_year)
        emp_balance: list = []
        if sel_gid:
            emp_balance = get_employee_leave_balance(conn, sel_gid, sel_year)
    finally:
        conn.close()

    _valid_tabs = {'dashboard', 'applications', 'balances', 'types'}
    return render(templates, request, 'leaves.html', {
        'tab':             tab if tab in _valid_tabs else 'applications',
        'leave_types':     ltypes,
        'leave_apps':      page_apps,
        'all_emps':        all_emps,
        'all_depts':       all_depts,
        'all_dirs':        all_dirs,
        'all_sects':       all_sects,
        'balances':        balances,
        'emp_balance':     emp_balance,
        'sel_status':      status or '',
        'sel_emp_key':     emp_key or '',
        'sel_gid':         sel_gid,
        'sel_from_bs':     from_bs_disp,
        'sel_to_bs':       to_bs_disp,
        'sel_year':        sel_year,
        'sel_month':       sel_month,
        'sel_dept':        f_dept  or '',
        'sel_sec':         f_sec   or '',
        'sel_dir':         f_dir   or '',
        'nepali_months':   NEPALI_MONTHS,
        'today_bs':        today_bs,
        'COMPANY_NAME':    COMPANY_NAME,
        'total_apps':      total_apps,
        'total_approved':  total_approved,
        'total_pending':   total_pending,
        'total_rejected':  total_rejected,
        'total_days':      total_days,
        'approved_days':   approved_days,
        'page_num':        page_num,
        'total_pages':     total_pages,
        'per_page':        per_page,
        'dash_by_type':    dash_by_type,
        'dash_by_dept':    dash_by_dept,
        'recent_apps':     recent_apps,
    })


@app.post("/leaves/add")
async def add_leave(request: Request):
    from db import (create_leave_application, get_holidays, count_leave_working_days)
    from nepali_utils import bs_to_ad, ad_to_bs_tuple
    form = await request.form()

    def _fp(k, d=''):  return (form.get(k) or d).strip()
    def _fi(k):
        try: return int(form.get(k) or '')
        except: return None

    global_user_id = _fi('global_user_id')
    leave_type_id  = _fi('leave_type_id')
    from_bs        = _fp('from_bs')
    to_bs          = _fp('to_bs')
    reason         = _fp('reason')
    status         = _fp('status') or 'approved'

    if not global_user_id or not leave_type_id or not from_bs or not to_bs:
        return redirect_with_flash('/leaves', 'error',
                                   'Employee, leave type, and dates are required.')

    from_ad = bs_to_ad(from_bs)
    to_ad   = bs_to_ad(to_bs)
    if not from_ad or not to_ad:
        return redirect_with_flash('/leaves', 'error', 'Invalid BS dates.')
    if from_ad > to_ad:
        return redirect_with_flash('/leaves', 'error', 'From date must be before to date.')

    try:
        days_val = float(_fp('days') or '0')
    except ValueError:
        days_val = 0.0
    if days_val <= 0:
        conn = get_connection()
        try:
            hols = get_holidays(conn, from_ad, to_ad)
        finally:
            conn.close()
        days_val = count_leave_working_days(from_ad, to_ad, hols)
    if days_val <= 0:
        return redirect_with_flash('/leaves', 'error',
                                   'No working days in selected range (check for holidays/weekend).')

    today_bs = _today_bs()
    conn = get_connection()
    try:
        create_leave_application(conn, global_user_id, leave_type_id,
                                  from_bs, to_bs, from_ad, to_ad,
                                  days_val, reason, today_bs, status,
                                  app_user_id=_current_user_id(request))
    except Exception as exc:
        return redirect_with_flash('/leaves', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/leaves', 'success',
                                f'Leave added ({days_val} days, {status}).')


@app.post("/leaves/{app_id}/approve")
async def approve_leave(request: Request, app_id: int):
    from db import update_leave_status
    form = await request.form()
    remarks = (form.get('remarks') or '').strip()
    conn = get_connection()
    try:
        cu = _current_user(request)
        approver = cu['display_name'] if cu else 'admin'
        update_leave_status(conn, app_id, 'approved', remarks, approver,
                            app_user_id=_current_user_id(request))
    except Exception as exc:
        return redirect_with_flash('/leaves', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/leaves', 'success', 'Leave approved.')


@app.post("/leaves/{app_id}/reject")
async def reject_leave(request: Request, app_id: int):
    from db import update_leave_status
    form = await request.form()
    remarks = (form.get('remarks') or '').strip()
    conn = get_connection()
    try:
        cu = _current_user(request)
        approver = cu['display_name'] if cu else 'admin'
        update_leave_status(conn, app_id, 'rejected', remarks, approver,
                            app_user_id=_current_user_id(request))
    except Exception as exc:
        return redirect_with_flash('/leaves', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/leaves', 'success', 'Leave rejected.')


@app.post("/leaves/{app_id}/delete")
async def delete_leave(request: Request, app_id: int):
    from db import delete_leave_application
    conn = get_connection()
    try:
        delete_leave_application(conn, app_id)
    except Exception as exc:
        return redirect_with_flash('/leaves', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/leaves', 'success', 'Leave record deleted.')


@app.post("/leaves/allocate")
async def allocate_leaves(request: Request):
    from db import allocate_annual_leaves
    form    = await request.form()
    bs_year = int(form.get('bs_year') or _today_bs()[:4])
    conn    = get_connection()
    try:
        count = allocate_annual_leaves(conn, bs_year)
    except Exception as exc:
        return redirect_with_flash('/leaves', 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash('/leaves', 'success',
                                f'Annual leaves allocated for {bs_year} BS ({count} records).')


@app.get("/api/leave-days")
def api_leave_days(from_ad: str, to_ad: str):
    from db import get_holidays, count_leave_working_days
    from fastapi.responses import JSONResponse
    conn = get_connection()
    try:
        hols = get_holidays(conn, from_ad, to_ad)
    finally:
        conn.close()
    days = count_leave_working_days(from_ad, to_ad, hols)
    return JSONResponse({'days': days})


# ─── Holiday / Monthly Calendar ───────────────────────────────────────────────


@app.get("/calendar")
def calendar_page(request: Request,
                  bs_year:  str | None = None,
                  bs_month: str | None = None):
    from db import get_holidays
    from nepali_utils import bs_month_info, ad_to_bs_tuple, NEPALI_MONTHS
    import datetime

    today_bs = _today_bs()
    if not bs_year or not bs_month:
        parts = today_bs.split('-') if today_bs else ['2082', '1', '1']
        bs_year  = bs_year  or parts[0]
        bs_month = bs_month or parts[1]

    try:
        y, m = int(bs_year), int(bs_month)
    except ValueError:
        y, m = 2082, 1

    mi = bs_month_info(y, m)
    if not mi:
        return render(templates, request, 'calendar.html', {
            'error': f'Invalid BS year/month: {y}/{m}',
            'COMPANY_NAME': COMPANY_NAME,
        })

    from_ad, to_ad = mi['first_ad'], mi['last_ad']
    conn = get_connection()
    try:
        holiday_rows = get_holidays(conn, from_ad, to_ad)
    finally:
        conn.close()

    holiday_map = {}
    for h in holiday_rows:
        k = h['holiday_ad'].isoformat() if hasattr(h['holiday_ad'], 'isoformat') else str(h['holiday_ad'])
        holiday_map[k] = h

    from datetime import date as _date, timedelta
    NEPAL_DAYS = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    days_info: list = []
    d     = _date.fromisoformat(from_ad)
    to_d  = _date.fromisoformat(to_ad)
    while d <= to_d:
        dow    = d.isoweekday() % 7
        is_sat = (dow == 6)
        ds     = d.isoformat()
        h      = holiday_map.get(ds)
        bs_t   = ad_to_bs_tuple(d)
        days_info.append({
            'date_ad':      ds,
            'bs_day':       bs_t[2] if bs_t else d.day,
            'day_name':     NEPAL_DAYS[dow],
            'is_weekend':   is_sat,
            'is_holiday':   bool(h),
            'holiday_name': h['name'] if h else '',
            'holiday_type': h['holiday_type'] if h else '',
        })
        d += timedelta(days=1)

    # Build calendar grid (weeks starting Sunday)
    first_dow = mi['first_weekday']
    cal_grid:  list = []
    row: list = [None] * first_dow
    for dd in days_info:
        row.append(dd)
        if len(row) == 7:
            cal_grid.append(row)
            row = []
    if row:
        while len(row) < 7:
            row.append(None)
        cal_grid.append(row)

    working_days = sum(1 for dd in days_info if not dd['is_weekend'] and not dd['is_holiday'])
    holiday_count = len(holiday_rows)
    weekend_count = sum(1 for dd in days_info if dd['is_weekend'])

    prev_m = m - 1 if m > 1 else 12
    prev_y = y if m > 1 else y - 1
    next_m = m + 1 if m < 12 else 1
    next_y = y if m < 12 else y + 1

    return render(templates, request, 'calendar.html', {
        'bs_year':       y,
        'bs_month':      m,
        'month_name':    mi['month_name'],
        'nepali_months': NEPALI_MONTHS,
        'days':          days_info,
        'cal_grid':      cal_grid,
        'holidays':      holiday_rows,
        'working_days':  working_days,
        'holiday_count': holiday_count,
        'weekend_count': weekend_count,
        'total_days':    len(days_info),
        'from_ad':       from_ad,
        'to_ad':         to_ad,
        'prev_y':        prev_y, 'prev_m': prev_m,
        'next_y':        next_y, 'next_m': next_m,
        'today_bs':      today_bs,
        'COMPANY_NAME':  COMPANY_NAME,
    })


@app.post("/calendar/holidays/add")
async def add_holiday_route(request: Request):
    from db import create_holiday
    from nepali_utils import bs_to_ad
    form = await request.form()
    name         = (form.get('name') or '').strip()
    holiday_bs   = (form.get('holiday_bs') or '').strip()
    holiday_ad   = (form.get('holiday_ad') or '').strip()
    holiday_type = (form.get('holiday_type') or 'public').strip()
    description  = (form.get('description') or '').strip()
    redirect_to  = (form.get('redirect_to') or '/calendar').strip()

    if not name or not holiday_bs:
        return redirect_with_flash(redirect_to, 'error', 'Holiday name and BS date are required.')
    if not holiday_ad:
        holiday_ad = bs_to_ad(holiday_bs)
    if not holiday_ad:
        return redirect_with_flash(redirect_to, 'error', 'Invalid BS date — cannot convert to AD.')

    # Redirect to the calendar month that contains the added holiday, not the current view
    hbs_parts = holiday_bs.replace('/', '-').split('-')
    if len(hbs_parts) >= 2:
        try:
            redirect_to = f"/calendar?bs_year={int(hbs_parts[0])}&bs_month={int(hbs_parts[1])}"
        except ValueError:
            pass

    conn = get_connection()
    try:
        create_holiday(conn, name, holiday_ad, holiday_bs, holiday_type, description,
                       app_user_id=_current_user_id(request))
    except Exception as exc:
        return redirect_with_flash(redirect_to, 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash(redirect_to, 'success', f'Holiday "{name}" added.')


@app.post("/calendar/holidays/{h_id}/delete")
async def delete_holiday_route(request: Request, h_id: int):
    from db import delete_holiday
    form        = await request.form()
    redirect_to = (form.get('redirect_to') or '/calendar').strip()
    conn = get_connection()
    try:
        delete_holiday(conn, h_id)
    except Exception as exc:
        return redirect_with_flash(redirect_to, 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash(redirect_to, 'success', 'Holiday deleted.')


@app.post("/calendar/holidays/{h_id}/edit")
async def edit_holiday_route(request: Request, h_id: int):
    from db import update_holiday
    from nepali_utils import bs_to_ad
    form         = await request.form()
    name         = (form.get('name') or '').strip()
    holiday_bs   = (form.get('holiday_bs') or '').strip()
    holiday_ad   = (form.get('holiday_ad') or '').strip()
    holiday_type = (form.get('holiday_type') or 'public').strip()
    description  = (form.get('description') or '').strip()
    redirect_to  = (form.get('redirect_to') or '/calendar').strip()

    if not name or not holiday_bs:
        return redirect_with_flash(redirect_to, 'error', 'Holiday name and BS date are required.')
    if not holiday_ad:
        holiday_ad = bs_to_ad(holiday_bs)
    if not holiday_ad:
        return redirect_with_flash(redirect_to, 'error', 'Invalid BS date — cannot convert to AD.')

    conn = get_connection()
    try:
        update_holiday(conn, h_id, name, holiday_ad, holiday_bs, holiday_type, description)
    except Exception as exc:
        return redirect_with_flash(redirect_to, 'error', str(exc))
    finally:
        conn.close()
    return redirect_with_flash(redirect_to, 'success', f'Holiday "{name}" updated.')


# ─── Daily Attendance Report ──────────────────────────────────────────────────


def _resolve_daily_date(date_bs, date_ad):
    from nepali_utils import bs_to_ad, ad_to_bs
    import datetime, zoneinfo as _zi
    if date_bs and not date_ad:
        date_ad = bs_to_ad(date_bs)
    if not date_ad:
        try:
            date_ad = datetime.datetime.now(_zi.ZoneInfo('Asia/Kathmandu')).date().isoformat()
        except Exception:
            date_ad = datetime.date.today().isoformat()
    if not date_bs:
        date_bs = ad_to_bs(date_ad) or ''
    return date_bs, date_ad


def _build_daily_data(conn, date_ad: str, name_q: str = '', dept_q: str = ''):
    """Shared data builder for both the HTML and Excel daily report routes."""
    import datetime, psycopg2.extras as _pex
    from db import (get_daily_present_list, get_daily_present_gu_ids,
                    get_leaves_for_date, get_all_global_users_with_dept)

    def _att_key(emp, field='att_id'):
        v = emp.get(field) or ''
        try:    return (0, int(v))
        except: return (1, str(v))

    # present grouped by al.user_id — same dedup as the attendance page
    present_all = get_daily_present_list(conn, date_ad)
    present_all.sort(key=lambda e: _att_key(e, 'att_id'))
    for emp in present_all:
        emp['company_id'] = emp.get('att_id', '')

    # separate query: which global_user IDs actually have punches today
    present_gu_ids = get_daily_present_gu_ids(conn, date_ad)

    leave_rows    = get_leaves_for_date(conn, date_ad)
    on_leave_ids  = {r['global_user_id'] for r in leave_rows}
    on_leave_list = [r for r in leave_rows if r['global_user_id'] not in present_gu_ids]

    d_obj       = datetime.date.fromisoformat(date_ad)
    is_saturday = (d_obj.isoweekday() % 7 == 6)
    with conn.cursor(cursor_factory=_pex.DictCursor) as cur:
        cur.execute("SELECT * FROM holidays WHERE holiday_ad = %s", (date_ad,))
        h_row = cur.fetchone()
    holiday = dict(h_row) if h_row else None

    all_emps    = get_all_global_users_with_dept(conn)
    absent_list = []
    if not is_saturday and not holiday:
        for emp in all_emps:
            uid = emp['id']
            if uid in present_gu_ids or uid in on_leave_ids:
                continue
            cid = emp.get('company_id') or ''
            absent_list.append({
                'emp_name':        emp.get('name') or '(unknown)',
                'att_id':          cid,
                'company_id':      cid,
                'department_name': emp.get('department_name') or '',
                'section_name':    emp.get('section_name') or '',
            })
        absent_list.sort(key=lambda e: _att_key(e, 'att_id'))

    on_leave_list.sort(key=lambda e: _att_key(e, 'company_id'))

    # dept summary from full (unfiltered) data
    dept_summary: dict = {}
    for emp in present_all:
        dn = emp.get('department_name') or 'No Department'
        dept_summary.setdefault(dn, {'present': 0, 'on_leave': 0, 'absent': 0})
        dept_summary[dn]['present'] += 1
    for emp in on_leave_list:
        dn = emp.get('department_name') or 'No Department'
        dept_summary.setdefault(dn, {'present': 0, 'on_leave': 0, 'absent': 0})
        dept_summary[dn]['on_leave'] += 1
    for emp in absent_list:
        dn = emp.get('department_name') or 'No Department'
        dept_summary.setdefault(dn, {'present': 0, 'on_leave': 0, 'absent': 0})
        dept_summary[dn]['absent'] += 1

    # apply name/dept filters after dept summary is built
    nq = name_q.strip().lower()
    dq = dept_q.strip().lower()
    if nq or dq:
        def _m(emp, nk='emp_name'):
            nm = (emp.get(nk) or emp.get('employee_name') or '').lower()
            dn = (emp.get('department_name') or '').lower()
            return (not nq or nq in nm) and (not dq or dq in dn)
        present_all   = [e for e in present_all   if _m(e)]
        on_leave_list = [e for e in on_leave_list if _m(e, 'employee_name')]
        absent_list   = [e for e in absent_list   if _m(e)]

    return {
        'present_all':     present_all,
        'present':         present_all,   # alias for routes that use summary.get('present')
        'on_leave':        on_leave_list,
        'absent':          absent_list,
        'dept_summary':    dict(sorted(dept_summary.items())),
        'total_employees': len(all_emps),
        'is_saturday':     is_saturday,
        'is_holiday':      bool(holiday),
        'holiday':         holiday,
        'totals': {
            'present':  len(present_all),
            'on_leave': len(on_leave_list),
            'absent':   len(absent_list),
        },
    }


@app.get("/reports/daily")
def daily_report(request: Request,
                 date_bs: str | None = None,
                 date_ad: str | None = None,
                 name_q:  str | None = None,
                 dept_q:  str | None = None,
                 page:    str | None = None):
    from nepali_utils import ad_to_bs

    date_bs, date_ad = _resolve_daily_date(date_bs, date_ad)
    nq = (name_q or '').strip()
    dq = (dept_q or '').strip()

    conn = get_connection()
    try:
        data = _build_daily_data(conn, date_ad, nq, dq)
    except Exception as exc:
        conn.close()
        return render(templates, request, 'reports_daily.html', {
            'error': str(exc), 'sel_date_bs': date_bs, 'sel_date_ad': date_ad,
            'name_q': nq, 'dept_q': dq,
            'COMPANY_NAME': COMPANY_NAME, 'COMPANY_ADDRESS': COMPANY_ADDRESS,
        })
    finally:
        conn.close()

    present_all = data['present_all']
    per_page    = 100
    page_num    = max(1, _int_param(page) or 1)
    total_pres  = len(present_all)
    total_pages = max(1, (total_pres + per_page - 1) // per_page)
    page_num    = min(page_num, total_pages)
    present_page = present_all[(page_num - 1) * per_page : page_num * per_page]

    fqs = urlencode({'date_bs': date_bs, 'date_ad': date_ad,
                     'name_q': nq, 'dept_q': dq})
    return render(templates, request, 'reports_daily.html', {
        'sel_date_bs':     date_bs,
        'sel_date_ad':     date_ad,
        'name_q':          nq,
        'dept_q':          dq,
        'present':         present_page,
        'on_leave':        data['on_leave'],
        'absent':          data['absent'],
        'dept_summary':    data['dept_summary'],
        'total_employees': data['total_employees'],
        'is_saturday':     data['is_saturday'],
        'is_holiday':      data['is_holiday'],
        'holiday':         data['holiday'],
        'totals': {
            'present':  total_pres,
            'on_leave': len(data['on_leave']),
            'absent':   len(data['absent']),
        },
        'page':        page_num,
        'total_pages': total_pages,
        'per_page':    per_page,
        'filter_qs':   fqs,
        'COMPANY_NAME':    COMPANY_NAME,
        'COMPANY_ADDRESS': COMPANY_ADDRESS,
    })


@app.get("/reports/daily/excel")
def daily_report_excel(
    date_bs: str | None = None,
    date_ad: str | None = None,
    name_q:  str | None = None,
    dept_q:  str | None = None,
):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    date_bs, date_ad = _resolve_daily_date(date_bs, date_ad)
    nq = (name_q or '').strip()
    dq = (dept_q or '').strip()

    conn = get_connection()
    try:
        data = _build_daily_data(conn, date_ad, nq, dq)
    finally:
        conn.close()

    present_all   = data['present_all']
    on_leave_list = data['on_leave']
    absent_list   = data['absent']
    dept_summary  = data['dept_summary']

    wb = openpyxl.Workbook()
    thin     = Side(style='thin')
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1769E0")
    hdr_font = Font(bold=True, color="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="EEF6FF")

    def _ws_header(ws, title, ncols):
        last = get_column_letter(ncols)
        ws.merge_cells(f'A1:{last}1')
        ws.cell(1,1).value = COMPANY_NAME or ''
        ws.cell(1,1).font = Font(bold=True, size=13)
        ws.cell(1,1).alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A2:{last}2')
        ws.cell(2,1).value = COMPANY_ADDRESS or ''
        ws.cell(2,1).font = Font(size=10)
        ws.cell(2,1).alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A3:{last}3')
        ws.cell(3,1).value = f"{title}  |  {date_ad}  ({date_bs} BS)"
        ws.cell(3,1).font = Font(italic=True, size=10)
        ws.cell(3,1).alignment = Alignment(horizontal='center')
        ws.append([])

    def _style_header(ws):
        for cell in ws[ws.max_row]:
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = bdr

    def _style_row(ws, alt=False):
        for cell in ws[ws.max_row]:
            cell.border = bdr
            cell.alignment = Alignment(vertical='top')
            if alt: cell.fill = alt_fill

    def _auto_width(ws):
        for i in range(1, ws.max_column + 1):
            col = get_column_letter(i)
            mx = max((len(str(c.value)) for row in ws.iter_rows(min_col=i, max_col=i)
                      for c in row if c.value is not None), default=8)
            ws.column_dimensions[col].width = min(mx + 4, 42)

    PRES_COLS = ['#','Name','Att. ID','Department','Section','Check-In','Check-Out','Hours','Punches']
    ws1 = wb.active; ws1.title = "Present"
    _ws_header(ws1, "Present Employees", len(PRES_COLS))
    ws1.append(PRES_COLS); _style_header(ws1)
    for i, emp in enumerate(present_all, 1):
        fp = emp.get('first_punch'); lp = emp.get('last_punch')
        ci = fp.strftime('%H:%M') if fp else '-'
        has_co = lp and fp and lp != fp
        co = lp.strftime('%H:%M') if has_co else '-'
        hrs = '-'
        if has_co:
            s = (lp - fp).total_seconds()
            hrs = f"{int(s//3600)}:{int((s%3600)//60):02d}"
        ws1.append([i, emp.get('emp_name') or '-', emp.get('att_id') or '-',
                    emp.get('department_name') or '-', emp.get('section_name') or '-',
                    ci, co, hrs, emp.get('punch_count') or 0])
        _style_row(ws1, i % 2 == 0)

    ABS_COLS = ['#','Name','Att. ID','Department','Section']
    ws2 = wb.create_sheet("Absent")
    _ws_header(ws2, "Absent Employees", len(ABS_COLS))
    ws2.append(ABS_COLS); _style_header(ws2)
    for i, emp in enumerate(absent_list, 1):
        ws2.append([i, emp.get('emp_name') or '-', emp.get('att_id') or '-',
                    emp.get('department_name') or '-', emp.get('section_name') or '-'])
        _style_row(ws2, i % 2 == 0)

    OL_COLS = ['#','Name','Att. ID','Department','Leave Type']
    ws3 = wb.create_sheet("On Leave")
    _ws_header(ws3, "On Leave", len(OL_COLS))
    ws3.append(OL_COLS); _style_header(ws3)
    for i, emp in enumerate(on_leave_list, 1):
        ws3.append([i, emp.get('employee_name') or '-', emp.get('company_id') or '-',
                    emp.get('department_name') or '-', emp.get('leave_type_name') or '-'])
        _style_row(ws3, i % 2 == 0)

    DS_COLS = ['Department','Present','On Leave','Absent','Total']
    ws4 = wb.create_sheet("Dept Summary")
    _ws_header(ws4, "Department-wise Summary", len(DS_COLS))
    ws4.append(DS_COLS); _style_header(ws4)
    for i, (dept, counts) in enumerate(sorted(dept_summary.items()), 1):
        p = counts.get('present',0); ol = counts.get('on_leave',0); ab = counts.get('absent',0)
        ws4.append([dept, p, ol, ab, p+ol+ab])
        _style_row(ws4, i % 2 == 0)

    for ws in [ws1, ws2, ws3, ws4]:
        _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"daily_attendance_{date_ad}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ─── Day-wise Absent Report ───────────────────────────────────────────────────


@app.get("/reports/absent")
def absent_report(request: Request,
                  date_bs: str | None = None,
                  date_ad: str | None = None,
                  name_q:  str | None = None,
                  dept_q:  str | None = None):
    nq = (name_q or '').strip()
    dq = (dept_q or '').strip()
    date_bs, date_ad = _resolve_daily_date(date_bs, date_ad)

    conn = get_connection()
    try:
        data = _build_daily_data(conn, date_ad, nq, dq)
    except Exception as exc:
        data = {'error': str(exc), 'absent': [], 'is_saturday': False,
                'is_holiday': False, 'holiday': None, 'total_employees': 0,
                'totals': {'present': 0, 'on_leave': 0, 'absent': 0}}
    finally:
        conn.close()

    return render(templates, request, 'reports_absent.html', {
        **data,
        'sel_date_bs':   date_bs,
        'sel_date_ad':   date_ad,
        'name_q':        nq,
        'dept_q':        dq,
        'COMPANY_NAME':  COMPANY_NAME,
        'COMPANY_ADDRESS': COMPANY_ADDRESS,
    })


@app.get("/reports/absent/excel")
def absent_report_excel(
    date_bs: str | None = None,
    date_ad: str | None = None,
    name_q:  str | None = None,
    dept_q:  str | None = None,
):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    nq = (name_q or '').strip()
    dq = (dept_q or '').strip()
    date_bs, date_ad = _resolve_daily_date(date_bs, date_ad)

    conn = get_connection()
    try:
        data = _build_daily_data(conn, date_ad, nq, dq)
    except Exception:
        data = {'absent': []}
    finally:
        conn.close()

    absent = data.get('absent', [])

    COLS = ['#', 'Name', 'Att. ID', 'Department', 'Section']
    ncols = len(COLS)
    last  = get_column_letter(ncols)

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Absent"

    thin = Side(style='thin')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1769E0")
    hdr_font = Font(bold=True, color="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="EEF6FF")

    ws.merge_cells(f'A1:{last}1')
    ws.cell(1, 1).value = f"{COMPANY_NAME or ''} — Day-wise Absent Report"
    ws.cell(1, 1).font      = Font(bold=True, size=13)
    ws.cell(1, 1).alignment = Alignment(horizontal='center')
    ws.merge_cells(f'A2:{last}2')
    ws.cell(2, 1).value = f"{COMPANY_ADDRESS or ''}     Date: {date_bs} BS  |  {date_ad}"
    ws.cell(2, 1).alignment = Alignment(horizontal='center')
    ws.append([])

    ws.append(COLS)
    for c in ws[ws.max_row]:
        c.fill = hdr_fill; c.font = hdr_font; c.border = bdr
        c.alignment = Alignment(horizontal='center')

    for i, emp in enumerate(absent, 1):
        ws.append([i,
                   emp.get('emp_name') or '-',
                   emp.get('company_id') or '-',
                   emp.get('department_name') or '-',
                   emp.get('section_name') or '-'])
        for c in ws[ws.max_row]:
            c.border = bdr
            if i % 2 == 0: c.fill = alt_fill

    for ci in range(1, ncols + 1):
        col_letter = get_column_letter(ci)
        max_len = max((len(str(cell.value or ''))
                       for row in ws.iter_rows(min_col=ci, max_col=ci)
                       for cell in row if getattr(cell, 'value', None) is not None), default=8)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 42)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"absent_report_{date_ad}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ─── Department Attendance Report ─────────────────────────────────────────────


@app.get("/reports/dept-attendance")
def dept_attendance_report(request: Request,
                            date_bs: str | None = None,
                            date_ad: str | None = None):
    date_bs, date_ad = _resolve_daily_date(date_bs, date_ad)

    conn = get_connection()
    try:
        data = _build_daily_data(conn, date_ad)
    except Exception as exc:
        data = {'error': str(exc), 'dept_summary': {}, 'is_saturday': False,
                'is_holiday': False, 'holiday': None, 'total_employees': 0,
                'present': [], 'absent': [], 'on_leave': [],
                'totals': {'present': 0, 'on_leave': 0, 'absent': 0}}
    finally:
        conn.close()

    dept_detail: dict = {}
    for emp in data.get('present', []):
        dn = emp.get('department_name') or 'No Department'
        dept_detail.setdefault(dn, {'present': [], 'absent': [], 'on_leave': []})
        dept_detail[dn]['present'].append(emp)
    for emp in data.get('on_leave', []):
        dn = emp.get('department_name') or 'No Department'
        dept_detail.setdefault(dn, {'present': [], 'absent': [], 'on_leave': []})
        dept_detail[dn]['on_leave'].append(emp)
    for emp in data.get('absent', []):
        dn = emp.get('department_name') or 'No Department'
        dept_detail.setdefault(dn, {'present': [], 'absent': [], 'on_leave': []})
        dept_detail[dn]['absent'].append(emp)

    return render(templates, request, 'reports_dept_attendance.html', {
        **data,
        'dept_detail':   dict(sorted(dept_detail.items())),
        'sel_date_bs':   date_bs,
        'sel_date_ad':   date_ad,
        'COMPANY_NAME':  COMPANY_NAME,
        'COMPANY_ADDRESS': COMPANY_ADDRESS,
    })


@app.get("/reports/dept-attendance/excel")
def dept_attendance_excel(
    date_bs: str | None = None,
    date_ad: str | None = None,
):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    date_bs, date_ad = _resolve_daily_date(date_bs, date_ad)

    conn = get_connection()
    try:
        summary = _build_daily_data(conn, date_ad)
    except Exception:
        summary = {'dept_summary': {}}
    finally:
        conn.close()

    thin     = Side(style='thin')
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1769E0")
    hdr_font = Font(bold=True, color="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="EEF6FF")
    co_name  = COMPANY_NAME or ''
    co_addr  = COMPANY_ADDRESS or ''

    def _auto_width(wsh):
        for ci in range(1, wsh.max_column + 1):
            cl = get_column_letter(ci)
            mx = max((len(str(cell.value or ''))
                      for row in wsh.iter_rows(min_col=ci, max_col=ci)
                      for cell in row if getattr(cell, 'value', None) is not None), default=8)
            wsh.column_dimensions[cl].width = min(mx + 4, 42)

    def _hdr_row(wsh):
        for c in wsh[wsh.max_row]:
            c.fill = hdr_fill; c.font = hdr_font; c.border = bdr
            c.alignment = Alignment(horizontal='center')

    wb = openpyxl.Workbook()

    # Sheet 1: Dept Summary
    DS_COLS = ['Department', 'Present', 'On Leave', 'Absent', 'Total']
    ws = wb.active; ws.title = "Dept Summary"
    last = get_column_letter(len(DS_COLS))
    ws.merge_cells(f'A1:{last}1')
    ws.cell(1,1).value = f"{co_name} — Department Attendance Report"
    ws.cell(1,1).font = Font(bold=True, size=13); ws.cell(1,1).alignment = Alignment(horizontal='center')
    ws.merge_cells(f'A2:{last}2')
    ws.cell(2,1).value = f"{co_addr}     Date: {date_bs} BS  |  {date_ad}"
    ws.cell(2,1).alignment = Alignment(horizontal='center')
    ws.append([])
    ws.append(DS_COLS); _hdr_row(ws)
    for i, (dept, counts) in enumerate(sorted((summary.get('dept_summary') or {}).items()), 1):
        p = counts.get('present',0); ol = counts.get('on_leave',0); ab = counts.get('absent',0)
        ws.append([dept, p, ol, ab, p+ol+ab])
        if i % 2 == 0:
            for c in ws[ws.max_row]: c.fill = alt_fill; c.border = bdr
        else:
            for c in ws[ws.max_row]: c.border = bdr
    _auto_width(ws)

    # Sheet 2: Present employees
    PR_COLS = ['Department', 'Name', 'Att. ID', 'Check-In', 'Check-Out', 'Hours']
    ws2 = wb.create_sheet("Present by Dept")
    ws2.append(PR_COLS); _hdr_row(ws2)
    for i, emp in enumerate(sorted(summary.get('present', []),
                                   key=lambda x: (x.get('department_name') or '', x.get('emp_name') or '')), 1):
        fp = emp.get('first_punch'); lp = emp.get('last_punch')
        ci = fp.strftime('%H:%M') if fp else '-'
        has_co = lp and fp and lp != fp
        co = lp.strftime('%H:%M') if has_co else '-'
        hrs = '-'
        if has_co:
            secs = (lp - fp).total_seconds()
            hrs = f"{int(secs//3600)}:{int((secs%3600)//60):02d}"
        ws2.append([emp.get('department_name') or '-', emp.get('emp_name') or '-',
                    emp.get('company_id') or '-', ci, co, hrs])
        for c in ws2[ws2.max_row]: c.border = bdr
        if i % 2 == 0:
            for c in ws2[ws2.max_row]: c.fill = alt_fill
    _auto_width(ws2)

    # Sheet 3: Absent employees
    AB_COLS = ['Department', 'Name', 'Att. ID', 'Section']
    ws3 = wb.create_sheet("Absent by Dept")
    ws3.append(AB_COLS); _hdr_row(ws3)
    for i, emp in enumerate(sorted(summary.get('absent', []),
                                   key=lambda x: (x.get('department_name') or '', x.get('emp_name') or '')), 1):
        ws3.append([emp.get('department_name') or '-', emp.get('emp_name') or '-',
                    emp.get('company_id') or '-', emp.get('section_name') or '-'])
        for c in ws3[ws3.max_row]: c.border = bdr
        if i % 2 == 0:
            for c in ws3[ws3.max_row]: c.fill = alt_fill
    _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"dept_attendance_{date_ad}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ─── Monthly Summary Report ───────────────────────────────────────────────────


@app.get("/reports/monthly-summary")
def monthly_summary_report(request: Request,
                            bs_year:  str | None = None,
                            bs_month: str | None = None):
    from db import get_monthly_attendance_summary
    from nepali_utils import NEPALI_MONTHS
    import datetime

    today_bs = _today_bs()
    if not bs_year or not bs_month:
        parts    = today_bs.split('-') if today_bs else ['2082', '1', '1']
        bs_year  = bs_year  or parts[0]
        bs_month = bs_month or parts[1]

    try:
        y, m = int(bs_year), int(bs_month)
    except ValueError:
        y, m = 2082, 1

    conn = get_connection()
    try:
        summary = get_monthly_attendance_summary(conn, y, m)
    except Exception as exc:
        summary = {'error': str(exc)}
    finally:
        conn.close()

    from nepali_utils import NEPALI_MONTHS as _NM
    summary.setdefault('bs_year',       y)
    summary.setdefault('bs_month',      m)
    summary.setdefault('month_name',    _NM[m] if 1 <= m <= 12 else '')
    summary.setdefault('from_ad',       '')
    summary.setdefault('to_ad',         '')
    summary.setdefault('total_days',    0)
    summary.setdefault('working_days',  0)
    summary.setdefault('holiday_count', 0)
    summary.setdefault('weekend_count', 0)
    summary.setdefault('holidays',      [])
    summary.setdefault('employees',     [])
    summary['nepali_months'] = NEPALI_MONTHS
    summary['today_bs']      = today_bs
    summary['COMPANY_NAME']  = COMPANY_NAME
    return render(templates, request, 'reports_monthly_summary.html', summary)


# ---- Schedule editor ----------------------------------------------------


def _parse_hhmm(value: str) -> tuple[int, int]:
    hh, mm = value.strip().split(':')
    hour, minute = int(hh), int(mm)
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        raise ValueError("Time out of range")
    return hour, minute


@app.get("/schedule")
def schedule_view(request: Request):
    conn = get_connection()
    try:
        rows = db_mod.get_pull_schedule(conn)
    finally:
        conn.close()
    # Map "HH:MM" -> next_run string so the template can look it up directly
    next_run_by_time = {}
    for job in _scheduler_jobs_info():
        m = re.search(r'(\d{2}:\d{2})$', job['name'])
        if m:
            next_run_by_time[m.group(1)] = job['next_run']
    return render(templates, request, 'schedule.html', {
        'rows': rows,
        'next_run_by_time': next_run_by_time,
        'scheduler_timezone': SCHEDULER_TIMEZONE,
    })


@app.post("/schedule/add")
def schedule_add(request: Request, time_hhmm: str = Form(...), label: str = Form('')):
    try:
        hour, minute = _parse_hhmm(time_hhmm)
        conn = get_connection()
        try:
            db_mod.add_pull_schedule(conn, hour, minute, label.strip() or None)
        finally:
            conn.close()
        _restart_web_scheduler()
        return redirect_with_flash("/schedule", "success", f"Added pull time {hour:02d}:{minute:02d}.")
    except psycopg2.IntegrityError:
        return redirect_with_flash("/schedule", "error", "That time is already in the schedule.")
    except Exception as exc:
        return redirect_with_flash("/schedule", "error", f"Could not add time: {exc}")


@app.post("/schedule/{sched_id}/edit")
def schedule_edit(request: Request, sched_id: int, time_hhmm: str = Form(...),
                   label: str = Form(''), is_active: str = Form(None)):
    try:
        hour, minute = _parse_hhmm(time_hhmm)
        conn = get_connection()
        try:
            db_mod.update_pull_schedule(conn, sched_id, hour, minute,
                                        label.strip() or None, bool(is_active))
        finally:
            conn.close()
        _restart_web_scheduler()
        return redirect_with_flash("/schedule", "success", "Pull time updated.")
    except psycopg2.IntegrityError:
        return redirect_with_flash("/schedule", "error", "That time is already in the schedule.")
    except Exception as exc:
        return redirect_with_flash("/schedule", "error", f"Could not update time: {exc}")


@app.post("/schedule/{sched_id}/toggle")
def schedule_toggle(request: Request, sched_id: int):
    conn = get_connection()
    try:
        db_mod.toggle_pull_schedule(conn, sched_id)
    finally:
        conn.close()
    _restart_web_scheduler()
    return redirect_with_flash("/schedule", "success", "Pull time updated.")


@app.post("/schedule/{sched_id}/delete")
def schedule_delete(request: Request, sched_id: int):
    conn = get_connection()
    try:
        db_mod.delete_pull_schedule(conn, sched_id)
    finally:
        conn.close()
    _restart_web_scheduler()
    return redirect_with_flash("/schedule", "success", "Pull time removed.")


@app.post("/schedule/pull-now")
def schedule_pull_now(request: Request):
    if _pull_lock.locked():
        return redirect_with_flash("/schedule", "warning", "A pull is already in progress — please wait.")

    def _bg():
        try:
            _safe_pull()
        except Exception:
            _traceback.print_exc()

    _threading.Thread(target=_bg, daemon=True).start()
    return redirect_with_flash("/schedule", "success",
                               "Manual pull started. Check Sync Runs for results in a minute.")


@app.get("/kaaj")
def kaaj_page(request: Request):
    from nepali_utils import today_bs, bs_month_info
    from db import get_all_global_users_with_dept, get_all_departments, get_connection
    bs_year = request.query_params.get('bs_year')
    bs_month = request.query_params.get('bs_month')
    emp_key = request.query_params.get('emp_key')
    department_id = request.query_params.get('department_id')
    paid_filter = request.query_params.get('paid_filter')
    # Defaults
    today_bs_str = today_bs()
    if not bs_year or not bs_month:
        parts = today_bs_str.split('-') if today_bs_str else ['2082', '1', '1']
        bs_year = bs_year or parts[0]
        bs_month = bs_month or parts[1]
    try:
        y, m = int(bs_year), int(bs_month)
    except ValueError:
        y, m = 2082, 1
    mi = bs_month_info(y, m) if bs_month_info else None
    from_ad = mi['first_ad'] if mi else None
    to_ad = mi['last_ad'] if mi else None
    # We'll fetch employees and departments for the filters
    conn = get_connection()
    try:
        all_emps_raw = get_all_global_users_with_dept(conn)
        all_depts_raw = get_all_departments(conn)
    finally:
        conn.close()
    # Map employees to the format expected by the template
    all_emps = []
    for emp in all_emps_raw:
        all_emps.append({
            'key': str(emp['global_user_id']),  # use global_user_id as key
            'display_name': emp['name'],
            'company_id': emp['global_user_id'],  # same as key
            'department_name': emp.get('department_name') or ''
        })
    # Map departments
    all_depts = [{'id': dept['id'], 'name': dept['name']} for dept in all_depts_raw]
    # Determine selected employee display
    sel_emp_display = '— All Employees —'
    if emp_key:
        # Find the employee with matching key
        for emp in all_emps:
            if emp['key'] == emp_key:
                sel_emp_display = emp['display_name']
                if emp['company_id']:
                    sel_emp_display += f" ({emp['company_id']})"
                break
    # For now, we don't have a source for kaaj records, so show empty list
    records = []
    return render(templates, request, "kaaj.html", {
        "records": records,
        "sel_bs_year": bs_year,
        "sel_bs_month": bs_month,
        "from_ad": from_ad,
        "to_ad": to_ad,
        "all_emps": all_emps,
        "all_depts": all_depts,
        "f_dept": department_id,
        "f_paid": paid_filter,
        "sel_emp_key": emp_key,
        "sel_emp_display": sel_emp_display,
        "COMPANY_NAME": COMPANY_NAME,
    })


@app.post("/kaaj/add")
async def kaaj_add(request: Request):
    form = await request.form()
    global_user_id = form.get('global_user_id')
    ad_date = form.get('ad_date')
    bs_date = form.get('bs_date')
    is_paid = form.get('is_paid')
    reason = form.get('reason')
    approved_by = form.get('approved_by')
    # Basic validation
    if not global_user_id or not ad_date:
        return redirect_with_flash(request, '/kaaj', 'error', 'Employee and AD date are required')
    # Convert is_paid to boolean
    is_paid_bool = is_paid == '1'
    conn = get_connection()
    try:
        # Insert into kaaj_records table (adjust table/column names as per your schema)
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO kaaj_records (global_user_id, ad_date, bs_date, is_paid, reason, approved_by, created_by, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
        """, (global_user_id, ad_date, bs_date, is_paid_bool, reason, approved_by, _current_user_id(request)))
        conn.commit()
        return redirect_with_flash(request, '/kaaj', 'success', 'Kaaj record added successfully')
    except Exception as exc:
        conn.rollback()
        return redirect_with_flash(request, '/kaaj', 'error', f'Failed to add kaaj record: {exc}')
    finally:
        conn.close()


@app.get("/leaves/opening-balance")
def leaves_opening_balance(request: Request):
    from nepali_utils import today_bs
    from db import get_all_departments, get_all_leave_types, get_all_global_users_with_dept, get_leave_balances
    bs_year = request.query_params.get('bs_year')
    department_id = request.query_params.get('department_id')
    if not bs_year:
        bs_year = today_bs()[0]  # Get current B.S. year
    else:
        try:
            bs_year = int(bs_year)
        except ValueError:
            bs_year = today_bs()[0]
    if department_id:
        try:
            department_id = int(department_id)
        except ValueError:
            department_id = None
    conn = get_connection()
    try:
        all_depts = get_all_departments(conn)
        leave_types = get_all_leave_types(conn)
        all_emps = get_all_global_users_with_dept(conn)
        if department_id:
            all_emps = [emp for emp in all_emps if emp.get('department_id') == department_id]
        balances_list = get_leave_balances(conn, bs_year)
        # Create a dict: employee_balances[global_user_id][leave_type_code] = balance_object
        employee_balances = {}
        for bal in balances_list:
            gui = bal.get('global_user_id')
            ltc = bal.get('leave_type_code')  # leave type code from the leave_types table
            if gui is not None and ltc is not None:
                if gui not in employee_balances:
                    employee_balances[gui] = {}
                employee_balances[gui][ltc] = bal
        employees_data = []
        for emp in all_emps:
            gui = emp.get('id')  # Assuming 'id' is global_user_id from get_all_global_users_with_dept
            # Get the balances dict for this employee, or an empty dict if none
            emp['balances'] = employee_balances.get(gui, {})
            employees_data.append(emp)
        return render(templates, request, "leave_opening_balance.html", {
            "sel_bs_year": bs_year,
            "all_depts": all_depts,
            "f_dept": department_id,
            "leave_types": leave_types,
            "employees_data": employees_data,
            "COMPANY_NAME": COMPANY_NAME,
        })
    finally:
        conn.close()


@app.post("/leaves/opening-balance/save")
async def save_leaves_opening_balance(request: Request):
    from db import get_connection
    from web.flash import redirect_with_flash
    form = await request.form()
    bs_year = form.get('bs_year')
    if not bs_year:
        return redirect_with_flash(request, '/leaves/opening-balance', 'error', 'Year is required')
    try:
        bs_year = int(bs_year)
    except ValueError:
        return redirect_with_flash(request, '/leaves/opening-balance', 'error', 'Invalid year')
    updates = []
    for key, value in form.items():
        if key.startswith('ob_'):
            parts = key.split('_')
            if len(parts) < 3:
                continue
            try:
                gui = int(parts[1])
                ltc = '_'.join(parts[2:])  # Handle leave type codes with underscores
            except ValueError:
                continue
            try:
                ob = float(value) if value else 0
            except ValueError:
                ob = 0
            updates.append((gui, ltc, ob))
    conn = get_connection()
    try:
        # Build a map from leave_type_code to leave_type_id
        leave_type_map = {}
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT code, id FROM leave_types")
                for row in cur.fetchall():
                    leave_type_map[row['code']] = row['id']
        except Exception as e:
            # If we can't build the map, we cannot proceed
            return redirect_with_flash(request, '/leaves/opening-balance', 'error', f'Could not load leave types: {e}')
        # Convert updates to use leave_type_id
        updates_with_id = []
        for gui, ltc, ob in updates:
            lt_id = leave_type_map.get(ltc)
            if lt_id is None:
                # Skip or log? We'll skip for now.
                continue
            updates_with_id.append((gui, lt_id, ob))
        # Try to import the update function; if not available, use raw SQL as fallback
        try:
            from db import update_leave_opening_balance
            for gui, lt_id, ob in updates_with_id:
                update_leave_opening_balance(conn, gui, lt_id, bs_year, ob)
            conn.commit()
            return redirect_with_flash(request, f'/leaves/opening-balance?bs_year={bs_year}&{"&department_id="+str(request.query_params.get("department_id")) if request.query_params.get("department_id") else ""}', 'success', 'Opening balances updated successfully')
        except ImportError:
            # Fallback: update leave_balances table directly (adjust table/column names as per your schema)
            cursor = conn.cursor()
            for gui, lt_id, ob in updates_with_id:
                # Try to update existing record; if not found, insert new
                cursor.execute("""
                    INSERT INTO leave_balances (global_user_id, leave_type_id, bs_year, opening_balance)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (global_user_id, leave_type_id, bs_year)
                    DO UPDATE SET opening_balance = EXCLUDED.opening_balance
                """, (gui, lt_id, bs_year, ob))
            conn.commit()
            return redirect_with_flash(request, f'/leaves/opening-balance?bs_year={bs_year}&{"&department_id="+str(request.query_params.get("department_id")) if request.query_params.get("department_id") else ""}', 'success', 'Opening balances updated successfully')
    finally:
        conn.close()



# ═══════════════════════════════════════════════════════════════════════════
#  Payroll, Overtime & Tax  (Phase 13)
# ═══════════════════════════════════════════════════════════════════════════

def _payroll_ctx(request, extra=None):
    ctx = {"company": _get_company_settings(), "session": request.session}
    if extra:
        ctx.update(extra)
    return ctx


@app.get("/payroll")
def payroll_home(request: Request):
    from db import list_payroll_runs
    from nepali_utils import NEPALI_MONTHS
    conn = get_connection()
    try:
        runs = list_payroll_runs(conn)
    finally:
        conn.close()
    for r in runs:
        r["month_name"] = NEPALI_MONTHS[r["bs_month"]] if 1 <= r["bs_month"] <= 12 else r["bs_month"]
    today_bs = _today_bs()
    cur_year = int(today_bs[:4]) if today_bs else 2082
    cur_month = int(today_bs[5:7]) if today_bs else 1
    return templates.TemplateResponse(request, "payroll_home.html", _payroll_ctx(request, {
        "runs": runs,
        "nepali_months": NEPALI_MONTHS,
        "cur_year": cur_year,
        "cur_month": cur_month,
        "total_net": sum(float(r["total_net"]) for r in runs),
    }))


@app.get("/payroll/salary-structures")
def payroll_salary_structures(request: Request, q: str | None = None):
    from db import get_all_salary_structures
    conn = get_connection()
    try:
        rows = get_all_salary_structures(conn)
    finally:
        conn.close()
    if q:
        ql = q.lower()
        rows = [r for r in rows if ql in (r.get("name") or "").lower()
                or ql in str(r.get("emp_code") or "").lower()]
    configured = sum(1 for r in rows if r.get("basic_salary") is not None)
    return templates.TemplateResponse(request, "payroll_salary_structures.html", _payroll_ctx(request, {
        "rows": rows, "q": q or "", "configured": configured, "total": len(rows),
    }))


@app.post("/payroll/salary-structures")
def payroll_save_structure(request: Request,
                           global_user_id: int = Form(...),
                           basic_salary: float = Form(0),
                           allowances: float = Form(0),
                           daily_hours: float = Form(8),
                           ot_multiplier: float = Form(1.5),
                           marital: str = Form("single"),
                           other_deductions: float = Form(0),
                           q: str = Form("")):
    from db import upsert_salary_structure
    marital = "married" if marital == "married" else "single"
    conn = get_connection()
    try:
        upsert_salary_structure(conn, global_user_id, basic_salary, allowances,
                                daily_hours, ot_multiplier, marital,
                                other_deductions, _today_bs())
    finally:
        conn.close()
    dest = "/payroll/salary-structures" + (f"?q={q}" if q else "")
    return redirect_with_flash(dest, "success", "Salary structure saved.")


@app.get("/payroll/runs/new")
def payroll_new_run(request: Request):
    from nepali_utils import NEPALI_MONTHS, bs_month_info
    today_bs = _today_bs()
    cur_year = int(today_bs[:4]) if today_bs else 2082
    cur_month = int(today_bs[5:7]) if today_bs else 1
    mi = bs_month_info(cur_year, cur_month)
    return templates.TemplateResponse(request, "payroll_run_new.html", _payroll_ctx(request, {
        "nepali_months": NEPALI_MONTHS,
        "cur_year": cur_year, "cur_month": cur_month,
        "default_days": mi["days"] if mi else 30,
    }))


@app.post("/payroll/runs/generate")
def payroll_generate(request: Request,
                     bs_year: int = Form(...),
                     bs_month: int = Form(...),
                     working_days: int = Form(0)):
    import payroll as pay
    from nepali_utils import bs_month_info
    from db import (get_all_salary_structures, create_payroll_run,
                    clear_payroll_items, insert_payroll_item,
                    get_ytd_tax_totals, get_month_ot_split, get_month_present_days,
                    get_holiday_ot_multiplier)

    mi = bs_month_info(bs_year, bs_month)
    if not mi:
        return redirect_with_flash("/payroll/runs/new", "error", "Invalid Nepali month.")
    start_ad, end_ad = mi["first_ad"], mi["last_ad"]
    wd = working_days or mi["days"]
    period_index = pay.fiscal_period_index(bs_month)

    conn = get_connection()
    try:
        run_id = create_payroll_run(conn, bs_year, bs_month, period_index, wd,
                                    request.session.get("username", ""))
        clear_payroll_items(conn, run_id)
        structures = [r for r in get_all_salary_structures(conn)
                      if r.get("basic_salary") is not None]
        count = 0
        for s in structures:
            gu = s["global_user_id"]
            ot = get_month_ot_split(conn, gu, start_ad, end_ad)
            present = get_month_present_days(conn, gu, start_ad, end_ad)
            ytd = get_ytd_tax_totals(conn, gu, bs_year, period_index)
            # Holiday-OT premium multiplier if the employee is eligible, else normal.
            hol_mult = get_holiday_ot_multiplier(conn, gu)
            slip = pay.compute_payslip(
                basic_salary=s["basic_salary"],
                allowances=s["allowances"] or 0,
                working_days=wd,
                present_days=present,
                daily_hours=s["daily_hours"] or 8,
                ot_hours=ot["regular_ot_hours"],
                ot_multiplier=s["ot_multiplier"] or 1.5,
                holiday_ot_hours=ot["holiday_ot_hours"],
                holiday_ot_multiplier=hol_mult,
                other_deductions=s["other_deductions"] or 0,
                marital=s["marital"] or "single",
                period_index=period_index,
                taxable_ytd_before=ytd["taxable_before"],
                tax_paid_before=ytd["tax_paid_before"],
            )
            insert_payroll_item(conn, run_id, gu, {
                "present_days": slip["present_days"],
                "ot_hours": round(slip["ot_hours"] + slip["holiday_ot_hours"], 2),
                "ot_manual": False,
                "earned_basic": slip["earned_basic"],
                "earned_allowance": slip["earned_allowance"],
                "ot_pay": slip["ot_pay"],
                "other_earnings": slip["other_earnings"],
                "gross": slip["gross"],
                "taxable_this": slip["taxable_this_month"],
                "taxable_ytd": slip["taxable_ytd"],
                "tax": slip["tax"],
                "other_deductions": slip["other_deductions"],
                "net_pay": slip["net_pay"],
                "detail": {"hourly_rate": float(slip["hourly_rate"]),
                           "ot_multiplier": slip["ot_multiplier"],
                           "regular_ot_hours": slip["ot_hours"],
                           "regular_ot_pay": float(slip["regular_ot_pay"]),
                           "holiday_ot_hours": slip["holiday_ot_hours"],
                           "holiday_ot_multiplier": slip["holiday_ot_multiplier"],
                           "holiday_ot_pay": float(slip["holiday_ot_pay"]),
                           "holiday_ot_eligible": hol_mult is not None,
                           "prorate": slip["prorate"]},
            })
            count += 1
    finally:
        conn.close()
    return redirect_with_flash(f"/payroll/runs/{run_id}", "success",
                               f"Payroll generated for {count} employees.")


@app.get("/payroll/runs/{run_id}")
def payroll_view_run(request: Request, run_id: int):
    from db import get_payroll_run, get_payroll_items
    from nepali_utils import NEPALI_MONTHS
    conn = get_connection()
    try:
        run = get_payroll_run(conn, run_id)
        if not run:
            return redirect_with_flash("/payroll", "error", "Payroll run not found.")
        items = get_payroll_items(conn, run_id)
    finally:
        conn.close()
    run["month_name"] = NEPALI_MONTHS[run["bs_month"]] if 1 <= run["bs_month"] <= 12 else run["bs_month"]
    totals = {
        "gross": sum(float(i["gross"]) for i in items),
        "ot_pay": sum(float(i["ot_pay"]) for i in items),
        "tax": sum(float(i["tax"]) for i in items),
        "net": sum(float(i["net_pay"]) for i in items),
    }
    return templates.TemplateResponse(request, "payroll_run.html", _payroll_ctx(request, {
        "run": run, "items": items, "totals": totals,
    }))


@app.post("/payroll/runs/{run_id}/item/{gu_id}")
def payroll_edit_item(request: Request, run_id: int, gu_id: int,
                      ot_hours: float = Form(...),
                      holiday_ot_hours: float = Form(0),
                      other_earnings: float = Form(0),
                      other_deductions: float = Form(0)):
    """Manual OT-hours / adjustments override — recomputes just this payslip."""
    import payroll as pay
    from db import (get_payroll_run, get_salary_structure, get_ytd_tax_totals,
                    insert_payroll_item, get_month_present_days,
                    get_holiday_ot_multiplier)
    from nepali_utils import bs_month_info
    conn = get_connection()
    try:
        run = get_payroll_run(conn, run_id)
        s = get_salary_structure(conn, gu_id)
        if not run or not s:
            return redirect_with_flash(f"/payroll/runs/{run_id}", "error", "Not found.")
        mi = bs_month_info(run["bs_year"], run["bs_month"])
        present = get_month_present_days(conn, gu_id, mi["first_ad"], mi["last_ad"])
        ytd = get_ytd_tax_totals(conn, gu_id, run["bs_year"], run["period_index"])
        hol_mult = get_holiday_ot_multiplier(conn, gu_id)
        slip = pay.compute_payslip(
            basic_salary=s["basic_salary"], allowances=s["allowances"] or 0,
            working_days=run["working_days"], present_days=present,
            daily_hours=s["daily_hours"] or 8, ot_hours=ot_hours,
            ot_multiplier=s["ot_multiplier"] or 1.5,
            holiday_ot_hours=holiday_ot_hours, holiday_ot_multiplier=hol_mult,
            other_earnings=other_earnings, other_deductions=other_deductions,
            marital=s["marital"] or "single", period_index=run["period_index"],
            taxable_ytd_before=ytd["taxable_before"], tax_paid_before=ytd["tax_paid_before"],
        )
        insert_payroll_item(conn, run_id, gu_id, {
            "present_days": slip["present_days"],
            "ot_hours": round(slip["ot_hours"] + slip["holiday_ot_hours"], 2),
            "ot_manual": True, "earned_basic": slip["earned_basic"],
            "earned_allowance": slip["earned_allowance"], "ot_pay": slip["ot_pay"],
            "other_earnings": slip["other_earnings"], "gross": slip["gross"],
            "taxable_this": slip["taxable_this_month"], "taxable_ytd": slip["taxable_ytd"],
            "tax": slip["tax"], "other_deductions": slip["other_deductions"],
            "net_pay": slip["net_pay"],
            "detail": {"hourly_rate": float(slip["hourly_rate"]),
                       "ot_multiplier": slip["ot_multiplier"],
                       "regular_ot_hours": slip["ot_hours"],
                       "regular_ot_pay": float(slip["regular_ot_pay"]),
                       "holiday_ot_hours": slip["holiday_ot_hours"],
                       "holiday_ot_multiplier": slip["holiday_ot_multiplier"],
                       "holiday_ot_pay": float(slip["holiday_ot_pay"]),
                       "holiday_ot_eligible": hol_mult is not None,
                       "prorate": slip["prorate"]},
        })
    finally:
        conn.close()
    return redirect_with_flash(f"/payroll/runs/{run_id}/payslip/{gu_id}", "success",
                               "Payslip recalculated.")


@app.get("/payroll/runs/{run_id}/payslip/{gu_id}")
def payroll_payslip(request: Request, run_id: int, gu_id: int):
    import payroll as pay
    from db import get_payroll_run, get_payroll_item, get_salary_structure
    from nepali_utils import NEPALI_MONTHS
    conn = get_connection()
    try:
        run = get_payroll_run(conn, run_id)
        item = get_payroll_item(conn, run_id, gu_id)
        s = get_salary_structure(conn, gu_id)
        if not run or not item:
            return redirect_with_flash(f"/payroll/runs/{run_id}", "error", "Payslip not found.")
    finally:
        conn.close()
    run["month_name"] = NEPALI_MONTHS[run["bs_month"]] if 1 <= run["bs_month"] <= 12 else run["bs_month"]
    marital = (s or {}).get("marital", "single")
    tax_bands = pay.slab_breakdown(item["taxable_ytd"], marital)
    return templates.TemplateResponse(request, "payroll_payslip.html", _payroll_ctx(request, {
        "run": run, "item": item, "structure": s, "marital": marital,
        "tax_bands": tax_bands,
    }))


@app.get("/payroll/tax-preview")
def payroll_tax_preview(request: Request, income: float = 1200000, marital: str = "single"):
    """Standalone annual-tax calculator/preview (JSON)."""
    import payroll as pay
    bands = pay.slab_breakdown(income, marital)
    return JSONResponse({
        "annual_income": income,
        "marital": marital,
        "annual_tax": float(pay.annual_tax(income, marital)),
        "bands": [{k: (float(v) if hasattr(v, "quantize") else v) for k, v in b.items()} for b in bands],
    })


@app.get("/payroll/holiday-ot-rules")
def payroll_holiday_ot_rules(request: Request):
    from db import (get_holiday_ot_rules, get_all_departments, get_all_sections,
                    get_all_global_users_with_dept)
    conn = get_connection()
    try:
        rules = get_holiday_ot_rules(conn)
        depts = get_all_departments(conn)
        sections = get_all_sections(conn)
        users = get_all_global_users_with_dept(conn)
    finally:
        conn.close()
    return templates.TemplateResponse(request, "payroll_holiday_ot_rules.html", _payroll_ctx(request, {
        "rules": rules, "departments": depts, "sections": sections, "users": users,
    }))


@app.post("/payroll/holiday-ot-rules")
def payroll_add_holiday_ot_rule(request: Request,
                                scope: str = Form("employee"),
                                global_user_id: str = Form(""),
                                department_id: str = Form(""),
                                section_id: str = Form(""),
                                multiplier: float = Form(1.5),
                                note: str = Form("")):
    from db import add_holiday_ot_rule
    gu = _int_param(global_user_id) if scope == "employee" else None
    dept = _int_param(department_id) if scope == "department" else None
    sec = _int_param(section_id) if scope == "section" else None
    if not (gu or dept or sec):
        return redirect_with_flash("/payroll/holiday-ot-rules", "error",
                                   "Select an employee, department, or section for the rule.")
    conn = get_connection()
    try:
        add_holiday_ot_rule(conn, gu, dept, sec, multiplier, note)
    finally:
        conn.close()
    return redirect_with_flash("/payroll/holiday-ot-rules", "success",
                               "Holiday-OT rule added.")


@app.post("/payroll/holiday-ot-rules/{rule_id}/delete")
def payroll_delete_holiday_ot_rule(request: Request, rule_id: int):
    from db import delete_holiday_ot_rule
    conn = get_connection()
    try:
        delete_holiday_ot_rule(conn, rule_id)
    finally:
        conn.close()
    return redirect_with_flash("/payroll/holiday-ot-rules", "success", "Rule removed.")
